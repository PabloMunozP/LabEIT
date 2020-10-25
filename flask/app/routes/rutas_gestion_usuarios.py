from flask import Flask,Blueprint,render_template,request,redirect,url_for,flash,session,jsonify,send_file
from config import db,BASE_DIR
import os,time,bcrypt,random,re,ast
from jinja2 import Environment
from uuid import uuid4  # Token
from werkzeug.utils import secure_filename
import glob
import platform
from itertools import cycle
from datetime import datetime, timedelta
from .email_sender import enviar_correo_notificacion
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
PATH = BASE_DIR  # obtiene la ruta del directorio actual
# reemplaza [\\] por [/] en windows
PROFILE_PICS_PATH = PATH.replace(os.sep, '/')+'/app/static/imgs/profile_pics/'


mod = Blueprint("rutas_gestion_usuarios", __name__)


def redirect_url(default='index'):  # Redireccionamiento desde donde vino la request
    return request.args.get('next') or \
        request.referrer or \
        url_for(default)


# Función para obtener el archivo de la foto de perfil
def obtener_foto_perfil(rut_usuario):
    if glob.glob(PROFILE_PICS_PATH + rut_usuario + '.*'):  # Si la foto existe
        # Nombre del archivo + extension
        filename = glob.glob(PROFILE_PICS_PATH + rut_usuario + '.*')
        head, tail = os.path.split(filename[0])
        return tail  # Retorna nombre del archivo + extension de la foto de perfil
    else:  # Si la foto no existe
        return 'default_pic.png'  # Retorna la foto default


@mod.route("/gestion_usuarios", methods=['GET'])
def ver_usuarios():
    if "usuario" not in session.keys():
        return redirect("/")
    if session["usuario"]["id_credencial"] != 3:
        return redirect("/")
    
    # Se obtienen los parámetros en caso de haber 
    # sido redireccionado desde la carga masiva de usuarios desde la base de datos con el ID respectivo
    if request.args:
        id_resultado_carga_masiva = request.args.get("id_resultado_carga_masiva")
        sql_query = """
            SELECT * FROM Resultados_carga_masiva_usuarios
             WHERE id = %s
        """
        cursor = db.query(sql_query,(id_resultado_carga_masiva,))

        resultados_carga_masiva = cursor.fetchone()

        if resultados_carga_masiva is not None:
            # Se transforman los strings a diccionario y lista para errores y resultados respectivamente
            resultados_carga_masiva["dict_errores"] = ast.literal_eval(resultados_carga_masiva["dict_errores"])
            resultados_carga_masiva["registros_exitosos"] = ast.literal_eval(resultados_carga_masiva["registros_exitosos"])
    else:
        resultados_carga_masiva = None

    query = """
            SELECT Usuario.nombres AS nombres, Usuario.apellidos AS apellidos, Usuario.rut AS rut, Credencial.nombre AS credencial, Usuario.email as correo, Usuario.region as region, 
            Usuario.comuna as comuna, Usuario.direccion as direccion, Usuario.id_credencial as id_credencial, Usuario.activo 
                FROM Usuario,Credencial
                    WHERE Usuario.id_credencial= Credencial.id
            """
    #cursor.execute(query)
    cursor = db.query(query,None)

    usuarios = cursor.fetchall()
    return render_template("/vistas_gestion_usuarios/ver_usuarios.html", 
                            usuarios=usuarios,
                            resultados_carga_masiva=resultados_carga_masiva)


def digito_verificador(rut):
    reversed_digits = map(int, reversed(str(rut)))
    factors = cycle(range(2, 8))
    s = sum(d * f for d, f in zip(reversed_digits, factors))
    ver = (-s) % 11
    if ver < 10:
        return str(ver)
    if ver == 10:
        return 'K'


@mod.route("/gestion_usuarios/anadir_usuario", methods=["POST"])
def anadir_usuario():
    if "usuario" not in session.keys():
        return redirect("/")
    if session["usuario"]["id_credencial"] != 3:
        return redirect("/")

    if request.method == 'POST':
        # Creacion del usuario
        # obtener datos del usuario en un diccionario
        datos_usuario = request.form.to_dict()
        # Comprobar que el usuario no exista previamente

        query = ''' SELECT Usuario.rut as rut FROM Usuario WHERE rut=%s '''
        #cursor.execute(query, (datos_usuario['rut'],))
        cursor = db.query(query, (datos_usuario['rut'],))

        duplicados_rut = cursor.fetchone()

        query = ''' SELECT Usuario.email as correo FROM Usuario WHERE email=%s '''
        #cursor.execute(query, (datos_usuario['correo'],))
        cursor = db.query(query, (datos_usuario['correo'],))

        duplicados_correo = cursor.fetchone()

        if duplicados_rut is not None:  # Ya existe un usuario con ese rut
            flash('error-anadir-rut')
            return redirect("/gestion_usuarios")
        if duplicados_correo is not None:  # Ya existe alguien registrado con ese rut
            flash('error-anadir-correo')
            return redirect("/gestion_usuarios")
        # No exista nadie con rut ni correo repetido.
        if duplicados_rut is None and duplicados_correo is None:

            # Comprobar que el rut y verificador sean validos.
            # print(datos_usuario['rut'][-1])
            # print(digito_verificador(datos_usuario['rut'][0:-1]))
            if datos_usuario['rut'][-1] == digito_verificador(datos_usuario['rut'][0:-1]):
                # El rut es correcto

                query = ''' INSERT INTO Usuario(rut, id_credencial, email, nombres, apellidos)
                    VALUES (%s, %s, %s, %s, %s)'''
                #cursor.execute(query, (datos_usuario['rut'], datos_usuario['credencial'],
                #                       datos_usuario['correo'], datos_usuario['nombres'], datos_usuario['apellidos']))
                db.query(query, (datos_usuario['rut'], datos_usuario['credencial'],datos_usuario['correo'], datos_usuario['nombres'], datos_usuario['apellidos']))

                # Una vez creado, se le notifica al usuario para que cambie su contraseña y complete sus datos

                # Se abre el template HTML correspondiente al restablecimiento de contraseña
                direccion_template = os.path.normpath(os.path.join(
                    os.getcwd(), "app/templates/vistas_exteriores/establecer_password_mail.html"))
                html_restablecimiento = open(
                    direccion_template, encoding="utf-8").read()

                # Se crea el token único para restablecimiento de contraseña
                token = str(uuid4())

                # Se eliminan los registros de token asociados al rut del usuario en caso de existir
                sql_query = """ DELETE FROM Token_recuperacion_password
                      WHERE rut_usuario = %s   """
                #cursor.execute(sql_query, (datos_usuario["rut"],))
                db.query(sql_query, (datos_usuario["rut"],))

                # Se reemplazan los datos del usuario en el template a enviar vía correo
                html_restablecimiento = html_restablecimiento.replace(
                    "%nombre_usuario%", datos_usuario["nombres"])
                html_restablecimiento = html_restablecimiento.replace(
                    "%token_restablecimiento%", token)

                #Se envia el correo 
                enviar_correo_notificacion(html_restablecimiento,"Establecer Contraseña - LabEIT UDP",datos_usuario["correo"])
            
                # Se registra el token en la base de datos según el RUT del usuario
                sql_query = """
                    INSERT INTO Token_recuperacion_password
                        (token,rut_usuario)
                            VALUES (%s,%s)
                """
                #cursor.execute(
                #    sql_query, (str(token), datos_usuario["rut"]))
                db.query(sql_query, (str(token), datos_usuario["rut"]))

                flash('agregar-correcto')
                return redirect("/gestion_usuarios")
            else:
                flash('error-digitoVerficador')
                return redirect("/gestion_usuarios")


@mod.route("/gestion_usuarios/editar_usuario", methods=["POST"])
def editar():
    if "usuario" not in session.keys():
        return redirect("/")
    if session["usuario"]["id_credencial"] != 3:
        return redirect("/")

    if request.method == 'POST':
        datos_usuario = request.form.to_dict()
        # Comprobar que no se repita el correo.
        query = ''' SELECT rut FROM Usuario WHERE email= %s'''
        #cursor.execute(query, (datos_usuario['correo'],))
        cursor = db.query(query, (datos_usuario['correo'],))

        error_correo = cursor.fetchone()
        if error_correo['rut'] != datos_usuario['rut']:
            # hay otro usuario que ya utiliza el correo
            flash('error-editar-correo')
            return redirect("/gestion_usuarios")
        # query para actualizar datos del usuario
        query = ''' UPDATE Usuario SET id_credencial = %s, nombres =%s, apellidos= %s, region = %s, comuna = %s, direccion = %s
                    WHERE rut= %s'''
        #cursor.execute(query, (datos_usuario['id_credencial'], datos_usuario['correo'], datos_usuario['nombres'],
        #                       datos_usuario['apellidos'], datos_usuario['region'], datos_usuario['comuna'], datos_usuario['direccion'], datos_usuario['rut']))
        db.query(query, (datos_usuario['id_credencial'], datos_usuario['nombres'],datos_usuario['apellidos'], datos_usuario['region'], datos_usuario['comuna'], datos_usuario['direccion'], datos_usuario['rut']))

        flash('editado-correcto')
        # se redirige de vuelta a la pagina principal de gestion usuarios
        return redirect("/gestion_usuarios")


@mod.route("/gestion_usuarios/eliminar_usuario", methods=["POST"])
def eliminar():
    if "usuario" not in session.keys():
        return redirect("/")
    if session["usuario"]["id_credencial"] != 3:
        return redirect("/")

    if request.method == 'POST':
        rut = request.form['rut']
        # Se verifica que el usuario no tenga detalles de solicitudes de equipo en los estados correspondientes
        # por retirar (1) / en posesión (2) / con atraso (3)
        query = """
            SELECT COUNT(*) AS cantidad_detalles_solicitud_activos
                FROM Solicitud,Detalle_solicitud
                    WHERE Detalle_solicitud.estado IN (1,2,3)
                    AND Solicitud.id = Detalle_solicitud.id_solicitud
                    AND Solicitud.rut_alumno = %s
        """
        #cursor.execute(query, (rut,))
        cursor = db.query(query, (rut,))
        detalles_solicitudes = cursor.fetchone()

        if detalles_solicitudes is not None:
            cantidad_detalles_solicitudes_activos = detalles_solicitudes["cantidad_detalles_solicitud_activos"]
        else:
            flash("error-eliminar")
            return redirect("/gestion_usuarios")
        
        if cantidad_detalles_solicitudes_activos != 0:
            # En caso de tener detalles de solicitud activos según los estados respectivos
            # Se notifica acerca del error
            flash("error-eliminar")
            return redirect("/gestion_usuarios")
        
        # Si tiene 0 detalles de solicitudes en actividad, se comprueban
        # las solicitudes de circuitos
        query = """
            SELECT COUNT(*) AS cantidad_detalles_circuitos_activos
                FROM Solicitud_circuito,Detalle_solicitud_circuito
                    WHERE Detalle_solicitud_circuito.estado IN (1,2,3)
                    AND Solicitud_circuito.id = Detalle_solicitud_circuito.id_solicitud_circuito
                    AND Solicitud_circuito.rut_alumno = %s
        """
        cursor = db.query(query, (rut,))
        detalles_solicitudes = cursor.fetchone()

        if detalles_solicitudes is not None:
            cantidad_detalles_circuitos_activos = detalles_solicitudes["cantidad_detalles_circuitos_activos"]
        else:
            flash("error-eliminar")
            return redirect("/gestion_usuarios")
        
        if cantidad_detalles_circuitos_activos != 0:
            # En caso de tener detalles de solicitud (CIRCUITOS) activos según los estados respectivos
            # Se notifica acerca del error
            flash("error-eliminar")
            return redirect("/gestion_usuarios")
        
        # En caso de que no tenga ninguna solicitud en los estados mencionados
        # Tanto de circuitos como equipos, se eliminan los registros del usuario

        # Se obtienen los IDs de las solicitudes (equipos y circuitos) para eliminar los registros
        sql_query = """
            SELECT id FROM Solicitud
                WHERE rut_alumno = %s
        """
        cursor = db.query(sql_query,(rut,))
        lista_ids_solicitudes_equipos = cursor.fetchall()

        if len(lista_ids_solicitudes_equipos) != 0:
            # Se genera una lista con los ids obtenidos en caso de obtener registros
            lista_ids = []
            for registro_id in lista_ids_solicitudes_equipos:
                lista_ids.append(registro_id["id"])
        
            sql_query = """
                DELETE FROM Detalle_solicitud
                    WHERE id_solicitud IN {0}
            """
            string_lista_ids = str(lista_ids).replace("[","(").replace("]",")")
            sql_query = sql_query.format(string_lista_ids)
            db.query(sql_query,None)
        
        sql_query = """
            SELECT id FROM Solicitud_circuito
                WHERE rut_alumno = %s
        """
        cursor = db.query(sql_query,(rut,))
        lista_ids_solicitudes_circuitos = cursor.fetchall()
        
        if len(lista_ids_solicitudes_circuitos) != 0:
            # Se genera una lista con los ids obtenidos en caso de obtener registros
            lista_ids = []
            for registro_id in lista_ids_solicitudes_circuitos:
                lista_ids.append(registro_id["id"])
            
            sql_query = """
                DELETE FROM Detalle_solicitud_circuito
                    WHERE id_solicitud_circuito IN {0}
            """
            string_lista_ids = str(lista_ids).replace("[","(").replace("]",")")
            sql_query = sql_query.format(string_lista_ids)
            db.query(sql_query,None)

        # Se deben eliminar registros de sanciones, seccion_alumno, token_recuperacion_password, wishlist

        # Eliminación de sanciones asociadas al usuario
        sql_query = """
            DELETE FROM Sanciones
                WHERE rut_alumno = %s
        """
        db.query(sql_query,(rut,))

        # Eliminación de la relación de secciones con el usuario
        sql_query = """
            DELETE FROM Seccion_alumno
                WHERE rut_alumno = %s
        """
        db.query(sql_query,(rut,))

        # Eliminación de registros de sanciones del usuario
        sql_query = """
            DELETE FROM Token_recuperacion_password
                WHERE rut_usuario = %s
        """
        db.query(sql_query,(rut,))

        # Eliminación de registros de wishlist

        # Se obtienen los ids de wishlist asociados al rut del usuario a eliminar
        sql_query = """
            SELECT id FROM Wishlist
                WHERE rut_solicitante = %s
        """
        cursor = db.query(sql_query,(rut,))
        lista_ids_wishlist = cursor.fetchall()

        if len(lista_ids_wishlist) != 0:
            lista_ids = []
            for registro_id in lista_ids_wishlist:
                lista_ids.append(registro_id["id"])

            # Se eliminan los registros de Wishlist
            sql_query = """
                DELETE FROM Wishlist
                    WHERE id IN {0}
            """
            string_lista_ids = str(lista_ids).replace("[","(").replace("]",")")
            sql_query = sql_query.format(string_lista_ids)
            db.query(sql_query,None)

            # Se eliminan las urls asociadas
            sql_query = """
                DELETE FROM Url_wishlist
                    WHERE id_wishlist IN {0}
            """
            string_lista_ids = str(lista_ids).replace("[","(").replace("]",")")
            sql_query = sql_query.format(string_lista_ids)
            db.query(sql_query,None)

            # Se eliminan los motivos académicos asociados
            sql_query = """
                DELETE FROM Motivo_academico_wishlist
                    WHERE id_wishlist IN {0}
            """
            string_lista_ids = str(lista_ids).replace("[","(").replace("]",")")
            sql_query = sql_query.format(string_lista_ids)
            db.query(sql_query,None)
        
        # Finalmente, se elimina al usuario
        sql_query = """
            DELETE FROM Usuario
                WHERE rut = %s
        """
        db.query(sql_query,(rut,))

        flash('eliminar-correcto')
        return redirect("/gestion_usuarios")

@mod.route("/gestion_usuarios/inhabilitar", methods=["POST"])
def inhabilitar():

    if "usuario" not in session.keys():
        return redirect("/")
    if session["usuario"]["id_credencial"] != 3:
        return redirect("/")

    if request.method == 'POST':
        rut = request.form['rut']
        query = """
            SELECT COUNT(*) AS cantidad_detalles_en_proceso FROM Detalle_solicitud,Solicitud
                WHERE Detalle_solicitud.id_solicitud = Solicitud.id
                AND Solicitud.rut_alumno = %s
                AND Detalle_solicitud.estado IN (0,1,2,3)
        """
        #cursor.execute(query, (rut,))
        cursor = db.query(query, (rut,))

        solicitudes = cursor.fetchone()

        if solicitudes is None:
            # Error entre tablas de detalle y solicitud
            flash("error-enlace-solicitud")
            return redirect("/gestion_usuarios")

        else:
            if solicitudes["cantidad_detalles_en_proceso"]:
                flash("error-inhabilitar")
                return redirect("/gestion_usuarios")
            else:
                query = '''UPDATE Usuario SET Usuario.activo = 0 WHERE rut= %s'''
                #cursor.execute(query, (rut,))
                db.query(query, (rut,))

                flash('inhabilitar-correcto')
                return redirect("/gestion_usuarios")


@mod.route("/gestion_usuarios/habilitar", methods=["POST"])
def habilitar_usuario():
    if "usuario" not in session.keys():
        return redirect("/")
    if session["usuario"]["id_credencial"] != 3:
        return redirect("/")

    # Se habilita la cuenta del usuario en caso de existir
    rut = request.form['rut']
    sql_query = """
        UPDATE Usuario
            SET activo = 1
                WHERE rut = %s
    """
    #cursor.execute(sql_query, (rut,))
    db.query(sql_query, (rut,))

    flash("cuenta-activada")
    return redirect("/gestion_usuarios")


@mod.route("/gestion_usuarios/ver_usuario/<string:rut>", methods=["GET"])
def detalle_usuario(rut):
    if "usuario" not in session.keys():
        return redirect("/")
    if session["usuario"]["id_credencial"] != 3:
        return redirect("/")

    if request.method == 'GET':
        # print(rut)
        query = ''' SELECT Usuario.nombres as nombres, Usuario.apellidos as apellidos, Credencial.nombre as credencial,
            Usuario.email as correo, Usuario.region as region, Usuario.comuna as comuna, Usuario.rut as rut,
            Usuario.direccion as direccion, Usuario.celular as celular FROM Usuario,Credencial WHERE Credencial.id=Usuario.id_credencial AND Usuario.rut=%s '''
        #cursor.execute(query, (rut,))
        cursor = db.query(query, (rut,))

        usuario = cursor.fetchone()

        query = ''' 
            SELECT Solicitud.id as id, Solicitud.rut_profesor as profesor, Solicitud.rut_alumno as alumno, Solicitud.motivo as motivo, Solicitud.fecha_registro as registro,
                Detalle_solicitud.estado as estado, Detalle_solicitud.id as id_detalle, Equipo.nombre as equipo, Equipo.modelo as modelo, Equipo.marca as marca_equipo, Estado_detalle_solicitud.nombre AS nombre_estado
                    FROM Solicitud, Detalle_solicitud, Equipo, Estado_detalle_solicitud
                        WHERE Solicitud.id = Detalle_solicitud.id_solicitud 
                        AND Solicitud.rut_alumno= %s 
                        AND Equipo.id = Detalle_solicitud.id_equipo 
                        AND Detalle_solicitud.estado = Estado_detalle_solicitud.id
        '''
        #cursor.execute(query, (rut,))
        cursor = db.query(query, (rut,))

        solicitudes = cursor.fetchall()

        query = """
            SELECT Curso.*,Seccion.codigo AS codigo_seccion
                FROM Curso,Seccion,Seccion_alumno
                    WHERE Curso.id = Seccion.id_curso
                    AND Seccion.id = Seccion_alumno.id_seccion
                    AND Seccion_alumno.rut_alumno = %s
                    ORDER BY Curso.nombre
        """
        #cursor.execute(query, (rut,))
        cursor = db.query(query, (rut,))

        cursos = cursor.fetchall()

        query = '''SELECT * FROM Sanciones WHERE rut_alumno = %s'''
        #cursor.execute(query, (rut,))
        cursor = db.query(query, (rut,))

        sancion = cursor.fetchone()

        archivo_foto_perfil = obtener_foto_perfil(rut)

        return render_template("/vistas_gestion_usuarios/detalle_usuario.html", usuario=usuario, solicitudes=solicitudes,
                               cursos=cursos, sancion=sancion, dir_foto_perfil=url_for('static', filename='imgs/profile_pics/'+archivo_foto_perfil))


def get_size(fobj): # Permite obtener el tamaño de un archivo
        if fobj.content_length:
            return fobj.content_length
        try:
            pos = fobj.tell()
            fobj.seek(0, 2)  #seek to end
            size = fobj.tell()
            fobj.seek(pos)  # back to original position
            return size
        except (AttributeError, IOError):
            pass

        # in-memory file object that doesn't support seeking or tell
        return 0  #assume small enough


@mod.route("/gestion_usuarios/anadir_masivo",methods=["POST"])
def masivo():
    if "usuario" not in session.keys():
        return redirect("/")
    if session["usuario"]["id_credencial"] != 3:
        return redirect("/")
    
    try:
        archivo = request.files["file"]
        if not "." in archivo.filename:
            flash("error-agregar-masivo")
            return redirect("/gestion_usuarios")
    
        if get_size(archivo) > 1024*1024*5: #si el archivo es mayor a 5 Mb se rechaza
            flash('error-tamano')
            return redirect("/gestion_usuarios")
    
        # Se obtiene la extensión del archivo subido
        ext= os.path.splitext(archivo.filename)[1]

        if ext != ".xlsx":
            # El archivo subido no corresponde a un archivo Excel
            flash('error-agregar-masivo')
            return redirect("/gestion_usuarios")
    
        # El archivo subido corresponde a un archivo Excel
        # Se abre el archivo con load_workbook
        wb = load_workbook(archivo)
        ws = wb.active # Se abre la hoja activa del libro

        fila_inicial = 2 # Los inicios de registros en la planilla inician en la fila 2
        dict_errores = {} # Permite almacenar las coordenadas de los errores detectados
        # Se inicializan las llaves con los mensajes de error
        dict_errores["RUT no valido"] = []
        dict_errores["RUT en uso"] = []
        dict_errores["Email no valido"] = []
        dict_errores["Email en uso"] = []
        dict_errores["Correos fallidos"] = []

        registros_exitosos = [] # Permite almacenar los usuarios registrados correctamente

        # Se recorre cada uno de los registros en el archivo (no nulos)
        for index_fila in range(fila_inicial,ws.max_row+1):
            # Se crean los arreglos con la información de cada usuario
            registro_usuario = []
            celdas_vacias = False
            for celda in ws[index_fila]:
                # Se registra el valor en la información del registro de usuario
                registro_usuario.append(str(celda.value))
                if celda.value is None:
                    # En caso de que existan celdas vacías, se continúa con la siguiente fila
                    celdas_vacias = True
                    break
                # En caso de llegar a la última fila (fila 5 correspondiente a credencial)
                # Se termina de agregar información
                if celda.column == 5:
                    break
            
            if celdas_vacias:
                # Si existen celdas vacías, se omite la fila actual y se continúa con la siguiente
                continue

            # Se continúa con las validaciones una vez obtenidos los datos del usuario
            # Se modifica a False en caso de detectar algún error y no registrar finalmente.
            # Además permite obtener todos los errores en las distintas columnas de la fila correspondiente.
            registro_valido = True
            
            # Validaciones de RUT
            # Se remueven los puntos o guiones en caso de que existan y se ponen en mayúscula los caracteres (en caso de la 'K')
            registro_usuario[2] = registro_usuario[2].replace(".","").replace("-","").upper()
            
            # Se verifica el largo del RUT ingresado
            if (len(registro_usuario[2]) <= 7 or len(registro_usuario[2]) > 9):
                # El largo del string de RUT no corresponde a la cantidad de caracteres
                dict_errores["RUT no valido"].append(get_column_letter(3)+str(celda.row))
                registro_valido = False
            
            # Se intenta revisar y calcular el dígito verificador y en caso de error se agrega al log.
            if registro_valido:
                try:
                    rut_base = registro_usuario[2][:len(registro_usuario[2])-1] # Sin contemplar dígito verificador
                    if digito_verificador(rut_base) != registro_usuario[2][len(registro_usuario[2])-1]:
                        # El rut ingresado no es válido
                        dict_errores["RUT no valido"].append(get_column_letter(3)+str(celda.row))
                        registro_valido = False

                except:
                    # El rut ingresado no es válido
                    dict_errores["RUT no valido"].append(get_column_letter(3)+str(celda.row))
                    registro_valido = False
                
                if registro_valido:
                    # Si el registro sigue siendo válido, se revisa la existencia del RUT
                    # en la base de datos 
                    sql_query = """
                        SELECT COUNT(*) AS cantidad_registros
                            FROM Usuario
                                WHERE rut = %s
                    """
                    cursor = db.query(sql_query,(registro_usuario[2],))
                    cantidad_registros = cursor.fetchone()["cantidad_registros"]

                    if cantidad_registros != 0:
                        # El RUT ingresado ya se encuentra asociado a un usuario registrado
                        dict_errores["RUT en uso"].append(get_column_letter(3)+str(celda.row))
                        registro_valido = False

            # Luego de validar el RUT, se revisa la existencia y formato del correo electrónico

            # Se verifica el dominio del correo (mail.udp.cl o udp.cl)
            dominio_correo = re.search("@[\w.]+", registro_usuario[3]).group()
            
            if dominio_correo != "@mail.udp.cl" and dominio_correo != "@udp.cl":
                dict_errores["Email no valido"].append(get_column_letter(4)+str(celda.row))
                registro_valido = False
            
            if registro_valido:
                # Se verifica si el correo electrónico se encuentra registrado en la base de datos
                sql_query = """
                    SELECT COUNT(*) AS cantidad_registros FROM Usuario
                        WHERE email = %s
                """
                cursor = db.query(sql_query,(registro_usuario[3],))
                registro_correo = cursor.fetchone()["cantidad_registros"]

                if registro_correo != 0:
                    # El correo ya se encuentra asociado a un usuario registrado
                    dict_errores["Email en uso"].append(get_column_letter(4)+str(celda.row))
                    registro_valido = False
                
                else:
                    # Si el correo no se encuentra registrado, se realiza el registro
                    # en base al último campo del tipo de usuario
                    
                    # Si la credencial no corresponde a 1 ni a 2, se deja como 1 (alumno) por defecto
                    if registro_usuario[4] not in ["1","2"]:
                        registro_usuario[4] = "1" # Se setea a alumno por default
                    
                    # Se realiza el registro del usuario y envío de token
                    sql_query = """
                        INSERT INTO Usuario (rut,id_credencial,email,nombres,apellidos)
                            VALUES (%s,%s,%s,%s,%s)
                    """
                    db.query(sql_query,(registro_usuario[2],registro_usuario[4],registro_usuario[3],registro_usuario[0],registro_usuario[1]))

                    # Se eliminan los posibles tokens asociados al usuario
                    sql_query = """
                        DELETE FROM Token_recuperacion_password
                            WHERE rut_usuario = %s
                    """
                    db.query(sql_query,(registro_usuario[2],))

                    # Se crea el token único para restablecimiento de contraseña
                    token = str(uuid4())
                    fecha_actual = datetime.now().replace(microsecond=0)

                    # Se registra el token del nuevo usuario en la base de datos
                    sql_query = """
                        INSERT INTO Token_recuperacion_password (token,rut_usuario,fecha_registro)
                            VALUES (%s,%s,%s)
                    """
                    db.query(sql_query,(token,registro_usuario[2],fecha_actual))

                    # Se abre el template HTML correspondiente al restablecimiento de contraseña
                    direccion_template = os.path.normpath(os.path.join(
                    os.getcwd(), "app/templates/vistas_exteriores/establecer_password_mail.html"))
                    html_restablecimiento = open(direccion_template, encoding="utf-8").read()

                    # Se reemplazan los datos del usuario en el template a enviar vía correo
                    html_restablecimiento = html_restablecimiento.replace("%nombre_usuario%", registro_usuario[0])
                    html_restablecimiento = html_restablecimiento.replace("%token_restablecimiento%", token)

                    # Se intenta enviar el correo y se revisa el estado
                    estado_envio = enviar_correo_notificacion(html_restablecimiento,"Establecer Contraseña - LabEIT UDP",registro_usuario[3],masivo=True)

                    if estado_envio:
                        # El correo fue enviado correctamente al usuario
                        registros_exitosos.append(str(celda.row))
                    else:
                        # Se produjo un error interno al intentar enviar el correo
                        dict_errores["Correos fallidos"].append(str(celda.row))
        
        # Se eliminan los posibles registros de resultados anteriores
        sql_query = """
            DELETE FROM Resultados_carga_masiva_usuarios
        """
        db.query(sql_query,None)

        # Se genera el registro de los resultados de la carga masiva en la base de datos
        fecha_actual = datetime.now().replace(microsecond=0)
        sql_query = """
            INSERT INTO Resultados_carga_masiva_usuarios (dict_errores,registros_exitosos,fecha_registro)
                VALUES (%s,%s,%s)
        """
        cursor = db.query(sql_query,(str(dict_errores),str(registros_exitosos),fecha_actual))              

        # Se obtiene el ID de resultado generado en la base de datos
        id_resultado_carga_masiva = cursor.lastrowid

        return redirect(url_for("rutas_gestion_usuarios.ver_usuarios",id_resultado_carga_masiva=id_resultado_carga_masiva))
    
    except BaseException as e:
        print(e)
        flash('error-desconocido-carga-masiva')
        return redirect("/gestion_usuarios")

@mod.route("/gestion_usuarios/anadir_masivo/formato",methods=["GET"])
def formato_xlsx():
    if "usuario" not in session.keys():
        return redirect("/")
    if session["usuario"]["id_credencial"] != 3:
        return redirect("/")
    
    ruta = os.path.normpath(os.path.join( PATH +'/app/static/files/plantillas/', 'carga_masiva_usuarios.xlsx' ))
    return send_file(ruta, as_attachment=True)

@mod.route("/gestion_usuarios/sancion", methods=["POST"])
def sancion():
    data = request.form.to_dict()
    if data['cant_sancion'] == '0':
        flash("sin-cambio")
    else:
        query = ''' SELECT cantidad_dias FROM Sanciones WHERE rut_alumno= %s'''
        #cursor.execute(query, (data['rut'],))
        cursor = db.query(query, (data['rut'],))

        dias = cursor.fetchone()
        dias = dias['cantidad_dias']
        if data['op_sancion'] == '1':  # Aumentar dias de sancion
            dias = dias+int(data['cant_sancion'])
        elif data['op_sancion'] == '2':  # Disminuir dias de sancion
            dias = dias-int(data['cant_sancion'])
        if dias < 0:
            flash("Error")
            return redirect(redirect_url())
        elif dias == 0:
            query = ''' DELETE FROM  Sanciones WHERE rut_alumno= %s '''
            #cursor.execute(query, (data['rut'],))
            db.query(query, (data['rut'],))

            flash("cambio-realizado")
            return redirect(redirect_url())
        # Se actualiza el registro y se comprueba si se debe eliminar.
        query = ''' UPDATE Sanciones SET cantidad_dias = %s WHERE rut_alumno = %s'''
        #cursor.execute(query, (dias, data['rut']))
        cursor = db.query(query, (dias, data['rut']))
        # print("cambio-realizado")
        flash("cambio-realizado")
    return redirect(redirect_url())
