import smtplib
import json
from uuid import uuid4  # Token
from email import encoders
from openpyxl import Workbook
from jinja2 import Environment
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
import os
import time
import bcrypt
import random
import timeago
import shutil
from email.mime.multipart import MIMEMultipart
from .email_sender import enviar_correo_notificacion
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from config import db, BASE_DIR, ALLOWED_EXTENSIONS, MAX_CONTENT_LENGTH
from flask import Flask, Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, send_file

mod = Blueprint("rutas_gestion_solicitudes", __name__)

def redirect_url(default='index'):  # Redireccionamiento desde donde vino la request
    return request.args.get('next') or \
        request.referrer or \
        url_for(default)

# ================================== GESTIÓN DE SOLICITUDES DE PRÉSTAMOS
@mod.route("/gestion_solicitudes_prestamos", methods=["GET"])
def gestion_solicitudes_prestamos():
    if "usuario" not in session.keys():
        return redirect("/")
    # El usuario debe ser un administrador (Credencial = 3)
    if session["usuario"]["id_credencial"] != 3:
        return redirect("/")

    # Se obtiene el listado de detalles de solicitudes por revisar
    sql_query = """
        SELECT Detalle_solicitud.*,Equipo.nombre,Equipo.marca,Equipo.modelo,Usuario.rut AS rut_alumno,Usuario.nombres AS nombres_usuario,Usuario.apellidos AS apellidos_usuario,Solicitud.fecha_registro
            FROM Detalle_solicitud,Equipo,Solicitud,Usuario
                WHERE Solicitud.id = Detalle_solicitud.id_solicitud
                AND Detalle_solicitud.id_equipo = Equipo.id
                AND Solicitud.rut_alumno = Usuario.rut
                AND Detalle_solicitud.estado = 0
    """
    #cursor.execute(sql_query)
    cursor = db.query(sql_query,None)
    lista_solicitudes_por_revisar = cursor.fetchall()

    # Se obtiene el listado de detalles de solicitudes activas
    sql_query = """
        SELECT Detalle_solicitud.*,Equipo.nombre,Equipo.marca,Equipo.modelo,Usuario.rut AS rut_alumno,Usuario.nombres AS nombres_usuario,Usuario.apellidos AS apellidos_usuario,Solicitud.fecha_registro,Estado_detalle_solicitud.nombre AS nombre_estado
            FROM Detalle_solicitud,Equipo,Solicitud,Usuario,Estado_detalle_solicitud
                WHERE Solicitud.id = Detalle_solicitud.id_solicitud
                AND Estado_detalle_solicitud.id = Detalle_solicitud.estado
                AND Detalle_solicitud.id_equipo = Equipo.id
                AND Solicitud.rut_alumno = Usuario.rut
                AND Detalle_solicitud.estado != 0
                AND Detalle_solicitud.estado < 5
    """
    #cursor.execute(sql_query)
    cursor = db.query(sql_query,None)
    lista_solicitudes_activas = cursor.fetchall()

    # Se obtiene la lista de solicitudes pertenecientes al historial
    sql_query = """
        SELECT Detalle_solicitud.*,Equipo.nombre,Equipo.marca,Equipo.modelo,Usuario.rut AS rut_alumno,Usuario.nombres AS nombres_usuario,Usuario.apellidos AS apellidos_usuario,Solicitud.fecha_registro,Estado_detalle_solicitud.nombre AS nombre_estado
            FROM Detalle_solicitud,Equipo,Solicitud,Usuario,Estado_detalle_solicitud
                WHERE Solicitud.id = Detalle_solicitud.id_solicitud
                AND Estado_detalle_solicitud.id = Detalle_solicitud.estado
                AND Detalle_solicitud.id_equipo = Equipo.id
                AND Solicitud.rut_alumno = Usuario.rut
                AND Detalle_solicitud.estado != 0
                AND Detalle_solicitud.estado >= 5
    """
    #cursor.execute(sql_query)
    cursor = db.query(sql_query,None)
    lista_historial_solicitudes = cursor.fetchall()

    # Lista de equipos para formulario ágil
    sql_query = """
        SELECT Equipo.id,Equipo.nombre,Equipo.marca,Equipo.modelo,Equipo.codigo,Equipo_diferenciado.codigo_sufijo
            FROM Equipo,Equipo_diferenciado
                WHERE Equipo.codigo = Equipo_diferenciado.codigo_equipo
                AND Equipo_diferenciado.activo = 1
                AND Equipo_diferenciado.codigo_sufijo
                    NOT IN (SELECT Detalle_solicitud.codigo_sufijo_equipo
                                FROM Detalle_solicitud
                                    WHERE Detalle_solicitud.id_equipo = Equipo.id
                                        AND codigo_sufijo_equipo IS NOT NULL)
                ORDER BY Equipo.nombre,Equipo.marca,Equipo.modelo
    """
    #cursor.execute(sql_query)
    cursor = db.query(sql_query,None)
    lista_equipos_disponibles = cursor.fetchall()

    # Se obtiene de forma general la lista de solicitudes registradas (vista de canasta)
    sql_query = """
        SELECT Solicitud.*,CONCAT(Usuario.nombres,' ',Usuario.apellidos) AS nombre_solicitante,
            (SELECT COUNT(*) FROM Detalle_solicitud WHERE Detalle_solicitud.id_solicitud = Solicitud.id ) AS cantidad_detalles
         FROM
            Solicitud,Usuario
                WHERE Solicitud.rut_alumno = Usuario.rut
                ORDER BY Solicitud.fecha_registro DESC
    """
    #cursor.execute(sql_query)
    cursor = db.query(sql_query,None)
    lista_solicitud_canasta = cursor.fetchall()

    # Se obtienen los cursos registrados para asociar a la solicitud ágil
    sql_query = """
        SELECT id,codigo_udp,nombre
            FROM Curso
    """
    #cursor.execute(sql_query)
    cursor = db.query(sql_query,None)
    lista_asignaturas = cursor.fetchall()

    return render_template("/vistas_gestion_solicitudes_prestamos/gestion_solicitudes.html",
                           lista_solicitudes_por_revisar=lista_solicitudes_por_revisar,
                           lista_solicitudes_activas=lista_solicitudes_activas,
                           lista_historial_solicitudes=lista_historial_solicitudes,
                           lista_equipos_disponibles=lista_equipos_disponibles,
                           lista_solicitud_canasta=lista_solicitud_canasta,
                           lista_asignaturas=lista_asignaturas)


@mod.route("/gestion_solicitudes_prestamos/detalle_solicitud/<string:id_detalle_solicitud>", methods=["GET"])
def detalle_solicitud(id_detalle_solicitud):
    if "usuario" not in session.keys():
        return redirect("/")
    # El usuario debe ser un administrador (Credencial = 3)
    if session["usuario"]["id_credencial"] != 3:
        return redirect("/")

    # Se obtiene la información del detalle de solicitud
    sql_query = """
        SELECT Detalle_solicitud.*,Estado_detalle_solicitud.nombre AS nombre_estado
            FROM Detalle_solicitud,Estado_detalle_solicitud
                WHERE Detalle_solicitud.id = %s
                AND Detalle_solicitud.estado = Estado_detalle_solicitud.id
    """
    #cursor.execute(sql_query, (id_detalle_solicitud,))
    cursor = db.query(sql_query, (id_detalle_solicitud,))
    datos_detalle_solicitud = cursor.fetchone()

    # Si ya no existe el detalle, se redirecciona y notifica al administrador
    if datos_detalle_solicitud is None:
        flash("solicitud-no-encontrada")  # El registro fue eliminado
        return redirect("/gestion_solicitudes_prestamos")

    # Se obtiene la información general de la solicitud
    sql_query = """
        SELECT *
            FROM Solicitud
                WHERE Solicitud.id = %s
    """
    #cursor.execute(sql_query, (datos_detalle_solicitud["id_solicitud"],))
    cursor = db.query(sql_query, (datos_detalle_solicitud["id_solicitud"],))
    datos_encabezado_solicitud = cursor.fetchone()

    # Se obtiene la información del usuario solicitante
    sql_query = """
        SELECT nombres,apellidos,rut,email
            FROM Usuario
                WHERE rut = %s
    """
    #cursor.execute(sql_query, (datos_encabezado_solicitud["rut_alumno"],))
    cursor = db.query(sql_query, (datos_encabezado_solicitud["rut_alumno"],))
    datos_alumno = cursor.fetchone()

    # Se obtiene la información del equipo
    sql_query = """
        SELECT *
            FROM Equipo
                WHERE id = %s
    """
    #cursor.execute(sql_query, (datos_detalle_solicitud["id_equipo"],))
    cursor = db.query(sql_query, (datos_detalle_solicitud["id_equipo"],))
    datos_equipo = cursor.fetchone()

    # Se obtienen las cantidades de prestados y total del equipo
    sql_query = """
        SELECT COUNT(*) AS cantidad_prestados
            FROM Equipo_diferenciado
                WHERE codigo_equipo = %s
                AND activo = 1
                AND codigo_sufijo IN (SELECT codigo_sufijo_equipo
                                            FROM Detalle_solicitud
                                                WHERE id_equipo = %s
                                                    AND codigo_sufijo_equipo IS NOT NULL)
    """
    #cursor.execute(sql_query, (datos_equipo["codigo"], datos_equipo["id"]))
    cursor = db.query(sql_query, (datos_equipo["codigo"], datos_equipo["id"]))
    datos_equipo["cantidad_prestados"] = cursor.fetchone()[
        "cantidad_prestados"]

    sql_query = """
        SELECT count(*) AS cantidad_total
            FROM Equipo_diferenciado
                WHERE codigo_equipo = %s
    """
    #cursor.execute(sql_query, (datos_equipo["codigo"],))
    cursor = db.query(sql_query, (datos_equipo["codigo"],))
    datos_equipo["cantidad_total"] = cursor.fetchone()["cantidad_total"]

    sql_query = """
        SELECT count(*) AS cantidad_funcionales
            FROM Equipo_diferenciado
                WHERE codigo_equipo = %s
                AND activo = 1
    """
    #cursor.execute(sql_query, (datos_equipo["codigo"],))
    cursor = db.query(sql_query, (datos_equipo["codigo"],))
    datos_equipo["cantidad_funcionales"] = cursor.fetchone()[
        "cantidad_funcionales"]

    datos_equipo["cantidad_disponible"] = datos_equipo["cantidad_funcionales"] - \
        datos_equipo["cantidad_prestados"]

    # Se obtiene la lista de equipos y usuarios para opción de modificar solicitud
    sql_query = """
        SELECT rut,nombres,apellidos
            FROM Usuario
                ORDER BY apellidos
    """
    #cursor.execute(sql_query)
    cursor = db.query(sql_query,None)
    lista_usuarios = cursor.fetchall()

    sql_query = """
        SELECT id,marca,modelo
            FROM Equipo
                GROUP BY codigo
    """
    #cursor.execute(sql_query)
    cursor = db.query(sql_query,None)
    lista_equipos = cursor.fetchall()

    # Se obtiene la lista de equipos disponibles para prestar
    sql_query = """
        SELECT Equipo_diferenciado.codigo_equipo,Equipo_diferenciado.codigo_sufijo
            FROM Equipo_diferenciado
                WHERE codigo_equipo = %s
                AND activo = 1
                AND codigo_sufijo NOT IN (SELECT codigo_sufijo_equipo
                                            FROM Detalle_solicitud
                                                WHERE id_equipo = %s
                                                    AND codigo_sufijo_equipo IS NOT NULL)
    """
    #cursor.execute(sql_query, (datos_equipo["codigo"], datos_equipo["id"]))
    cursor = db.query(sql_query, (datos_equipo["codigo"], datos_equipo["id"]))
    lista_equipos_prestamo = cursor.fetchall()

    # Se verifica si el usuario solicitante se encuentra sancionado
    # En caso de que se encuentre sancionado, se notifica al administrador y se prohibe cambiar de estado la solicitud

    sql_query = """
        SELECT COUNT(*) AS cantidad_sanciones
            FROM Sanciones
                WHERE rut_alumno = %s
    """
    #cursor.execute(sql_query, (datos_alumno["rut"],))
    cursor = db.query(sql_query, (datos_alumno["rut"],))
    registros_sanciones = cursor.fetchone()["cantidad_sanciones"]

    if registros_sanciones:
        usuario_sancionado = True
    else:
        usuario_sancionado = False

    # Se obtiene el registro del curso asociado en caso de existir
    if datos_detalle_solicitud["id_curso_asociado"]:
        sql_query = """
            SELECT id,codigo_udp,nombre
                FROM Curso
                    WHERE id = %s
        """
        #cursor.execute(sql_query,(datos_detalle_solicitud["id_curso_asociado"],))
        cursor = db.query(sql_query,(datos_detalle_solicitud["id_curso_asociado"],))
        curso_asociado = cursor.fetchone()
    else:
        curso_asociado = None

    return render_template("/vistas_gestion_solicitudes_prestamos/detalle_solicitud.html",
                           datos_detalle_solicitud=datos_detalle_solicitud,
                           datos_encabezado_solicitud=datos_encabezado_solicitud,
                           datos_alumno=datos_alumno,
                           datos_equipo=datos_equipo,
                           lista_usuarios=lista_usuarios,
                           lista_equipos=lista_equipos,
                           lista_equipos_prestamo=lista_equipos_prestamo,
                           usuario_sancionado=usuario_sancionado,
                           curso_asociado=curso_asociado)


@mod.route("/rechazar_solicitud/<string:id_detalle>", methods=["POST"])
def rechazar_solicitud(id_detalle):
    if "usuario" not in session.keys():
        return redirect("/")
    # El usuario debe ser un administrador (Credencial = 3)
    if session["usuario"]["id_credencial"] != 3:
        return redirect("/")

    fecha_revision_solicitud = str(datetime.now().replace(microsecond=0))

    # Razón de rechazo de solicitud
    razon_rechazo = request.form.to_dict()["razon_rechazo"]

    # Se obtienen los datos al equipo y detalle de solicitud para notificar al usuario vía correo
    sql_query = """
        SELECT Equipo.nombre,Equipo.marca,Equipo.modelo,Solicitud.fecha_registro,Solicitud.rut_alumno,Detalle_solicitud.id_solicitud
            FROM Equipo,Solicitud,Detalle_solicitud
                WHERE Solicitud.id = Detalle_solicitud.id_solicitud
                AND Detalle_solicitud.id_equipo = Equipo.id
                AND Detalle_solicitud.id = %s
    """
    #cursor.execute(sql_query, (id_detalle,))
    cursor = db.query(sql_query, (id_detalle,))
    datos_solicitud_rechazada = cursor.fetchone()

    # Si la solicitud no existe, se redirecciona a la sección de gestión de solicitudes
    # Además, se notifica al administrador

    if datos_solicitud_rechazada is None:
        flash("solicitud-no-encontrada")
        return redirect("/gestion_solicitudes_prestamos")

    razon_rechazo = razon_rechazo.strip()

    # Si existe la solicitud, es marcada como rechazada (Historial)
    sql_query = """
        UPDATE Detalle_solicitud
            SET estado = 5,fecha_rechazo = %s,razon_termino = %s
                WHERE id = %s
    """
    #cursor.execute(sql_query, (datetime.now().replace(
    #    microsecond=0), razon_rechazo, id_detalle))

    db.query(sql_query,(datetime.now().replace(microsecond=0),razon_rechazo,id_detalle))

    # Por último, se notifica al usuario sobre el rechazo de la solicitud
    # Se obtienen los datos del usuario
    sql_query = """
        SELECT nombres,email FROM Usuario
            WHERE rut = %s
    """
    #cursor.execute(
    #    sql_query, (datos_solicitud_rechazada["rut_alumno"],))  # Modificar por el rut del solicitante

    cursor = db.query(sql_query,(datos_solicitud_rechazada["rut_alumno"],))
    
    datos_usuario = cursor.fetchone()

    direccion_template = os.path.normpath(os.path.join(os.getcwd(
    ), "app/templates/vistas_gestion_solicitudes_prestamos/templates_mail/rechazo_solicitud.html"))
    archivo_html = open(direccion_template, encoding="utf-8").read()

    # Se reemplazan los datos correspondientes en el archivo html
    archivo_html = archivo_html.replace("%id_solicitud%", str(
        datos_solicitud_rechazada["id_solicitud"]))
    archivo_html = archivo_html.replace("%id_detalle%", id_detalle)
    archivo_html = archivo_html.replace(
        "%nombre_usuario%", datos_usuario["nombres"])
    archivo_html = archivo_html.replace(
        "%equipo_solicitado%", datos_solicitud_rechazada["nombre"]+" "+datos_solicitud_rechazada["marca"]+" "+datos_solicitud_rechazada["modelo"])
    archivo_html = archivo_html.replace("%fecha_registro%", str(
        datos_solicitud_rechazada["fecha_registro"]))
    archivo_html = archivo_html.replace(
        "%fecha_revision_solicitud%", fecha_revision_solicitud)

    if len(razon_rechazo) == 0:
        razon_rechazo = "** No se ha adjuntado un motivo de rechazo de solicitud. **"
    else:
        razon_rechazo = razon_rechazo.replace("\n", "<br>")

    archivo_html = archivo_html.replace("%razon_rechazo%", razon_rechazo)

    enviar_correo_notificacion(
        archivo_html, "[LabEIT] Rechazo de solicitud de préstamo [IDD:"+str(id_detalle)+"]", datos_usuario["email"])

    flash("solicitud-rechazada-correctamente")
    return redirect(redirect_url())


@mod.route("/aprobar_solicitud/<string:id_detalle>", methods=["POST"])
def aprobar_solicitud(id_detalle):
    if "usuario" not in session.keys():
        return redirect("/")
    # El usuario debe ser un administrador (Credencial = 3)
    if session["usuario"]["id_credencial"] != 3:
        return redirect("/")

    datos_formulario = request.form.to_dict()

    # Se obtienen los datos del equipo según el código de sufijo del formulario
    sql_query = """
        SELECT nombre,marca,modelo
            FROM Equipo
                WHERE codigo = %s
    """
    #cursor.execute(sql_query, (datos_formulario["codigo_equipo"],))
    cursor = db.query(sql_query, (datos_formulario["codigo_equipo"],))
    datos_equipo = cursor.fetchone()

    # Se verifica que el equipo aún se encuentre registrado
    sql_query = """
        SELECT count(*) AS cantidad_registrados
            FROM Equipo_diferenciado
                WHERE codigo_equipo = %s
                AND codigo_sufijo = %s
    """
    #cursor.execute(
    #    sql_query, (datos_formulario["codigo_equipo"], datos_formulario["codigo_sufijo_prestado"]))

    cursor = db.query(sql_query,(datos_formulario["codigo_equipo"], datos_formulario["codigo_sufijo_prestado"]))
    
    cantidad_registrados = cursor.fetchone()["cantidad_registrados"]

    # El equipo fue eliminado (coincidencia de tiempos)
    if not cantidad_registrados:
        flash("equipo-no-existente")
        return redirect("/gestion_solicitudes_prestamos")

    # Se verifica que el equipo no haya sido prestado (coincidencia de tiempos)
    sql_query = """
        SELECT count(*) AS cantidad_prestados
            FROM Detalle_solicitud,Equipo
                WHERE Detalle_solicitud.id_equipo = Equipo.id
                AND Equipo.codigo = %s
                AND Detalle_solicitud.codigo_sufijo_equipo = %s
    """
    #cursor.execute(
    #    sql_query, (datos_formulario["codigo_equipo"], datos_formulario["codigo_sufijo_prestado"]))

    cursor = db.query(sql_query,(datos_formulario["codigo_equipo"], datos_formulario["codigo_sufijo_prestado"]))
    
    equipo_disponible = not bool(cursor.fetchone()["cantidad_prestados"])

    if not equipo_disponible:  # El equipo fue prestado mientras se recibía el formulario
        flash("equipo-prestado")
        return redirect(redirect_url())

    # En caso de que pueda prestarse:
    # Se agrega la hora (18:30) a la fecha de vencimiento
    datos_formulario["fecha_vencimiento_solicitud"] = datetime.strptime(
        datos_formulario["fecha_vencimiento_solicitud"], "%Y-%m-%d")
    datos_formulario["fecha_vencimiento_solicitud"] = str(
        datos_formulario["fecha_vencimiento_solicitud"].replace(hour=18, minute=30, second=0))

    # Se agrega el código sufijo a la solicitud, la fecha de vencimiento y se modifica el estado a 'Por retirar'
    sql_query = """
        UPDATE Detalle_solicitud
            SET codigo_sufijo_equipo = %s,estado = 1,fecha_vencimiento = %s
                WHERE id = %s
    """
    #cursor.execute(sql_query, (datos_formulario["codigo_sufijo_prestado"],
    #                           datos_formulario["fecha_vencimiento_solicitud"], id_detalle))

    db.query(sql_query,(datos_formulario["codigo_sufijo_prestado"],datos_formulario["fecha_vencimiento_solicitud"], id_detalle))

    # Se notifica al usuario
    # Se obtienen los datos del usuario
    sql_query = """
        SELECT rut_alumno,fecha_registro,id FROM Solicitud
            WHERE id = %s
    """
    #cursor.execute(sql_query, (datos_formulario["id_encabezado_solicitud"],))

    cursor = db.query(sql_query, (datos_formulario["id_encabezado_solicitud"],))

    datos_encabezado_solicitud = cursor.fetchone()

    sql_query = """
        SELECT nombres,email FROM Usuario
            WHERE rut = %s
    """
    #cursor.execute(
    #    sql_query, (datos_encabezado_solicitud["rut_alumno"],))  # Modificar por el rut del solicitante

    cursor = db.query(sql_query,(datos_encabezado_solicitud["rut_alumno"],))
    
    datos_usuario = cursor.fetchone()

    direccion_template = os.path.normpath(os.path.join(os.getcwd(
    ), "app/templates/vistas_gestion_solicitudes_prestamos/templates_mail/aprobacion_solicitud.html"))
    archivo_html = open(direccion_template, encoding="utf-8").read()

    # Se reemplazan los datos correspondientes en el archivo html
    archivo_html = archivo_html.replace(
        "%id_solicitud%", str(datos_encabezado_solicitud["id"]))
    archivo_html = archivo_html.replace("%id_detalle%", id_detalle)
    archivo_html = archivo_html.replace(
        "%nombre_usuario%", datos_usuario["nombres"])
    archivo_html = archivo_html.replace(
        "%equipo_solicitado%", datos_equipo["nombre"]+" "+datos_equipo["marca"]+" "+datos_equipo["modelo"])
    archivo_html = archivo_html.replace(
        "%codigo_equipo%", datos_formulario["codigo_equipo"])
    archivo_html = archivo_html.replace(
        "%codigo_sufijo%", datos_formulario["codigo_sufijo_prestado"])
    archivo_html = archivo_html.replace("%fecha_registro%", str(
        datos_encabezado_solicitud["fecha_registro"]))

    fecha_vencimiento = datetime.strptime(
        datos_formulario["fecha_vencimiento_solicitud"], "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M:%S")
    archivo_html = archivo_html.replace(
        "%fecha_vencimiento_solicitud%", str(fecha_vencimiento))

    fecha_revision_solicitud = str(datetime.now().replace(microsecond=0))
    archivo_html = archivo_html.replace(
        "%fecha_revision_solicitud%", fecha_revision_solicitud)

    enviar_correo_notificacion(
        archivo_html, "[LabEIT] Aprobación de solicitud de préstamo [IDD:"+str(id_detalle)+"]", datos_usuario["email"])

    flash("solicitud-aprobada-correctamente")
    return redirect(redirect_url())


@mod.route("/eliminar_solicitud/<string:id_detalle>", methods=["GET"])
def eliminar_solicitud(id_detalle):
    if "usuario" not in session.keys():
        return redirect("/")
    # El usuario debe ser un administrador (Credencial = 3)
    if session["usuario"]["id_credencial"] != 3:
        return redirect("/")

    # Permite eliminar el detalle de solicitud con id = id_solicitud

    # Se obtiene el ID de solicitud (id general) para determinar si se debe eliminar el encabezado
    sql_query = """
        SELECT id_solicitud
            FROM Detalle_solicitud
                WHERE id = %s
    """
    #cursor.execute(sql_query, (id_detalle,))
    cursor = db.query(sql_query, (id_detalle,))

    datos_encabezado = cursor.fetchone()

    # Se retorna en caso de que se haya eliminado el detalle de solicitud
    if datos_encabezado is None:
        return redirect(redirect_url())

    id_encabezado_solicitud = datos_encabezado["id_solicitud"]

    # Se elimina el detalle de solicitud
    sql_query = """
        DELETE FROM
            Detalle_solicitud
                WHERE id = %s
    """
    #cursor.execute(sql_query, (id_detalle,))
    db.query(sql_query,(id_detalle,))

    # Se verifica si existen más detalles asociados a la solicitud
    # Si no existen registros, entonces se elimina el encabezado

    sql_query = """
        SELECT COUNT(*) as cantidad_detalles
            FROM Detalle_solicitud
                WHERE id_solicitud = %s
    """
    #cursor.execute(sql_query, (id_encabezado_solicitud,))
    cursor = db.query(sql_query, (id_encabezado_solicitud,))

    cantidad_detalles = cursor.fetchone()["cantidad_detalles"]

    if cantidad_detalles == 0:
        sql_query = """
            DELETE FROM
                Solicitud
                    WHERE id = %s
        """
        #cursor.execute(sql_query, (id_encabezado_solicitud,))
        db.query(sql_query, (id_encabezado_solicitud,))
    
    flash("detalle-eliminado")
    return redirect(redirect_url())


@mod.route("/eliminar_detalles_seleccionados", methods=["POST"])
def eliminar_detalles_seleccionados():
    # Permite eliminar los detalles que hayan sido seleccionados
    # en las tablas de detalles pendientes, activos e historial.
    lista_detalles_solicitudes = request.form.getlist(
        "eliminar_detalle_solicitud")

    # Se retorna en caso de que no se encuentren detalles seleccionados
    if not len(lista_detalles_solicitudes):
        return redirect(redirect_url())

    # Se obtienen los ids de encabezado de solicitudes para posteriormente
    # eliminar los que tengan 0 detalles
    lista_encabezados_solicitudes = []
    for id_detalle in lista_detalles_solicitudes:
        sql_query = """
            SELECT id_solicitud
                FROM Detalle_solicitud
                    WHERE id = %s
        """
        #cursor.execute(sql_query, (id_detalle,))
        cursor = db.query(sql_query, (id_detalle,))

        registro_encabezado_solicitud = cursor.fetchone()

        if registro_encabezado_solicitud is None:
            continue

        id_encabezado_solicitud = registro_encabezado_solicitud["id_solicitud"]

        if id_encabezado_solicitud not in lista_encabezados_solicitudes:
            lista_encabezados_solicitudes.append(id_encabezado_solicitud)
        
    # Se eliminan todos los detalles de solicitudes seleccionados
    sql_query = "DELETE FROM Detalle_solicitud WHERE id IN (%s)" % ','.join(
        ['%s'] * len(lista_detalles_solicitudes))
    #cursor.execute(sql_query, lista_detalles_solicitudes)
    db.query(sql_query,lista_detalles_solicitudes)

    # Se eliminan los encabezados de solicitud en caso de que se hayan eliminado todos sus detalles
    for id_encabezado_solicitud in lista_encabezados_solicitudes:
        sql_query = """
            DELETE FROM Solicitud
                WHERE id = %s
                AND (SELECT COUNT(*) FROM Detalle_solicitud WHERE id_solicitud = %s) = 0
        """
        #cursor.execute(
        #    sql_query, (id_encabezado_solicitud, id_encabezado_solicitud))
        db.query(sql_query,(id_encabezado_solicitud, id_encabezado_solicitud))

    flash("detalles-seleccionados-eliminados")
    return redirect(redirect_url())


@mod.route("/eliminar_solicitud_canasta/<string:id_solicitud>", methods=["POST"])
def eliminar_solicitud_canasta(id_solicitud):
    # Permite eliminar un encabezado de solicitud según id_solicitud en caso de que queden 0 detalles
    # asociados a la solicitud correspondiente
    sql_query = """
        DELETE FROM
            Solicitud
                WHERE id = %s
    """
    #cursor.execute(sql_query, (id_solicitud,))
    db.query(sql_query, (id_solicitud,))

    return redirect("/gestion_solicitudes_prestamos")


@mod.route("/cancelar_solicitud/<string:id_detalle>", methods=["POST"])
def cancelar_solicitud(id_detalle):
    # Al cancelar una solicitud, se debe liberar el equipo reservado
    # ID de estado 'cancelada': 7

    datos_formulario = request.form.to_dict()

    fecha_cancelacion_solicitud = datetime.now()
    fecha_cancelacion_solicitud = str(fecha_cancelacion_solicitud.date(
    ))+" "+str(fecha_cancelacion_solicitud.hour)+":"+str(fecha_cancelacion_solicitud.minute)

    # Se libera el codigo de sufijo de equipo del detalle de solicitud y se modifica el estado
    razon_cancelacion = datos_formulario["razon_cancelacion"]
    razon_cancelacion = razon_cancelacion.strip()

    sql_query = """
        UPDATE Detalle_solicitud
            SET estado = 7,codigo_sufijo_equipo = NULL,fecha_cancelacion = %s,razon_termino = %s
                WHERE id = %s
    """
    #cursor.execute(sql_query, (datetime.now().replace(
    #    microsecond=0), razon_cancelacion, id_detalle))
    db.query(sql_query, (datetime.now().replace(microsecond=0), razon_cancelacion, id_detalle))

    # Se obtienen los datos necesarios para el correo

    # Datos del usuario
    # Se obtienen los datos del usuario
    sql_query = """
        SELECT rut_alumno,fecha_registro,id FROM Solicitud
            WHERE id = %s
    """
    #cursor.execute(sql_query, (datos_formulario["id_encabezado_solicitud"],))
    cursor = db.query(sql_query, (datos_formulario["id_encabezado_solicitud"],))

    datos_encabezado_solicitud = cursor.fetchone()

    sql_query = """
        SELECT nombres,email FROM Usuario
            WHERE rut = %s
    """
    #cursor.execute(
    #    sql_query, (datos_encabezado_solicitud["rut_alumno"],))  # Modificar por el rut del solicitante
    cursor = db.query(sql_query, (datos_encabezado_solicitud["rut_alumno"],))
    
    datos_usuario = cursor.fetchone()

    # Se obtienen los datos del equipo según el código de sufijo del formulario
    sql_query = """
        SELECT nombre,marca,modelo
            FROM Equipo
                WHERE codigo = %s
    """
    #cursor.execute(sql_query, (datos_formulario["codigo_equipo"],))
    cursor = db.query(sql_query, (datos_formulario["codigo_equipo"],))

    datos_equipo = cursor.fetchone()

    direccion_template = os.path.normpath(os.path.join(os.getcwd(
    ), "app/templates/vistas_gestion_solicitudes_prestamos/templates_mail/cancelacion_solicitud.html"))
    archivo_html = open(direccion_template, encoding="utf-8").read()

    # Se reemplazan los datos correspondientes en el archivo html
    archivo_html = archivo_html.replace(
        "%id_solicitud%", str(datos_encabezado_solicitud["id"]))
    archivo_html = archivo_html.replace("%id_detalle%", id_detalle)
    archivo_html = archivo_html.replace(
        "%nombre_usuario%", datos_usuario["nombres"])
    archivo_html = archivo_html.replace(
        "%equipo_solicitado%", datos_equipo["nombre"]+" "+datos_equipo["marca"]+" "+datos_equipo["modelo"])
    archivo_html = archivo_html.replace(
        "%codigo_equipo%", datos_formulario["codigo_equipo"])
    archivo_html = archivo_html.replace(
        "%codigo_sufijo%", datos_formulario["codigo_sufijo_equipo"])
    archivo_html = archivo_html.replace("%fecha_registro%", str(
        datos_encabezado_solicitud["fecha_registro"]))

    archivo_html = archivo_html.replace(
        "%fecha_cancelacion_solicitud%", str(fecha_cancelacion_solicitud))

    if len(razon_cancelacion) == 0:
        razon_cancelacion = "** No se ha adjuntado una razón de cancelación de solicitud. **"
    else:
        razon_cancelacion = razon_cancelacion.replace("\n", "<br>")

    archivo_html = archivo_html.replace(
        "%razon_cancelacion%", razon_cancelacion)

    enviar_correo_notificacion(
        archivo_html, "[LabEIT] Cancelación de solicitud de préstamo [IDD:"+str(id_detalle)+"]", datos_usuario["email"])

    flash("solicitud-cancelada")
    return redirect(redirect_url())


@mod.route("/entregar_equipo/<string:id_detalle>", methods=["POST"])
def entregar_equipo(id_detalle):
    # Se entrega el equipo al usuario solicitante, cambiando el estado de la solicitud
    # y agregando las fechas de inicio y término correspondientes al préstamo
    datos_formulario = request.form.to_dict()

    # Fecha en la que se marcó el retiro del equipo
    fecha_retiro = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fecha_inicio_prestamo = datetime.strptime(
        fecha_retiro, "%Y-%m-%d %H:%M:%S")

    # Se comprueba si el administrador especificó otra fecha
    # En caso de que no se haya especificado, se calcula la fecha según la cantidad de días especificada por el equipo
    if len(datos_formulario["fecha_termino_prestamo"]) == 0:
        fecha_termino_prestamo = fecha_inicio_prestamo + \
            timedelta(days=int(datos_formulario["dias_max_prestamo"]))
        # Entrega día de la semana en formato 0-4: Lunes-Viernes / 5-6: Sábado-Domingo
        dia_habil = fecha_termino_prestamo.weekday()

        if dia_habil >= 5:  # Si el día de la fecha de término es un fin de semana se debe recalcular
            if dia_habil == 6:
                delta = 1
            else:
                delta = 2
            # Se calcula la fecha de término como el Lunes inmediatamente siguiente al fin de semana
            fecha_termino_prestamo = fecha_termino_prestamo + \
                timedelta(days=delta)
    else:
        fecha_termino_prestamo = datetime.strptime(
            datos_formulario["fecha_termino_prestamo"], "%Y-%m-%d")

    # Se agrega la hora a la fecha de término, por default (18:30 PM)
    fecha_termino_prestamo = fecha_termino_prestamo.replace(
        hour=18, minute=30, second=0)

    # Se actualizan los datos del detalle de la solicitud
    sql_query = """
        UPDATE Detalle_solicitud
            SET fecha_inicio = %s,fecha_termino = %s,estado = 2,fecha_vencimiento = NULL
                WHERE id = %s
    """
    #cursor.execute(sql_query, (fecha_inicio_prestamo,
    #                           fecha_termino_prestamo, id_detalle))
    db.query(sql_query, (fecha_inicio_prestamo,fecha_termino_prestamo, id_detalle))

    # Se envía un nuevo correo al usuario con las indicaciones correspondientes
    sql_query = """
        SELECT Detalle_solicitud.*,Solicitud.rut_alumno,Solicitud.fecha_registro as fecha_registro_encabezado,Equipo.nombre,Equipo.marca,Equipo.modelo,Equipo.codigo AS codigo_equipo
            FROM Solicitud,Detalle_solicitud,Equipo
                WHERE Detalle_solicitud.id_solicitud = Solicitud.id
                AND Detalle_solicitud.id_equipo = Equipo.id
                AND Detalle_solicitud.id = %s
    """
    #cursor.execute(sql_query, (id_detalle,))
    cursor = db.query(sql_query, (id_detalle,))

    # Info de solicitud + detalle + equipo
    datos_generales_solicitud = cursor.fetchone()

    # Se obtienen los datos del usuario
    sql_query = """
        SELECT nombres,email
            FROM Usuario
                WHERE rut = %s
    """
    #cursor.execute(sql_query, (datos_generales_solicitud["rut_alumno"],))
    cursor = db.query(sql_query, (datos_generales_solicitud["rut_alumno"],))

    datos_usuario = cursor.fetchone()

    direccion_template = os.path.normpath(os.path.join(os.getcwd(
    ), "app/templates/vistas_gestion_solicitudes_prestamos/templates_mail/indicaciones_retiro_equipo.html"))
    archivo_html = open(direccion_template, encoding="utf-8").read()

    # Se reemplazan los datos correspondientes en el archivo html
    archivo_html = archivo_html.replace("%id_solicitud%", str(
        datos_formulario["id_encabezado_solicitud"]))
    archivo_html = archivo_html.replace("%id_detalle%", id_detalle)
    archivo_html = archivo_html.replace(
        "%nombre_usuario%", datos_usuario["nombres"])
    archivo_html = archivo_html.replace(
        "%equipo_prestado%", datos_generales_solicitud["nombre"]+" "+datos_generales_solicitud["marca"]+" "+datos_generales_solicitud["modelo"])
    archivo_html = archivo_html.replace(
        "%codigo_equipo%", datos_generales_solicitud["codigo_equipo"])
    archivo_html = archivo_html.replace(
        "%codigo_sufijo%", datos_generales_solicitud["codigo_sufijo_equipo"])
    archivo_html = archivo_html.replace("%fecha_registro%", str(
        datos_generales_solicitud["fecha_registro_encabezado"]))
    archivo_html = archivo_html.replace(
        "%fecha_retiro_equipo%", str(fecha_retiro))

    # Fechas de inicio y término de préstamo
    fecha_inicio_prestamo = str(datetime.strptime(
        str(fecha_inicio_prestamo), "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M:%S"))
    fecha_termino_prestamo = str(datetime.strptime(
        str(fecha_termino_prestamo), "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M:%S"))
    archivo_html = archivo_html.replace(
        "%fecha_inicio_prestamo%", fecha_inicio_prestamo)
    archivo_html = archivo_html.replace(
        "%fecha_termino_prestamo%", fecha_termino_prestamo)

    enviar_correo_notificacion(
        archivo_html, "[LabEIT] Comprobante de retiro de equipo [IDD:"+str(id_detalle)+"]", datos_usuario["email"])

    flash("retiro-correcto")
    return redirect(redirect_url())


@mod.route("/devolucion_equipo/<string:id_detalle>", methods=["POST"])
def devolucion_equipo(id_detalle):

    datos_formulario = request.form.to_dict()

    if len(datos_formulario["fecha_devolucion_equipo"]) == 0:
        fecha_devolucion_equipo = datetime.now().replace(microsecond=0)
    else:
        fecha_devolucion_equipo = datos_formulario["fecha_devolucion_equipo"]

    # Se verifica si el detalle presentaba sancion_activa = 1
    sql_query = """
        SELECT sancion_activa
            FROM Detalle_solicitud
                WHERE id = %s
    """
    #cursor.execute(sql_query, (id_detalle,))
    cursor = db.query(sql_query, (id_detalle,))

    sancion_activa_detalle = cursor.fetchone()

    detalle_sancionado = False

    if sancion_activa_detalle is not None:
        if sancion_activa_detalle["sancion_activa"]:
            detalle_sancionado = True

    # Se obtienen los datos de la solicitud para notificación vía correo electrónico
    sql_query = """
        SELECT Detalle_solicitud.*,Solicitud.rut_alumno,Solicitud.fecha_registro as fecha_registro_encabezado,Equipo.nombre,Equipo.marca,Equipo.modelo,Equipo.codigo AS codigo_equipo
            FROM Solicitud,Detalle_solicitud,Equipo
                WHERE Detalle_solicitud.id_solicitud = Solicitud.id
                AND Detalle_solicitud.id_equipo = Equipo.id
                AND Detalle_solicitud.id = %s
    """
    #cursor.execute(sql_query, (id_detalle,))
    cursor = db.query(sql_query, (id_detalle,))

    # Info de solicitud + detalle + equipo
    datos_generales_solicitud = cursor.fetchone()

    # Se modifica el detalle de solicitud
    # Se deja en 0 sancion_activa, en caso de que el detalle se encontrara bajo una sanción activa
    sql_query = """
        UPDATE Detalle_solicitud
            SET codigo_sufijo_equipo = NULL,estado = 4,fecha_devolucion = %s,sancion_activa = 0
                WHERE id = %s
    """
    #cursor.execute(sql_query, (fecha_devolucion_equipo, id_detalle))
    db.query(sql_query, (fecha_devolucion_equipo, id_detalle))

    # Se obtienen los datos del usuario
    sql_query = """
        SELECT nombres,email
            FROM Usuario
                WHERE rut = %s
    """
    #cursor.execute(sql_query, (datos_generales_solicitud["rut_alumno"],))
    cursor = db.query(sql_query, (datos_generales_solicitud["rut_alumno"],))

    datos_usuario = cursor.fetchone()

    # Si el detalle se encontraba sancionado, se verifica si existen más detalles con sanción
    if detalle_sancionado:
        # En caso de que no existan, se modifica el registro de la sanción a inactiva (activa = 0)
        sql_query = """
            SELECT COUNT(*) AS cantidad_detalles_sancionados
                FROM Detalle_solicitud,Solicitud
                    WHERE Detalle_solicitud.id_solicitud = Solicitud.id
                    AND Solicitud.rut_alumno = %s
                    AND Detalle_solicitud.sancion_activa = 1
        """
        #cursor.execute(sql_query, (datos_generales_solicitud["rut_alumno"],))
        cursor = db.query(sql_query, (datos_generales_solicitud["rut_alumno"],))

        cantidad_detalles_sancionados = int(
            cursor.fetchone()["cantidad_detalles_sancionados"])

        if cantidad_detalles_sancionados == 0:
            # Se actualiza la sanción del usuario de activa (1) a inactiva (0)
            sql_query = """
                UPDATE Sanciones
                    SET activa = 0
                        WHERE activa = 1
                        AND rut_alumno = %s
            """
            #cursor.execute(
            #    sql_query, (datos_generales_solicitud["rut_alumno"],))
            db.query(sql_query, (datos_generales_solicitud["rut_alumno"],))

    direccion_template = os.path.normpath(os.path.join(os.getcwd(
    ), "app/templates/vistas_gestion_solicitudes_prestamos/templates_mail/comprobante_devolucion.html"))
    archivo_html = open(direccion_template, encoding="utf-8").read()

    archivo_html = archivo_html.replace("%id_solicitud%", str(
        datos_formulario["id_encabezado_solicitud"]))
    archivo_html = archivo_html.replace("%id_detalle%", id_detalle)
    archivo_html = archivo_html.replace(
        "%nombre_usuario%", datos_usuario["nombres"])
    archivo_html = archivo_html.replace(
        "%equipo_devuelto%", datos_generales_solicitud["nombre"]+" "+datos_generales_solicitud["marca"]+" "+datos_generales_solicitud["modelo"])
    archivo_html = archivo_html.replace(
        "%codigo_equipo%", datos_generales_solicitud["codigo_equipo"])
    archivo_html = archivo_html.replace(
        "%codigo_sufijo%", datos_generales_solicitud["codigo_sufijo_equipo"])
    archivo_html = archivo_html.replace(
        "%fecha_devolucion_equipo%", str(fecha_devolucion_equipo))

    enviar_correo_notificacion(
        archivo_html, "[LabEIT] Comprobante de devolución de equipo [IDD:"+str(id_detalle)+"]", datos_usuario["email"])
    flash("equipo-devuelto")
    return redirect(redirect_url())


@mod.route("/finalizar_solicitud/<string:id_detalle>", methods=["GET"])
def finalizar_solicitud(id_detalle):
    if "usuario" not in session.keys():
        return redirect("/")
    # El usuario debe ser un administrador (Credencial = 3)
    if session["usuario"]["id_credencial"] != 3:
        return redirect("/")

    # Se actualiza el estado del detalle de solicitud a Finalizada.
    sql_query = """
        UPDATE Detalle_solicitud
            SET estado = 6
                WHERE id = %s
    """
    #cursor.execute(sql_query, (id_detalle,))
    db.query(sql_query, (id_detalle,))

    flash("detalle-finalizado")
    return redirect(redirect_url())

# ======== EXPORTACIONES DE SOLICITUDES ============


@mod.route("/exportar_solicitudes/<int:id_exportacion>", methods=["GET"])
def exportar_solicitudes(id_exportacion):
    if "usuario" not in session.keys():
        return redirect("/")
    # El usuario debe ser un administrador (Credencial = 3)
    if session["usuario"]["id_credencial"] != 3:
        return redirect("/")

    # El ID de exportación corresponde a la lista que se va a exportar
    # 1: Solicitudes entrantes
    # 2: Préstamos activos
    # 3: Historial de solicitudes

    # Se ejecuta la consulta SQL según el ID de exportación

    if id_exportacion == 1:
        sql_query = """
            SELECT Detalle_solicitud.id_solicitud AS 'ID de solicitud',
                Detalle_solicitud.id AS 'ID de detalle',
                Usuario.rut AS 'RUT del solicitante',
                CONCAT(Usuario.nombres," ",Usuario.apellidos) AS 'Nombre del solicitante',
                CONCAT(Equipo.nombre," ",Equipo.marca," ",Equipo.modelo) AS 'Equipo solicitado',
                Equipo.codigo AS 'Código de equipo',
                Solicitud.fecha_registro AS 'Fecha de registro',
                (SELECT Curso.nombre FROM Curso WHERE id = Detalle_solicitud.id_curso_asociado
                ) as 'Asignatura asociada'
                    FROM Detalle_solicitud,Solicitud,Equipo,Usuario
                        WHERE Detalle_solicitud.id_solicitud = Solicitud.id
                        AND Detalle_solicitud.id_equipo = Equipo.id
                        AND Solicitud.rut_alumno = Usuario.rut
                        AND Detalle_solicitud.estado = 0
                        ORDER BY Solicitud.fecha_registro DESC
        """
        #cursor.execute(sql_query)
        cursor = db.query(sql_query,None)

        lista_detalles = cursor.fetchall()

    elif id_exportacion == 2:
        sql_query = """
            SELECT Detalle_solicitud.id_solicitud AS 'ID de solicitud',
                Detalle_solicitud.id AS 'ID de detalle',
                Usuario.rut AS 'RUT del solicitante',
                CONCAT(Usuario.nombres," ",Usuario.apellidos) AS 'Nombre del solicitante',
                CONCAT(Equipo.nombre," ",Equipo.marca," ",Equipo.modelo) AS 'Equipo solicitado',
                Equipo.codigo AS 'Código de equipo',
                Detalle_solicitud.codigo_sufijo_equipo AS 'Código sufijo de equipo',
                Estado_detalle_solicitud.nombre AS Estado,
                Solicitud.fecha_registro AS 'Fecha de registro',
                Detalle_solicitud.fecha_inicio AS 'Fecha de inicio',
                Detalle_solicitud.fecha_termino AS 'Fecha de término',
                Detalle_solicitud.fecha_devolucion AS 'Fecha de devolución',
                Detalle_solicitud.fecha_vencimiento AS 'Fecha de vencimiento',
                Detalle_solicitud.renovaciones AS 'Cantidad de renovaciones',
                (SELECT Curso.nombre FROM Curso WHERE id = Detalle_solicitud.id_curso_asociado
                ) as 'Asignatura asociada'
                    FROM Detalle_solicitud,Solicitud,Equipo,Usuario,Estado_detalle_solicitud
                        WHERE Detalle_solicitud.id_solicitud = Solicitud.id
                        AND Detalle_solicitud.id_equipo = Equipo.id
                        AND Solicitud.rut_alumno = Usuario.rut
                        AND Detalle_solicitud.estado = Estado_detalle_solicitud.id
                        AND Detalle_solicitud.estado != 0
                        AND Detalle_solicitud.estado < 5
                        ORDER BY Solicitud.fecha_registro DESC
                        """
        #cursor.execute(sql_query)
        cursor = db.query(sql_query,None)

        lista_detalles = cursor.fetchall()

    elif id_exportacion == 3:
        sql_query = """
            SELECT Detalle_solicitud.id_solicitud AS 'ID de solicitud',
                Detalle_solicitud.id AS 'ID de detalle',
                Usuario.rut AS 'RUT del solicitante',
                CONCAT(Usuario.nombres," ",Usuario.apellidos) AS 'Nombre del solicitante',
                CONCAT(Equipo.nombre," ",Equipo.marca," ",Equipo.modelo) AS 'Equipo solicitado',
                Equipo.codigo AS 'Código de equipo',
                Estado_detalle_solicitud.nombre AS Estado,
                Solicitud.fecha_registro AS 'Fecha de registro',
                Detalle_solicitud.fecha_inicio AS 'Fecha de inicio',
                Detalle_solicitud.fecha_termino AS 'Fecha de término',
                Detalle_solicitud.fecha_devolucion AS 'Fecha de devolución',
                Detalle_solicitud.fecha_vencimiento AS 'Fecha de vencimiento',
                Detalle_solicitud.fecha_rechazo AS 'Fecha de rechazo',
                Detalle_solicitud.fecha_cancelacion AS 'Fecha de cancelación',
                Detalle_solicitud.renovaciones AS 'Cantidad de renovaciones',
                Detalle_solicitud.razon_termino AS 'Motivo de término',
                (SELECT Curso.nombre FROM Curso WHERE id = Detalle_solicitud.id_curso_asociado
                ) as 'Asignatura asociada'
                    FROM Detalle_solicitud,Solicitud,Equipo,Usuario,Estado_detalle_solicitud
                        WHERE Detalle_solicitud.id_solicitud = Solicitud.id
                        AND Detalle_solicitud.id_equipo = Equipo.id
                        AND Solicitud.rut_alumno = Usuario.rut
                        AND Detalle_solicitud.estado = Estado_detalle_solicitud.id
                        AND Detalle_solicitud.estado != 0
                        AND Detalle_solicitud.estado >= 5
                        ORDER BY Solicitud.fecha_registro DESC
                        """
        #cursor.execute(sql_query)
        cursor = db.query(sql_query,None)

        lista_detalles = cursor.fetchall()

    wb = Workbook()  # Instancia de libro Excel
    ws = wb.active  # Hoja activa para escribir

    if id_exportacion == 1:
        ws.title = "Solicitudes por revisar"
        nombre_archivo = "solicitudes_entrantes.xlsx"
    elif id_exportacion == 2:
        ws.title = "Solicitudes activas"
        nombre_archivo = "solicitudes_activas.xlsx"
    else:
        ws.title = "Historial de solicitudes"
        nombre_archivo = "historial_solicitudes.xlsx"

    # Creación de estilo con borde para Excel
    borde_delgado = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000'))

    # Columnas de la tabla
    lista_columnas = []
    if len(lista_detalles):
        lista_columnas = [
            nombre_columna for nombre_columna in lista_detalles[0].keys()]
    else:
        return redirect(redirect_url())

    for i in range(len(lista_columnas)):
        celda = ws.cell(row=1, column=i+1)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.border = borde_delgado
        celda.fill = PatternFill("solid", fgColor="4D4D4D")
        celda.alignment = Alignment(horizontal="left")
        celda.value = lista_columnas[i]

    # Se agregan los registros
    index_row = 2
    index_column = 1
    for detalle in lista_detalles:
        for key in detalle:
            celda = ws.cell(row=index_row, column=index_column)
            celda.value = detalle[key]
            celda.border = borde_delgado
            celda.alignment = Alignment(horizontal="left")
            index_column += 1
        index_column = 1
        index_row += 1

    # Ajuste automático de columnas en Excel
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length

    direccion_archivo = os.path.normpath(os.path.join(
        BASE_DIR, "app/static/files/exportaciones/"+nombre_archivo))
    wb.save(direccion_archivo)

    return send_file(direccion_archivo, as_attachment=True)


# ================= CANASTAS ============================
# Rutas solicitudes ágiles
@mod.route("/registrar_solicitud_agil", methods=["POST"])
def registrar_solicitud_agil():
    # Permite registrar solicitudes por parte del administrador (pensado para casos informales)
    # La solicitud se crea con todos los detalles y pasa automáticamente al estado de posesión
    rut_usuario = request.form.get("rut_usuario")
    fecha_termino = request.form.get("fecha_termino")
    lista_equipos_seleccionados = request.form.getlist(
        "id_equipo_seleccionado")
    lista_codigos_sufijos_equipos_seleccionados = request.form.getlist(
        "codigo_sufijo_equipo_seleccionado")

    # Se comprueba que el RUT sea válido
    sql_query = """
        SELECT COUNT(*) AS cantidad_usuarios
            FROM Usuario
                WHERE rut = %s
    """
    #cursor.execute(sql_query, (rut_usuario,))
    cursor = db.query(sql_query, (rut_usuario,))

    usuario_existente = bool(cursor.fetchone()["cantidad_usuarios"])

    # Si el rut ingresado no coincide con ningún usuario, se notifica el error.
    if not usuario_existente:
        flash("usuario-no-existente")
        return redirect(redirect_url())

    # Se verifica que el usuario no presente sanciones
    sql_query = """
        SELECT COUNT(*) AS cantidad_sanciones
            FROM Sanciones
                WHERE rut_alumno = %s
    """
    #cursor.execute(sql_query, (rut_usuario,))
    cursor = db.query(sql_query, (rut_usuario,))

    registros_sanciones = cursor.fetchone()["cantidad_sanciones"]

    # El usuario presenta una sanción y no puede acceder a préstamos
    if registros_sanciones:
        flash("usuario-sancionado")
        return redirect(redirect_url())

    # Se verifica que ambas listas (de equipos y códigos) coincidan en la cantidad de elementos
    if len(lista_equipos_seleccionados) != len(lista_codigos_sufijos_equipos_seleccionados):
        flash("error-listas-agiles")
        return redirect(redirect_url())

    # Se modifican las fechas para agregar la hora
    fecha_inicio = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fecha_termino_especificada = False

    if len(fecha_termino):
        fecha_termino_especificada = True
        # Si se especificó una fecha en el formulario, se considera esa.
        # En caso contrario, se registra para cada uno de los detalles según la cantidad de días
        fecha_termino = datetime.strptime(
            fecha_termino, "%Y-%m-%d").replace(hour=18, minute=30, second=0)

    # Se crea la solicitud de préstamo con el rut y la fecha correspondiente
    sql_query = """
        INSERT INTO Solicitud (rut_alumno,fecha_registro)
            VALUES (%s,%s)
    """
    #cursor.execute(sql_query, (rut_usuario, fecha_inicio))
    cursor = db.query(sql_query, (rut_usuario, fecha_inicio))

    id_solicitud = cursor.lastrowid  # Se obtiene el id de solicitud recién creada

    # Se crean los detalles de solicitud según las listas de equipos y códigos
    for id_equipo, codigo_sufijo in zip(lista_equipos_seleccionados, lista_codigos_sufijos_equipos_seleccionados):

        if not fecha_termino_especificada:
            sql_query = """
                SELECT dias_max_prestamo
                    FROM Equipo
                        WHERE id = %s
            """
            #cursor.execute(sql_query, (id_equipo,))
            cursor = db.query(sql_query, (id_equipo,))

            cantidad_dias_prestamo = cursor.fetchone()["dias_max_prestamo"]

            if cantidad_dias_prestamo is None:
                cantidad_dias_prestamo = 5

            fecha_termino = datetime.strptime(
                fecha_inicio, "%Y-%m-%d %H:%M:%S") + timedelta(days=cantidad_dias_prestamo)
            # Entrega día de la semana en formato 0-4: Lunes-Viernes / 5-6: Sábado-Domingo
            dia_habil = fecha_termino.weekday()

            if dia_habil >= 5:  # Si el día de la fecha de término es un fin de semana se debe recalcular
                if dia_habil == 6:
                    delta = 1
                else:
                    delta = 2
                # Se calcula la fecha de término como el Lunes inmediatamente siguiente al fin de semana
                fecha_termino += timedelta(days=delta)

            fecha_termino = fecha_termino.replace(hour=18, minute=30, second=0)

        sql_query = """
            INSERT INTO Detalle_solicitud
                (id_solicitud,id_equipo,fecha_inicio,fecha_termino,estado,codigo_sufijo_equipo)
                    VALUES (%s,%s,%s,%s,2,%s)
        """
        #cursor.execute(sql_query, (id_solicitud, id_equipo,
        #                           fecha_inicio, fecha_termino, codigo_sufijo))
        db.query(sql_query, (id_solicitud, id_equipo,fecha_inicio, fecha_termino, codigo_sufijo))

    flash("solicitud-registrada")
    return redirect(redirect_url())

# Canasta de solicitud


@mod.route("/gestion_solicitudes_prestamos/canasta_solicitud/<string:id_solicitud>", methods=["GET"])
def detalle_canasta_solicitud(id_solicitud):
    if "usuario" not in session.keys():
        return redirect("/")
    # El usuario debe ser un administrador (Credencial = 3)
    if session["usuario"]["id_credencial"] != 3:
        return redirect("/")

    # Se obtiene la solicitud
    sql_query = """
        SELECT *
            FROM Solicitud
                WHERE id = %s
    """
    #cursor.execute(sql_query, (id_solicitud,))
    cursor = db.query(sql_query, (id_solicitud,))

    solicitud = cursor.fetchone()

    if solicitud is None:
        return redirect("/gestion_solicitudes_prestamos")

    # Se obtienen los datos principales del usuario
    sql_query = """
        SELECT rut,nombres,apellidos,email
            FROM Usuario
                WHERE rut = %s
    """
    #cursor.execute(sql_query, (solicitud["rut_alumno"],))
    cursor = db.query(sql_query, (solicitud["rut_alumno"],))

    datos_usuario = cursor.fetchone()

    # Se obtienen cada uno de los detalles asociados a la solicitud
    sql_query = """
        SELECT Detalle_solicitud.*,Equipo.codigo AS codigo_equipo,Equipo.nombre AS nombre_equipo,Equipo.modelo AS modelo_equipo,Equipo.marca AS marca_equipo,
            Equipo.dias_max_prestamo AS dias_max_prestamo_equipo,Estado_detalle_solicitud.nombre AS nombre_estado
            FROM Detalle_solicitud,Equipo,Estado_detalle_solicitud
                WHERE id_solicitud = %s
                AND Detalle_solicitud.id_equipo = Equipo.id
                AND Detalle_solicitud.estado = Estado_detalle_solicitud.id
    """
    #cursor.execute(sql_query, (id_solicitud,))
    cursor = db.query(sql_query, (id_solicitud,))

    lista_detalles_solicitud = cursor.fetchall()

    # Se almacena como llave el ID del detalle de solicitud y como valor la lista de equipos
    # disponibles para prestar
    equipos_disponibles = {}
    for detalle_solicitud in lista_detalles_solicitud:
        # En caso de que el detalle se encuentre aún pendiente, se obtiene la lista
        # de equipos disponibles para prestar
        if detalle_solicitud["estado"] == 0:
            sql_query = """
                SELECT Equipo_diferenciado.codigo_equipo,Equipo_diferenciado.codigo_sufijo
                    FROM Equipo_diferenciado,Equipo
                        WHERE Equipo.codigo = Equipo_diferenciado.codigo_equipo
                        AND Equipo.id = %s
                        AND activo = 1
                        AND codigo_sufijo NOT IN (SELECT codigo_sufijo_equipo
                                                    FROM Detalle_solicitud
                                                        WHERE id_equipo = %s
                                                            AND codigo_sufijo_equipo IS NOT NULL)
            """
            #cursor.execute(
            #    sql_query, (detalle_solicitud["id_equipo"], detalle_solicitud["id_equipo"]))
            cursor = db.query(sql_query, (detalle_solicitud["id_equipo"], detalle_solicitud["id_equipo"]))
            
            equipos_disponibles[detalle_solicitud["id"]] = cursor.fetchall()
        
        # Se obtiene el registro del curso asociado en caso de existir
        if detalle_solicitud["id_curso_asociado"]:
            sql_query = """
                SELECT Curso.id AS id_curso,Curso.codigo_udp AS curso_codigo_udp,Curso.nombre AS nombre_curso
                    FROM Curso
                        WHERE id = %s
            """
            #cursor.execute(sql_query,(detalle_solicitud["id_curso_asociado"],))
            cursor = db.query(sql_query,(detalle_solicitud["id_curso_asociado"],))

            curso_asociado = cursor.fetchone()
            # Se agrega la info del curso al detalle de solicitud en caso de que exista
            if curso_asociado is not None:
                detalle_solicitud.update(curso_asociado)

    # Se verifica si el usuario solicitante se encuentra sancionado
    # En caso de que se encuentre sancionado, se notifica al administrador y se prohibe cambiar de estado la solicitud

    sql_query = """
        SELECT COUNT(*) AS cantidad_sanciones
            FROM Sanciones
                WHERE rut_alumno = %s
    """
    #cursor.execute(sql_query, (datos_usuario["rut"],))
    cursor = db.query(sql_query, (datos_usuario["rut"],))

    registros_sanciones = cursor.fetchone()["cantidad_sanciones"]

    if registros_sanciones:
        usuario_sancionado = True
    else:
        usuario_sancionado = False

    return render_template("/vistas_gestion_solicitudes_prestamos/detalle_canasta_solicitud.html",
                           solicitud=solicitud,
                           equipos_disponibles=equipos_disponibles,
                           datos_usuario=datos_usuario,
                           lista_detalles_solicitud=lista_detalles_solicitud,
                           usuario_sancionado=usuario_sancionado)
