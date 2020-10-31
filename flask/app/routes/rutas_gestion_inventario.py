import json
import os,time
from openpyxl import Workbook
from jinja2 import Environment
from config import db,BASE_DIR
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from flask import Flask,Blueprint,render_template,request,redirect,url_for,flash,session,jsonify, make_response,send_file

def redirect_url(default='index'): # Redireccionamiento desde donde vino la request
    return request.args.get('next') or \
           request.referrer or \
           url_for(default)

mod = Blueprint("rutas_gestion_inventario",__name__)

#*********************************************************************************************#

# ***** Importante VISTA PRINCIPAL GESTION INVENTARIO **** #

# Consulta una lista con los equipos
# Es usada en la vista general de todos los equipos
def consultar_lista_equipos_general():
    # Nota importante:
    # actualizar la query papra definir equipos_disponibles y total_equipos
    query = ('''
        SELECT *,
            COUNT(CASE WHEN Detalle_solicitud.estado IN (1, 2, 3) THEN 1
            ELSE NULL END) AS en_prestamo
        FROM (SELECT
                Equipo.id AS equipo_id,
                Equipo.codigo,
                Equipo.nombre,
                Equipo.modelo,
                Equipo.marca,
                Equipo.descripcion,
                Equipo.dias_max_prestamo,
                Equipo.dias_renovacion,
                Equipo.imagen,
                COUNT(CASE WHEN Equipo_diferenciado.activo = 1 THEN 1 ELSE NULL END) AS disponibles,
                COUNT(Equipo_diferenciado.activo) AS total_equipos
                FROM Equipo
                    LEFT JOIN Equipo_diferenciado ON  Equipo.codigo = Equipo_diferenciado.codigo_equipo
                GROUP BY Equipo.id) AS inventario_general
            LEFT JOIN Detalle_solicitud ON inventario_general.equipo_id = Detalle_solicitud.id_equipo
        GROUP BY inventario_general.equipo_id
            ''')
    #cursor.execute(query)
    cursor = db.query(query,None)
    equipos = cursor.fetchall()
    return equipos
def consultar_lista_equipos(): # funcion para poder conusultar toda la lista de equiipos detallados
    query = ("""
        SELECT Equipo_diferenciado.id,
            Equipo_diferenciado.codigo_equipo,
            Equipo_diferenciado.codigo_sufijo,
            Equipo_diferenciado.codigo_activo,
            Equipo_diferenciado.fecha_compra,
            Equipo.nombre,
            Equipo.modelo,
            Equipo.marca,
            Equipo_diferenciado.activo,
            Equipo_diferenciado.razon_inactivo,
            CASE WHEN Equipo_diferenciado.activo = 0 THEN 'No disponible'
                WHEN Detalle_solicitud.estado = 1 THEN 'Por retirar'
                WHEN Detalle_solicitud.estado = 2 THEN 'En posesión'
                WHEN Detalle_solicitud.estado = 3 THEN 'Con atraso'
                ELSE 'Disponible' END AS estado,
            Detalle_solicitud.codigo_sufijo_equipo
        FROM Equipo_diferenciado
        LEFT JOIN Equipo ON Equipo.codigo = Equipo_diferenciado.codigo_equipo
        LEFT JOIN (SELECT * FROM Detalle_solicitud WHERE Detalle_solicitud.estado IN (1,2,3)) AS Detalle_solicitud
            ON Detalle_solicitud.id_equipo = Equipo.id
            AND Detalle_solicitud.codigo_sufijo_equipo = Equipo_diferenciado.codigo_sufijo
        ORDER BY Equipo_diferenciado.codigo_equipo
    """)
    #cursor.execute(query)
    cursor = db.query(query,None)
    resultado = cursor.fetchall()
    return resultado

@mod.route("/gestion_inventario_admin")
def gestion_inventario_admin():
    if 'usuario' not in session or session["usuario"]["id_credencial"] != 3:
        return redirect('/')
    else:
        equipos = consultar_lista_equipos_general()
        equipos_detalle = consultar_lista_equipos()
        circuitos = consultar_lista_circuito()
        return render_template('vistas_gestion_inventario/gestion_inventario.html',
            lista_equipo = equipos,
            lista_equipo_detalle = equipos_detalle,
            lista_circuitos = circuitos)


# **** VISTA_PRINCIPAL/MODAL "AGREGAR EQUIPO" **** #

def insertar_lista_equipos_general(valores_a_insertar):
    query = ('''
        INSERT INTO Equipo (codigo, nombre, modelo, marca, descripcion, imagen, dias_max_prestamo, dias_renovacion)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s);
    ''')
    db.query(query,(valores_a_insertar['codigo'],
                    valores_a_insertar['nombre'],
                    valores_a_insertar['modelo'],
                    valores_a_insertar['marca'],
                    valores_a_insertar['descripcion'],
                    valores_a_insertar['imagen'],
                    valores_a_insertar['dias_maximo_prestamo'],
                    valores_a_insertar['dias_renovacion']))

    return valores_a_insertar['codigo']

def insertar_lista_equipos_detalle_masiva(codigo_equipo, fecha_compra, unidad_inicial, unidad_final):
    for codigo_sufijo in range(unidad_inicial, unidad_final + 1):
        equipo = {  'codigo_sufijo': codigo_sufijo,
                    'codigo_activo': '',
                    'fecha_compra': fecha_compra}
        insertar_lista_equipos_detalle(codigo_equipo, equipo) 
    return
# Funcion en AJAX para validar que el codigo del equipo a insertar no esté duplicado 
@mod.route("/gestion_inventario_admin/insertar/validar_codigo", methods = ['POST']) # AJAX
def validar_codigo_equipo_nuevo():
    if request.method == "POST" and session["usuario"]["id_credencial"] == 3:
        
        try:
            
            CODIGO = request.form["codigo"]
            query = ('''SELECT *
                        FROM Equipo
                        WHERE Equipo.codigo = %s''')
            cursor = db.query(query,(CODIGO,))
            
            if len(cursor.fetchall()) > 0: # Si hay coincidencias retorna verdadero
                
                return jsonify({'match':'True'})
            
            else: # Si no hay coicidencias retorna False
                
                return jsonify({'match':'False'})
        except:
            
            return jsonify({'error':'missing_data'})
            
    return jsonify({'error':'missing data!'})

@mod.route("/gestion_inventario_admin/insert", methods = ['POST'])
def funcion_añadir_equipo_form():
    if request.method == "POST" and session["usuario"]["id_credencial"] == 3:
        try:

            informacion_a_insertar = request.form.to_dict() 

            if "checkbox-agregar-varios" in informacion_a_insertar:
                
                unidad_inicial = int(informacion_a_insertar["unidad_inicio"])   # Obtiene el sufijo inicial
                unidad_final = int(informacion_a_insertar["unidad_final"])      # Obtiene el sufijo final
                
                insertar_lista_equipos_detalle_masiva(informacion_a_insertar["codigo"], 
                                                      informacion_a_insertar["fecha_compra"], 
                                                      unidad_inicial, unidad_final)
                

            insertar_lista_equipos_general(informacion_a_insertar)
            flash("equipo-agregado")
            
        except:
            flash("error")

    return redirect('/gestion_inventario_admin')        


# **** VISTA_PRINCIPAL/MODAL "EDITAR EQUIPO" **** #
# Funcion para actualizar la informacion de equipo, junto a los equipos relacionados 
def editar_equipo_general(informacion_a_actualizar):
    query = ('''UPDATE Equipo
                LEFT JOIN Equipo_diferenciado ON Equipo_diferenciado.codigo_equipo = Equipo.codigo
                    SET Equipo_diferenciado.codigo_equipo = %s,
                        Equipo.codigo = %s,
                        Equipo.nombre = %s,
                        Equipo.modelo = %s,
                        Equipo.marca = %s,
                        Equipo.imagen = %s,
                        Equipo.descripcion = %s,
                        Equipo.dias_max_prestamo = %s,
                        Equipo.dias_renovacion = %s
                    WHERE
                        Equipo.codigo = %s''')
    # A partir del codigo original se actualizan los equipos
    db.query(query,(informacion_a_actualizar['codigo'],
                    informacion_a_actualizar['codigo'],
                    informacion_a_actualizar['nombre'],
                    informacion_a_actualizar['modelo'],
                    informacion_a_actualizar['marca'],
                    informacion_a_actualizar['imagen'],
                    informacion_a_actualizar['descripcion'],
                    informacion_a_actualizar['dias_max_prestamo'],
                    informacion_a_actualizar['dias_renovacion'],
                    informacion_a_actualizar['codigo_original']))
    return informacion_a_actualizar

def editar_equipo_especifico(informacion_a_actualizar):
    query = ('''UPDATE Equipo_diferenciado
                SET Equipo_diferenciado.codigo_sufijo = %s,
                    Equipo_diferenciado.fecha_compra = %s,
                    Equipo_diferenciado.codigo_activo = %s,
                    Equipo_diferenciado.activo = %s,
                    Equipo_diferenciado.razon_inactivo = %s
                WHERE Equipo_diferenciado.codigo_sufijo = %s
                AND Equipo_diferenciado.codigo_equipo = %s''')


    db.query(query,(informacion_a_actualizar['codigo_sufijo'],
                    informacion_a_actualizar['fecha_compra'],
                    informacion_a_actualizar['codigo_activo'],
                    informacion_a_actualizar['activo'],
                    informacion_a_actualizar['razon_inactivo'],
                    informacion_a_actualizar['codigo_sufijo_original'],
                    informacion_a_actualizar['codigo_equipo']))

    return informacion_a_actualizar



@mod.route("/gestion_inventario_admin/lista_equipo_diferenciado/actualizar_informacion/validar_codigo", methods = ['POST'])
def validar_codigo_sufijo_editar():
    if request.method == "POST" and session["usuario"]["id_credencial"] == 3:
        
        try:
            
            CODIGO_EQUIPO = request.form["codigo_equipo"]
            CODIGO_SUFIJO = request.form["codigo_sufijo"]
            CODIGO_SUFIJO_NUEVO = request.form["codigo_sufijo_nuevo"]

            if CODIGO_SUFIJO_NUEVO == CODIGO_SUFIJO: # Si el codigo no varia, valida el codigo
                return jsonify({'match':'False'})
            
            query = ('''SELECT * FROM Equipo_diferenciado
                        WHERE Equipo_diferenciado.codigo_equipo = %s 
                            AND Equipo_diferenciado.codigo_sufijo = %s''')
            cursor = db.query(query,(CODIGO_EQUIPO, CODIGO_SUFIJO_NUEVO)) # Busca si hay un equipo con el codigo:sufijo nuevo
            
            
            
            if len(cursor.fetchall()) > 0: # Si hay coincidencias retorna verdadero
                
                return jsonify({'match':'True'})
            
            else: # Si no hay coicidencias retorna False
                
                return jsonify({'match':'False'})
        except:
            
            return jsonify({'error':'missing_data'})
            
    return jsonify({'error':'missing data!'})

#Actualizar informacion equipo diferenciado
@mod.route('/gestion_inventario_admin/lista_equipo_diferenciado/actualizar_informacion', methods = ['POST'])
def funcion_editar_equipo_diferenciado_form():
    if request.method == "POST" and session["usuario"]["id_credencial"] == 3:
        informacion_a_actualizar = request.form.to_dict()
        equipos = consultar_lista_equipos_detalle(informacion_a_actualizar["codigo_equipo"])
        if(informacion_a_actualizar["codigo_sufijo_original"]==informacion_a_actualizar["codigo_sufijo"]):
            flash("equipo-editado")
            editar_equipo_especifico(informacion_a_actualizar)
            return redirect("/gestion_inventario_admin/lista_equipo_diferenciado/"+informacion_a_actualizar["codigo_equipo"])
        for val in equipos:
            if (val["codigo_sufijo"]==informacion_a_actualizar["codigo_sufijo"]):
                    flash("codigo-equipo-existente")
                    return redirect("/gestion_inventario_admin/lista_equipo_diferenciado/"+informacion_a_actualizar["codigo_equipo"])
        flash("equipo-editado")
        editar_equipo_especifico(informacion_a_actualizar)
        return redirect("/gestion_inventario_admin/lista_equipo_diferenciado/"+informacion_a_actualizar["codigo_equipo"])

#Actualizar información del equipo

@mod.route('/gestion_inventario_admin/actualizar_informacion', methods = ['POST'])
def funcion_editar_equipo():
    if request.method == "POST" and session["usuario"]["id_credencial"] == 3:
        informacion_a_actualizar = request.form.to_dict()
        equipos = consultar_lista_equipos_general()
        if (informacion_a_actualizar["codigo"]==informacion_a_actualizar["codigo_original"]):
            flash("equipo-editado")
            editar_equipo_general(informacion_a_actualizar)
            return redirect("/gestion_inventario_admin")
        for val in equipos:
            if (val["codigo"]==informacion_a_actualizar["codigo"]):
                    flash("codigo-equipo-existente")
                    return redirect('/gestion_inventario_admin')

        editar_equipo_general(informacion_a_actualizar)
        flash("equipo-editado")
        return redirect("/gestion_inventario_admin")

# **** VISTA_PRINCIPAL/MODAL "BORRAR EQUIPO" **** #

def eliminar_equipo_general(equipo):
    query = ('''
        DELETE Equipo, Equipo_diferenciado
        FROM Equipo
        LEFT JOIN Equipo_diferenciado ON Equipo_diferenciado.codigo_equipo = Equipo.codigo
        WHERE Equipo.codigo = %s
    ''')
    #cursor.execute(query,(equipo['codigo'],))

    db.query(query,(equipo['codigo'],))

    return 'ok'

#Ruta eliminar equipo

@mod.route("/gestion_inventario_admin/delete",methods=["POST"])
def funcion_eliminar_equipo():
    if request.method == "POST" and session["usuario"]["id_credencial"] == 3:
        equipo_a_eliminar = request.form.to_dict()
        equipos = consultar_lista_equipos_detalle_estado(equipo_a_eliminar["codigo"])
        for val in equipos:
            if (val["estado"]==1 or val["estado"]==2 or val["estado"]==3):
                flash("equipo-ocupado")
                return redirect("/gestion_inventario_admin")
        eliminar_equipo_general(equipo_a_eliminar)
        flash("equipo-eliminado")
        return redirect("/gestion_inventario_admin")




#*********************************************************************************************#

# **** VISTA_PRINCIPAL/MODAL "BORRAR EQUIPO DIFERENCIADO" **** #

def eliminar_equipo_vista_diferenciado(equipo):
    query = ('''
        DELETE Equipo_diferenciado
        FROM Equipo_diferenciado
        WHERE Equipo_diferenciado.codigo_equipo = %s
        AND Equipo_diferenciado.codigo_sufijo = %s
    ''')
    #cursor.execute(query,(
    #equipo['codigo_equipo'],
    #equipo['codigo_sufijo']
    #))

    db.query(query,(equipo['codigo_equipo'],equipo['codigo_sufijo']))

    return 'ok'

#Ruta eliminar equipo
def consultar_equipo_especifico_estado(codigo_equipo,codigo_sufijo):
    query = ('''
        SELECT
            Equipo_diferenciado.codigo_equipo,
            Equipo_diferenciado.codigo_sufijo,
            Equipo_diferenciado.activo,
            Equipo.id AS equipo_id,
            Detalle_solicitud.id AS detalle_sol_id,
            Detalle_solicitud.codigo_sufijo_equipo,
            Detalle_solicitud.estado,
            Estado_detalle_solicitud.nombre,
            Solicitud.rut_alumno,
            Detalle_solicitud.fecha_inicio,
            Detalle_solicitud.fecha_termino,
            Detalle_solicitud.fecha_vencimiento
        FROM Equipo_diferenciado
        LEFT JOIN Equipo ON Equipo.codigo = Equipo_diferenciado.codigo_equipo
        LEFT JOIN Detalle_solicitud ON Detalle_solicitud.id_equipo = Equipo.id
            AND Detalle_solicitud.codigo_sufijo_equipo = Equipo_diferenciado.codigo_sufijo
        LEFT JOIN Solicitud ON Solicitud.id=Detalle_solicitud.id_solicitud
        LEFT JOIN Estado_detalle_solicitud ON Estado_detalle_solicitud.id=Detalle_solicitud.estado
        WHERE Equipo_diferenciado.codigo_equipo = %s
            AND Equipo_diferenciado.codigo_sufijo = %s

    '''
    )
    #cursor.execute(query,(codigo_equipo,codigo_sufijo,))
    cursor = db.query(query,(codigo_equipo,codigo_sufijo,))
    equipos_detalle = cursor.fetchone()
    return equipos_detalle

@mod.route("/gestion_inventario_admin/lista_equipo_diferenciado/delete",methods=["POST"])
def funcion_eliminar_equipo_diferenciado():
    if request.method == "POST" and session["usuario"]["id_credencial"] == 3:
        equipo_a_eliminar = request.form.to_dict()
        equipos = consultar_equipo_especifico_estado(equipo_a_eliminar["codigo_equipo"],equipo_a_eliminar["codigo_sufijo"])
        if (equipos["estado"]==1 or equipos["estado"]==2 or equipos["estado"]==3):
            flash("equipo-ocupado")
            return redirect("/gestion_inventario_admin/lista_equipo_diferenciado/"+equipo_a_eliminar["codigo_equipo"])
        eliminar_equipo_vista_diferenciado(equipo_a_eliminar)
        flash("equipo-eliminado")
        return redirect("/gestion_inventario_admin/lista_equipo_diferenciado/"+equipo_a_eliminar["codigo_equipo"])




#*********************************************************************************************#
#Informacion de equipo unico para listado de equipo_diferenciado
def consultar_equipo_descripcion(codigo):
        query = ('''
            SELECT  Equipo.id,
                    Equipo.codigo,
                    Equipo.nombre,
                    Equipo.modelo,
                    Equipo.marca,
                    Equipo.descripcion,
                    Equipo.dias_max_prestamo,
                    Equipo.dias_renovacion,
                    Equipo.imagen,
                    COUNT(CASE WHEN Equipo_diferenciado.activo = 1 THEN 1 ELSE NULL END) AS disponibles,
                    COUNT(Equipo_diferenciado.activo) AS total_equipos
                    FROM Equipo
                    LEFT JOIN Equipo_diferenciado ON  Equipo.codigo = Equipo_diferenciado.codigo_equipo
                    WHERE Equipo.codigo= %s
                ''')
        #cursor.execute(query,(codigo,))
        cursor = db.query(query,(codigo,))
        equipo_detalle = cursor.fetchone()
        return equipo_detalle

# Consulta una  lista de equipos especificos
def consultar_lista_equipos_detalle_estado(codigo_equipo):
    query = ('''
        SELECT
            Equipo_diferenciado.codigo_equipo,
            Equipo_diferenciado.codigo_sufijo,
            Equipo_diferenciado.activo,
            Equipo.id AS equipo_id,
            Detalle_solicitud.id AS detalle_sol_id,
            Detalle_solicitud.codigo_sufijo_equipo,
            Detalle_solicitud.estado,
            Estado_detalle_solicitud.nombre,
            Solicitud.rut_alumno,
            Detalle_solicitud.fecha_inicio,
            Detalle_solicitud.fecha_termino,
            Detalle_solicitud.fecha_vencimiento
        FROM Equipo_diferenciado
        LEFT JOIN Equipo ON Equipo.codigo = Equipo_diferenciado.codigo_equipo
        LEFT JOIN Detalle_solicitud ON Detalle_solicitud.id_equipo = Equipo.id
            AND Detalle_solicitud.codigo_sufijo_equipo = Equipo_diferenciado.codigo_sufijo
        LEFT JOIN Solicitud ON Solicitud.id=Detalle_solicitud.id_solicitud
        LEFT JOIN Estado_detalle_solicitud ON Estado_detalle_solicitud.id=Detalle_solicitud.estado
        WHERE Equipo_diferenciado.codigo_equipo = %s
            AND Detalle_solicitud.estado IN (1,2,3)
    '''
    )
    #cursor.execute(query,(codigo_equipo,))
    cursor = db.query(query,(codigo_equipo,))
    equipos_detalle = cursor.fetchall()
    return equipos_detalle


#Vista más informacion equipo , sin tabla solicitudes definida
@mod.route("/gestion_inventario_admin/detalles_equipo/<string:codigo_equipo>",methods=["GET"])
def detalle_info_equipo(codigo_equipo):
    equipos = consultar_lista_equipos_detalle_estado(codigo_equipo)
    equipo_descripcion = consultar_equipo_descripcion(codigo_equipo)
    return render_template("/vistas_gestion_inventario/detalles_equipo.html",
    equipo_descripcion=equipo_descripcion, equipos_detalle = equipos)


# Importante FUNCION() ENCARGADA DE INGRESAR LOS VALORES DEL FORMULARIO "AGREGAR" EN VISTA GESTION INVENTARIO DIFERENCIAD0 ** #

# Funcion en AJAX para validar que el codigo_sufijo del equipo no exista
@mod.route("/gestion_inventario_admin/lista_equipo_diferenciado/agregar_equipo/validar_codigo", methods = ['POST'])
def validar_codigo_sufijo_nuevo():
    if request.method == "POST" and session["usuario"]["id_credencial"] == 3:
        
        try:
            
            CODIGO_EQUIPO = request.form["codigo_equipo"]
            CODIGO_SUFIJO = request.form["codigo_sufijo"]

            
            query = ('''SELECT * FROM Equipo_diferenciado
                        WHERE Equipo_diferenciado.codigo_equipo = %s 
                            AND Equipo_diferenciado.codigo_sufijo = %s''')
            
            cursor = db.query(query,(CODIGO_EQUIPO, CODIGO_SUFIJO))
            
            if len(cursor.fetchall()) > 0: # Si hay coincidencias retorna verdadero
                
                return jsonify({'match':'True'})
            
            else: # Si no hay coicidencias retorna False
                
                return jsonify({'match':'False'})
        except:
            
            return jsonify({'error':'missing_data'})
            
    return jsonify({'error':'missing data!'})


@mod.route("/gestion_inventario_admin/lista_equipo_diferenciado/agregar_equipo", methods = ['POST'])
def validar_form_añadir_equipo_espeficio():
    if request.method == "POST" and session["usuario"]["id_credencial"] == 3: # si es admin y metodo POST
        try:
            
            informacion_a_insertar = request.form.to_dict() # Obtiene la informacion del formulario
            
            codigo_equipo = informacion_a_insertar["codigo_equipo"] # obtiene el codigo general del equipo
            
            insertar_lista_equipos_detalle(codigo_equipo, informacion_a_insertar)
            
            flash("equipo-agregado")
        except:
            flash("error")

        return redirect("/gestion_inventario_admin/lista_equipo_diferenciado/" + codigo_equipo)


# Agrega un equipo a partir de la relacion que tenga DIF
def insertar_lista_equipos_detalle(codigo_equipo, valores_a_insertar):
    query = ('''INSERT INTO Equipo_diferenciado (codigo_equipo, codigo_sufijo, codigo_activo, fecha_compra)
                VALUES (%s, %s, %s, %s)''')
    db.query(query,(codigo_equipo,
                    valores_a_insertar['codigo_sufijo'],
                    valores_a_insertar['codigo_activo'],
                    valores_a_insertar['fecha_compra']))

    return


def consultar_lista_equipos_detalle(codigo_equipo):
    query = ('''
        SELECT *
        FROM Equipo_diferenciado
        WHERE Equipo_diferenciado.codigo_equipo = %s
    '''
    )
    #cursor.execute(query,(codigo_equipo,))
    cursor = db.query(query,(codigo_equipo,))
    equipos_detalle = cursor.fetchall()
    return equipos_detalle

# ** Importante VISTA GESTION INVENTARIO DIFERENCIADO ** #
@mod.route("/gestion_inventario_admin/lista_equipo_diferenciado/<string:codigo_equipo>")
def gestion_inventario_admin_equipo(codigo_equipo):
    if 'usuario' not in session or session["usuario"]["id_credencial"] != 3: # Si no está logeado, redirigie al login
        return redirect('/')
    equipos = consultar_lista_equipos_detalle(codigo_equipo)
    equipos_descripcion = consultar_equipo_descripcion(codigo_equipo)
    return render_template('vistas_gestion_inventario/similares_tabla.html',
    equipos_descripcion=equipos_descripcion, equipos_detalle = equipos, equipo_padre = codigo_equipo) #Cambiar render por el que corresponda



#*********************************************************************************************#

# ***** Importante VISTA CIRCUITOS **** #

#Consulta tabla principal
def consultar_lista_circuito():
    query = ('''
        SELECT *
        FROM Circuito
    '''
    )
    #cursor.execute(query)
    cursor = db.query(query,None)
    circuitos = cursor.fetchall()
    return circuitos


#Consulta para insertar circuito
def insertar_lista_circuitos(valores_a_insertar):
    query = ('''
        INSERT INTO Circuito (nombre, cantidad, descripcion, dias_max_prestamo, dias_renovacion, imagen)
        VALUES (%s, %s, %s, %s, %s, %s);
    ''')
    #cursor.execute(query,(
    #    valores_a_insertar['nombre_circuito'],
    #    valores_a_insertar['cantidad_circuito'],
    #    valores_a_insertar['descripcion_circuito'],
    #    valores_a_insertar['dias_max_prestamo'],
    #    valores_a_insertar['dias_renovacion'],
    #    valores_a_insertar['imagen_circuito']))

    db.query(query,(
        valores_a_insertar['nombre_circuito'],
        valores_a_insertar['cantidad_circuito'],
        valores_a_insertar['descripcion_circuito'],
        valores_a_insertar['dias_max_prestamo'],
        valores_a_insertar['dias_renovacion'],
        valores_a_insertar['imagen_circuito']))

    return 'OK'

#funcion insertar nuevo circuito
@mod.route("/gestion_inventario_admin/insert_circuito", methods = ['POST'])
def funcion_añadir_circuito_form():
    if request.method == "POST" and session["usuario"]["id_credencial"] == 3:
        informacion_a_insertar = request.form.to_dict()
        circuitos=consultar_lista_circuito()
        for val in circuitos:
            if (informacion_a_insertar["nombre_circuito"]==val["nombre"]
            and informacion_a_insertar["descripcion_circuito"]==val["descripcion"] ):
                flash("equipo-existente")
                return redirect('/gestion_inventario_admin')
        insertar_lista_circuitos(informacion_a_insertar)
        flash("equipo-agregado")
        return redirect('/gestion_inventario_admin')

#Consulta editar circuito
def editar_circuito(informacion_a_actualizar):  # Query UPDATE
            query = ('''
                UPDATE Circuito
                SET Circuito.nombre = %s,
                    Circuito.cantidad = %s,
                    Circuito.descripcion = %s,
                    Circuito.dias_max_prestamo= %s,
                    Circuito.dias_renovacion= %s,
                    Circuito.imagen= %s
                WHERE Circuito.id = %s
            ''')
            #cursor.execute(query,(
            #    informacion_a_actualizar['nombre_circuito'],
            #    informacion_a_actualizar['cantidad_circuito'],
            #    informacion_a_actualizar['descripcion_circuito'],
            #    informacion_a_actualizar['dias_max_prestamo'],
            #    informacion_a_actualizar['dias_renovacion'],
            #    informacion_a_actualizar['imagen_circuito'],
            #    informacion_a_actualizar['id_circuito']
            #    ))

            db.query(query,(
                informacion_a_actualizar['nombre_circuito'],
                informacion_a_actualizar['cantidad_circuito'],
                informacion_a_actualizar['descripcion_circuito'],
                informacion_a_actualizar['dias_max_prestamo'],
                informacion_a_actualizar['dias_renovacion'],
                informacion_a_actualizar['imagen_circuito'],
                informacion_a_actualizar['id_circuito']
                ))

            return informacion_a_actualizar

#Funcion editar circuito
@mod.route('/gestion_inventario_admin/actualizar_informacion_circuito', methods = ['POST'])
def funcion_editar_circuito():
    if request.method == "POST" and session["usuario"]["id_credencial"] == 3:
        informacion_a_actualizar = request.form.to_dict()
        flash("equipo-editado")
        editar_circuito(informacion_a_actualizar)
        return redirect("/gestion_inventario_admin")

#Consulta eliminar circuito
def eliminar_circuito(circuito):
    query = ('''
        DELETE Circuito
        FROM Circuito
        WHERE Circuito.id = %s
        AND Circuito.prestados = %s
    ''')
    #cursor.execute(query,(
    #circuito['id'],
    #circuito['prestados'],
    #))

    db.query(query,(
    circuito['id'],
    circuito['prestados'],
    ))

    return 'ok'

#Funcion eliminar circuito

@mod.route("/gestion_inventario_admin/delete_circuito",methods=["POST"])
def funcion_eliminar_circuito():
    if request.method == "POST" and session["usuario"]["id_credencial"] == 3:
        equipo_a_eliminar = request.form.to_dict()
        circuitos=consultar_lista_circuito()
        if (int(equipo_a_eliminar["prestados"]) > 0):
                flash("equipo-ocupado")
                return redirect("/gestion_inventario_admin")
        eliminar_circuito(equipo_a_eliminar)
        flash("equipo-eliminado")
        return redirect("/gestion_inventario_admin")






def consultar_circuito_especifico(id_circuito):
    query = ('''
        SELECT *
        FROM Circuito
        WHERE Circuito.id = %s
    '''
    )
    #cursor.execute(query,(nombre_circuito,))
    cursor = db.query(query,(id_circuito,))
    circuito = cursor.fetchone()
    return circuito

def consultar_circuito_estado(id_circuito):
    query = ('''
            SELECT Detalle_solicitud_circuito.id AS IDD,
                Detalle_solicitud_circuito.id_solicitud_circuito AS IDS,
                Detalle_solicitud_circuito.cantidad,
                Detalle_solicitud_circuito.id_circuito AS id_componente,
                Detalle_solicitud_circuito.fecha_inicio,
                Detalle_solicitud_circuito.fecha_termino,
                Detalle_solicitud_circuito.fecha_vencimiento,
                Detalle_solicitud_circuito.renovaciones,
                Solicitud_circuito.fecha_registro,
                Solicitud_circuito.rut_alumno AS rut,
                Estado_detalle_solicitud.nombre AS estado,
                CASE WHEN Solicitud_circuito.id_curso_motivo = 0 THEN 'Personal'
                    ELSE CONCAT(Curso.codigo_udp, ' ',Curso.nombre)
                    END AS curso_motivo,
                Solicitud_circuito.motivo,
                Usuario.nombres,
                Usuario.apellidos,
                Usuario.email,
                Circuito.nombre as componente,
                Circuito.cantidad - Circuito.prestados AS disponibles,
                Circuito.cantidad AS total,
                Circuito.dias_max_prestamo,
                Circuito.dias_renovacion,
                Circuito.descripcion
            FROM Detalle_solicitud_circuito
                LEFT JOIN Solicitud_circuito
                    ON Solicitud_circuito.id = Detalle_solicitud_circuito.id_solicitud_circuito
                LEFT JOIN Usuario
                    ON Usuario.rut = Solicitud_circuito.rut_alumno
                LEFT JOIN Circuito
                    ON Circuito.id = Detalle_solicitud_circuito.id_circuito
                LEFT JOIN Curso
                    ON Solicitud_circuito.id_curso_motivo = Curso.id
                LEFT JOIN Estado_detalle_solicitud
                    ON Estado_detalle_solicitud.id = Detalle_solicitud_circuito.estado
            WHERE Detalle_solicitud_circuito.estado IN (1,2,3)
            AND Circuito.id = %s
             ''')
    cursor = db.query(query,(id_circuito,))
    return cursor.fetchall()


@mod.route("/gestion_inventario_admin/detalles_circuito/<string:id_circuito>",methods=["GET"])
def detalle_info_circuito(id_circuito):
    circuito=consultar_circuito_especifico(id_circuito)
    circuito_estado=consultar_circuito_estado(id_circuito)
    return render_template("/vistas_gestion_inventario/detalles_circuito.html",
    circuito_descripcion=circuito, circuito_prestamos=circuito_estado)


@mod.route("/exportar_inventario/<int:id_exportacion>",methods=["GET"])
def exportar_inventario(id_exportacion):
        if "usuario" not in session.keys():
            return redirect("/")
        if session["usuario"]["id_credencial"] != 3: # El usuario debe ser un administrador (Credencial = 3)
            return redirect("/")

            # 1: Equipos agrupados
            # 2: Equipos separados
            # 3: Circuitos

        if id_exportacion == 1:
            query =     ('''
                        SELECT Equipo_id,
                        Código_equipo,
                        Nombre,
                        Modelo,
                        Marca,
                        Días_de_prestamo,
                        Días_para_renovar,
                        COUNT(CASE WHEN Detalle_solicitud.estado IN (1, 2, 3) THEN 1
                        ELSE NULL END) AS "Prestados",
                        Funcional,
                        Total
                    FROM (SELECT
                            Equipo.id AS Equipo_id,
                            Equipo.codigo as Código_equipo,
                            Equipo.nombre as Nombre,
                            Equipo.modelo as Modelo,
                            Equipo.marca as Marca,
                            Equipo.descripcion,
                            Equipo.dias_max_prestamo as Días_de_prestamo,
                            Equipo.dias_renovacion as Días_para_renovar,
                            Equipo.imagen,
                            COUNT(CASE WHEN Equipo_diferenciado.activo = 1 THEN 1 ELSE NULL END) AS Funcional,
                            COUNT(Equipo_diferenciado.activo) AS Total
                            FROM Equipo
                                LEFT JOIN Equipo_diferenciado ON  Equipo.codigo = Equipo_diferenciado.codigo_equipo
                            GROUP BY Equipo.id) AS inventario_general
                        LEFT JOIN Detalle_solicitud ON inventario_general.Equipo_id = Detalle_solicitud.id_equipo
                    GROUP BY inventario_general.Equipo_id
                    ORDER BY Código_equipo
                        ''')
            cursor = db.query(query,None)
            lista_detalles = cursor.fetchall()

        elif id_exportacion == 2:
            query = ("""
                SELECT Equipo_diferenciado.id AS "ID equipo",
                    Equipo_diferenciado.codigo_equipo AS "Código equipo",
                    Equipo_diferenciado.codigo_sufijo AS "Código sufijo",
                    Equipo_diferenciado.codigo_activo AS "Código activo",
                    Equipo_diferenciado.fecha_compra AS "Fecha de compra",
                    Equipo.nombre AS "Nombre",
                    Equipo.modelo "Modelo",
                    Equipo.marca "Marca",
                    CASE WHEN Equipo_diferenciado.activo = 0 THEN 'No disponible'
                        WHEN Detalle_solicitud.estado = 1 THEN 'Por retirar'
                        WHEN Detalle_solicitud.estado = 2 THEN 'En posesión'
                        WHEN Detalle_solicitud.estado = 3 THEN 'Con atraso'
                        ELSE 'Disponible' END AS "Estado"
                FROM Equipo_diferenciado
                LEFT JOIN Equipo ON Equipo.codigo = Equipo_diferenciado.codigo_equipo
                LEFT JOIN (SELECT * FROM Detalle_solicitud WHERE Detalle_solicitud.estado IN (1,2,3)) AS Detalle_solicitud
                    ON Detalle_solicitud.id_equipo = Equipo.id
                    AND Detalle_solicitud.codigo_sufijo_equipo = Equipo_diferenciado.codigo_sufijo
                ORDER BY Equipo_diferenciado.codigo_equipo
            """)
            #cursor.execute(query)
            cursor = db.query(query,None)
            lista_detalles = cursor.fetchall()

        elif id_exportacion == 3:
                query = ('''
                    SELECT Circuito.id as "ID",
                    Circuito.nombre as "Nombre",
                    Circuito.descripcion as "Descripción",
                    Circuito.cantidad as "Cantidad",
                    Circuito.prestados as "Prestados",
                    Circuito.dias_max_prestamo as "Días max de prestamo",
                    Circuito.dias_renovacion as "Días para renovar"
                    FROM Circuito
                '''
                )
                #cursor.execute(query)
                cursor = db.query(query,None)
                lista_detalles = cursor.fetchall()


        wb = Workbook() # Instancia de libro Excel
        ws = wb.active # Hoja activa para escribir
        if id_exportacion == 1:
            ws.title = "Equipos"
            nombre_archivo = "lista_de_equipos.xlsx"
        elif id_exportacion == 2:
            ws.title = "Equipos separados"
            nombre_archivo = "lista_equipos_separados.xlsx"
        else:
            ws.title = "Circuitos"
            nombre_archivo = "lista_circuitos.xlsx"


        borde_delgado = Border(
            left=Side(border_style=BORDER_THIN, color='00000000'),
            right=Side(border_style=BORDER_THIN, color='00000000'),
            top=Side(border_style=BORDER_THIN, color='00000000'),
            bottom=Side(border_style=BORDER_THIN, color='00000000'))

        lista_columnas = []
        if len(lista_detalles):
            lista_columnas = [nombre_columna for nombre_columna in lista_detalles[0].keys()]
        else:
            return redirect(redirect_url())

        for i in range(len(lista_columnas)):
            celda = ws.cell(row=2,column=i+1)
            celda.font = Font(bold=True,color="FFFFFF")
            celda.border = borde_delgado
            celda.fill = PatternFill("solid", fgColor="4D4D4D")
            celda.alignment = Alignment(horizontal="left")
            celda.value = lista_columnas[i]

        # Se agregan los registros
        index_row = 3
        index_column = 1
        for detalle in lista_detalles:
            for key in detalle:
                celda = ws.cell(row=index_row,column=index_column)
                celda.value = detalle[key]
                celda.border = borde_delgado
                celda.alignment = Alignment(horizontal="left")
                index_column += 1
            index_column = 1
            index_row += 1

        # Ajuste automático de columnas en Excel
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length

        direccion_archivo = os.path.normpath(os.path.join(BASE_DIR, "app/static/files/exportaciones/"+nombre_archivo))
        wb.save(direccion_archivo)

        return send_file(direccion_archivo,as_attachment=True)
