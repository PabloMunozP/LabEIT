from flask import Flask,Blueprint,render_template,request,redirect,url_for,flash,session,jsonify,send_file
from config import db,cursor
import os,time,bcrypt,random
import smtplib
from email import encoders
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from jinja2 import Environment
from uuid import uuid4 # Token
from datetime import datetime,timedelta
from openpyxl import Workbook

mod = Blueprint("rutas_seba",__name__)

def redirect_url(default='index'): # Redireccionamiento desde donde vino la request
    return request.args.get('next') or \
           request.referrer or \
           url_for(default)

def enviar_correo_notificacion(archivo,str_para,str_asunto,correo_usuario): # Envío de correo (notificaciones de solicitudes de préstamo)
    # Se crea el mensaje
    correo = MIMEText(archivo,"html")
    correo.set_charset("utf-8")
    correo["From"] = "labeit.udp@gmail.com"
    correo["To"] = correo_usuario
    correo["Subject"] = str_asunto

    try:
        server = smtplib.SMTP("smtp.gmail.com",587)
        server.starttls()
        server.login("labeit.udp@gmail.com","LabEIT_UDP_2020")
        str_correo = correo.as_string()
        server.sendmail("labeit.udp@gmail.com",correo_usuario,str_correo)
        server.close()
        flash("correo-exito") # Notificación de éxito al enviar el correo

    except Exception as e:
        print(e)
        flash("correo-fallido") # Notificación de fallo al enviar el correo

# Vista principal de la plataforma.
# En caso de estar autenticado, redirecciona dentro del sistema.
# En caso contrario, se mantiene en el login.
@mod.route("/",methods=["GET"])
def principal():
    if "usuario" not in session.keys():
        return render_template("/vistas_exteriores/login.html")
    else:
        return redirect('/perfil')

# Luego de ingresar los datos en el formulario del login, se reciben para la autenticación.
@mod.route("/iniciar_sesion",methods=["POST"])
def iniciar_sesion():
    datos_solicitante = request.form.to_dict() # Se obtienen los datos del formulario
    datos_solicitante["rut"] = datos_solicitante["rut"]
    datos_solicitante["password"] = datos_solicitante["password"]

    # Se obtienen los datos del colaborador (contraseña --> hash de contraseña)
    sql_query = """
        SELECT rut,nombres,apellidos,id_credencial,email,contraseña
            FROM Usuario
                WHERE rut = %s
    """
    cursor.execute(sql_query,(datos_solicitante["rut"],))
    # Se obtienen los datos asociados al rut ingresado en el formulario
    datos_usuario_registrado = cursor.fetchone()

    # Si no se obtiene un registro, entonces el rut no se encuentra registrado en el sistema
    if datos_usuario_registrado is None:
        flash("credenciales-invalidas") # Se notifica al front-end acerca del error para alertar al usuario
        return redirect(url_for("rutas_seba.principal"))

    # En caso de que la cuenta exista, se comprueban las contraseñas
    # Se codifica la password ingresada en el formulario para comparación de hash
    datos_solicitante["password"] = datos_solicitante["password"].encode(encoding="UTF-8")

    # Si las contraseñas no coinciden, entonces se devuelve al login y se notifica el error
    if not bcrypt.checkpw(datos_solicitante["password"],datos_usuario_registrado["contraseña"].encode(encoding="UTF-8")):
        flash("credenciales-invalidas") # Se notifica al front-end acerca del error para alertar al usuario
        return redirect(url_for("rutas_seba.principal"))

    # En caso de que se compruebe la validez de la contraseña, se crea la sesión
    # Adicionalmente, se redirecciona al perfil de usuario
    del datos_usuario_registrado["contraseña"]
    session["usuario"] = {} # Creación de sesión para usuario
    for atributo in datos_usuario_registrado.keys():
        session["usuario"][str(atributo)] = datos_usuario_registrado[str(atributo)]

    # Se verifica si el usuario presenta sanciones
    sql_query = """
        SELECT *
            FROM Sanciones
                WHERE rut_alumno = %s
                AND activa = 1
    """
    cursor.execute(sql_query,(session["usuario"]["rut"],))
    sancion_usuario = cursor.fetchone()

    if sancion_usuario is not None:
        session["usuario"]["sancionado"] = True
    else: # Si no se recibe nada de la consulta, no tiene sanciones
        session["usuario"]["sancionado"] = False

    return redirect('/perfil')

# Vista para recuperación de contraseña.
@mod.route("/recuperacion_password",methods=["GET"])
def recuperacion_password():
    return render_template("/vistas_exteriores/recuperacion_password.html")

@mod.route("/enviar_recuperacion_password",methods=["POST"])
def enviar_recuperacion_password():

    # Se obtienen los datos del formulario
    datos_recuperacion = request.form.to_dict()
    datos_recuperacion["identificacion_usuario"] = datos_recuperacion["identificacion_usuario"]

    # Se revisa si el RUT o correo coincide con el registro de usuarios
    sql_query = """
        SELECT rut,nombres,email
            FROM Usuario
                WHERE rut = %s
                OR email = %s
    """
    cursor.execute(sql_query,(datos_recuperacion["identificacion_usuario"],datos_recuperacion["identificacion_usuario"]))
    datos_usuario = cursor.fetchone()

    # Si el correo o el rut no se encuentran registrados, se alerta al usuario
    if datos_usuario is None:
        flash("recuperacion-invalida") # Se notifica al front-end acerca del error para alertar al usuario
        return redirect(url_for("rutas_seba.recuperacion_password"))

    # En caso de existir registro, se envía el correo de recuperación y se alerta al usuario

    # Se abre el template HTML correspondiente al restablecimiento de contraseña
    direccion_template = os.path.normpath(os.path.join(os.getcwd(), "app/templates/vistas_exteriores/recuperacion_password_mail.html"))
    archivo_html = open(direccion_template,encoding="utf-8").read()

    # Se crea el token único para restablecimiento de contraseña
    token = str(uuid4())

    # Se eliminan los registros de token asociados al rut del usuario en caso de existir
    sql_query = """
        DELETE FROM Token_recuperacion_password
            WHERE rut_usuario = %s
    """
    cursor.execute(sql_query,(datos_usuario["rut"],))

    # Se reemplazan los datos del usuario en el template a enviar vía correo
    archivo_html = archivo_html.replace("%nombre_usuario%",datos_usuario["nombres"])
    archivo_html = archivo_html.replace("%codigo_restablecimiento%",str(random.randint(0,1000)))
    archivo_html = archivo_html.replace("%token_restablecimiento%",token)

    # Se crea el mensaje
    correo = MIMEText(archivo_html,"html")
    correo.set_charset("utf-8")
    correo["From"] = "labeit.udp@gmail.com"
    correo["To"] = datos_usuario["email"]
    correo["Subject"] = "Recuperación de contraseña"

    try:
        server = smtplib.SMTP("smtp.gmail.com",587)
        server.starttls()
        server.login("labeit.udp@gmail.com","LabEIT_UDP_2020")
        str_correo = correo.as_string()
        server.sendmail("labeit.udp@gmail.com",datos_usuario["email"],str_correo)
        server.close()
        # Se registra el token en la base de datos según el RUT del usuario
        sql_query = """
            INSERT INTO Token_recuperacion_password
                (token,rut_usuario)
                    VALUES (%s,%s)
        """
        cursor.execute(sql_query,(str(token),datos_usuario["rut"]))
        flash("correo-recuperacion-exito") # Notificación de éxito al enviar el correo

    except Exception as e:
        print(e)
        flash("correo-recuperacion-fallido") # Notificación de fallo al enviar el correo

    return redirect(url_for("rutas_seba.recuperacion_password"))

# Se redirecciona al formulario con el token respectivo al recuperar contraseña
@mod.route("/restablecer_password/<string:token>",methods=["GET"])
def restablecer_password(token):

    # Se verifica si el token se encuentra registrado
    # En caso de que esté registrado, se permite el acceso al formulario
    # En caso contrario, se redirecciona a la sección de recuperar contraseña

    sql_query = """
        SELECT token_id,rut_usuario
            FROM Token_recuperacion_password
                WHERE token = %s
    """
    cursor.execute(sql_query,(token,))
    registro_token = cursor.fetchone()

    # En caso de que el token sea inválido, se redirecciona a la sección de recuperación de contraseña
    if registro_token is None:
        return redirect(url_for("rutas_seba.recuperacion_password"))

    # Se obtiene el nombre del usuario
    sql_query = """
        SELECT nombres,apellidos
            FROM Usuario
                WHERE rut = %s
    """
    cursor.execute(sql_query,(registro_token["rut_usuario"],))
    registro_nombre_usuario = cursor.fetchone()

    if registro_nombre_usuario is not None:
        registro_token["nombres_usuario"] = registro_nombre_usuario["nombres"]
        registro_token["apellidos_usuario"] = registro_nombre_usuario["apellidos"]
    else:
        registro_token["nombres_usuario"] = ""
        registro_token["apellidos_usuario"] = ""

    return render_template("/vistas_exteriores/formulario_restablecer_password.html",
        registro_token = registro_token)

@mod.route("/recuperacion/modificar_password",methods=["POST"])
def modificar_password_recuperacion():
    datos_formulario = request.form.to_dict()

    # Se obtiene la identificación del usuario mediante el ID de token
    sql_query = """
        SELECT rut_usuario
            FROM Token_recuperacion_password
                WHERE token_id = %s
    """
    cursor.execute(sql_query,(int(datos_formulario["token_id"]),))
    datos_usuario = cursor.fetchone()

    # Si se ha obtenido un registro, se obtiene el rut
    if datos_usuario is not None:
        rut_usuario = datos_usuario["rut_usuario"]
    else:
        # No existe un rut de usuario asociado al token (posible error)
        # Se notifica al usuario y se redirecciona
        flash("error-id-token")
        return redirect(url_for("rutas_seba.principal"))

    # ------------------------------------ El registro de ID de token y usuario existe
    # Se comprueba que ambas contraseñas (nueva y confirmación) coincidan
    if datos_formulario["nueva_contraseña"] != datos_formulario["confirmacion_contraseña"]:
        # Si no coindicen, se notifica y se retorna al formulario desde donde vino
        flash("contraseñas-no-coinciden")
        return redirect(redirect_url())

    # Si coindicen, se elimina la confirmación
    del datos_formulario["confirmacion_contraseña"]

    # Se realiza la encriptación de la nueva contraseña para guardar en base de datos
    hashpass = bcrypt.hashpw(datos_formulario["nueva_contraseña"].encode(encoding="UTF-8"), bcrypt.gensalt())

    # Se almacena la nueva contraseña en la base de datos
    sql_query = """
        UPDATE Usuario
            SET contraseña = %s
                WHERE rut = %s
    """
    cursor.execute(sql_query,(hashpass.decode("UTF-8"),rut_usuario))

    # Se elimina el token generado de la base de datos
    sql_query = """
        DELETE FROM Token_recuperacion_password
            WHERE token_id = %s
    """
    cursor.execute(sql_query,(int(datos_formulario["token_id"]),))

    flash("contraseña-actualizada") # Se notifica el éxito al modificar la contraseña
    return redirect(url_for("rutas_seba.principal")) # Se redirecciona al login

@mod.route("/cerrar_sesion",methods=["GET"])
def cerrar_sesion():
    if "usuario" not in session.keys():
        return redirect(redirect_url())

    # Se elimina al usuario de la sesión
    del session["usuario"]
    # Se elimina el carro de pedidos en caso de existir
    if "carro_pedidos" in session.keys():
        del session["carro_pedidos"]

    # Se redirecciona al login una vez eliminada la sesión de usuario
    return redirect("/")


# ================================== GESTIÓN DE SOLICITUDES DE PRÉSTAMOS
@mod.route("/gestion_solicitudes_prestamos",methods=["GET"])
def gestion_solicitudes_prestamos():
    if "usuario" not in session.keys():
        return redirect("/")
    if session["usuario"]["id_credencial"] != 3: # El usuario debe ser un administrador (Credencial = 3)
        return redirect("/")

    # Se obtiene el listado de detalles de solicitudes por revisar
    sql_query = """
        SELECT Detalle_solicitud.*,Equipo.nombre,Equipo.marca,Equipo.modelo,Usuario.rut AS rut_alumno,Usuario.nombres AS nombres_usuario,Usuario.apellidos AS apellidos_usuario,Solicitud.fecha_registro
            FROM Detalle_solicitud,Equipo,Solicitud,Usuario
                WHERE Solicitud.id = Detalle_solicitud.id_solicitud
                AND Detalle_solicitud.id_equipo = Equipo.id
                AND Solicitud.rut_alumno = Usuario.rut
                AND Detalle_solicitud.estado = 0
                ORDER BY Solicitud.fecha_registro DESC
    """
    cursor.execute(sql_query)
    lista_solicitudes_por_revisar = cursor.fetchall()

    # Se obtiene el listado de detalles de solicitudes activas
    sql_query = """
        SELECT Detalle_solicitud.*,Equipo.nombre,Equipo.marca,Equipo.modelo,Usuario.rut AS rut_alumno,Usuario.nombres AS nombres_usuario,Usuario.apellidos AS apellidos_usuario,Solicitud.fecha_registro,Estado_detalle_solicitud.nombre AS nombre_estado
            FROM Detalle_solicitud,Equipo,Solicitud,Usuario,Estado_detalle_solicitud
                WHERE Solicitud.id = Detalle_solicitud.id_solicitud
                AND Estado_detalle_solicitud.id = Detalle_solicitud.estado
                AND Detalle_solicitud.id_equipo = Equipo.id
                AND Solicitud.rut_alumno = Usuario.rut
                AND Detalle_solicitud.estado != 0
                AND Detalle_solicitud.estado < 5
                ORDER BY Solicitud.fecha_registro DESC
    """
    cursor.execute(sql_query)
    lista_solicitudes_activas = cursor.fetchall()

    # Se obtiene la lista de solicitudes pertenecientes al historial
    sql_query = """
        SELECT Detalle_solicitud.*,Equipo.nombre,Equipo.marca,Equipo.modelo,Usuario.rut AS rut_alumno,Usuario.nombres AS nombres_usuario,Usuario.apellidos AS apellidos_usuario,Solicitud.fecha_registro,Estado_detalle_solicitud.nombre AS nombre_estado
            FROM Detalle_solicitud,Equipo,Solicitud,Usuario,Estado_detalle_solicitud
                WHERE Solicitud.id = Detalle_solicitud.id_solicitud
                AND Estado_detalle_solicitud.id = Detalle_solicitud.estado
                AND Detalle_solicitud.id_equipo = Equipo.id
                AND Solicitud.rut_alumno = Usuario.rut
                AND Detalle_solicitud.estado != 0
                AND Detalle_solicitud.estado >= 5
                ORDER BY Solicitud.fecha_registro DESC
    """
    cursor.execute(sql_query)
    lista_historial_solicitudes = cursor.fetchall()

    # Lista de equipos para formulario ágil
    sql_query = """
        SELECT Equipo.id,Equipo.nombre,Equipo.marca,Equipo.modelo,Equipo.codigo,Equipo_diferenciado.codigo_sufijo
            FROM Equipo,Equipo_diferenciado
                WHERE Equipo.codigo = Equipo_diferenciado.codigo_equipo
                AND Equipo_diferenciado.activo = 1
                AND Equipo_diferenciado.codigo_sufijo
                    NOT IN (SELECT Detalle_solicitud.codigo_sufijo_equipo
                                FROM Detalle_solicitud
                                    WHERE Detalle_solicitud.id_equipo = Equipo.id
                                        AND codigo_sufijo_equipo IS NOT NULL)
                ORDER BY Equipo.nombre,Equipo.marca,Equipo.modelo
    """
    cursor.execute(sql_query)
    lista_equipos_disponibles = cursor.fetchall()

    # Se obtiene de forma general la lista de solicitudes registradas (vista de canasta)
    sql_query = """
        SELECT Solicitud.*,CONCAT(Usuario.nombres,' ',Usuario.apellidos) AS nombre_solicitante,
            (SELECT COUNT(*) FROM Detalle_solicitud WHERE Detalle_solicitud.id_solicitud = Solicitud.id ) AS cantidad_detalles
         FROM
            Solicitud,Usuario
                WHERE Solicitud.rut_alumno = Usuario.rut
                ORDER BY Solicitud.fecha_registro DESC
    """
    cursor.execute(sql_query)
    lista_solicitud_canasta = cursor.fetchall()

    return render_template("/vistas_gestion_solicitudes_prestamos/gestion_solicitudes.html",
    lista_solicitudes_por_revisar=lista_solicitudes_por_revisar,
    lista_solicitudes_activas=lista_solicitudes_activas,
    lista_historial_solicitudes=lista_historial_solicitudes,
    lista_equipos_disponibles=lista_equipos_disponibles,
    lista_solicitud_canasta=lista_solicitud_canasta)

@mod.route("/gestion_solicitudes_prestamos/detalle_solicitud/<string:id_detalle_solicitud>",methods=["GET"])
def detalle_solicitud(id_detalle_solicitud):
    if "usuario" not in session.keys():
        return redirect("/")
    if session["usuario"]["id_credencial"] != 3: # El usuario debe ser un administrador (Credencial = 3)
        return redirect("/")

    # Se obtiene la información del detalle de solicitud
    sql_query = """
        SELECT Detalle_solicitud.*,Estado_detalle_solicitud.nombre AS nombre_estado
            FROM Detalle_solicitud,Estado_detalle_solicitud
                WHERE Detalle_solicitud.id = %s
                AND Detalle_solicitud.estado = Estado_detalle_solicitud.id
    """
    cursor.execute(sql_query,(id_detalle_solicitud,))
    datos_detalle_solicitud = cursor.fetchone()

    # Si ya no existe el detalle, se redirecciona y notifica al administrador
    if datos_detalle_solicitud is None:
        flash("solicitud-no-encontrada") # El registro fue eliminado
        return redirect("/gestion_solicitudes_prestamos")


    # Se obtiene la información general de la solicitud
    sql_query = """
        SELECT *
            FROM Solicitud
                WHERE Solicitud.id = %s
    """
    cursor.execute(sql_query,(datos_detalle_solicitud["id_solicitud"],))
    datos_encabezado_solicitud = cursor.fetchone()

    # Se obtiene la información del usuario solicitante
    sql_query = """
        SELECT nombres,apellidos,rut,email
            FROM Usuario
                WHERE rut = %s
    """
    cursor.execute(sql_query,(datos_encabezado_solicitud["rut_alumno"],))
    datos_alumno = cursor.fetchone()

    # Se obtiene la información del equipo
    sql_query = """
        SELECT *
            FROM Equipo
                WHERE id = %s
    """
    cursor.execute(sql_query,(datos_detalle_solicitud["id_equipo"],))
    datos_equipo = cursor.fetchone()

    # Se obtienen las cantidades de prestados y total del equipo
    sql_query = """
        SELECT COUNT(*) AS cantidad_prestados
            FROM Equipo_diferenciado
                WHERE codigo_equipo = %s
                AND activo = 1
                AND codigo_sufijo IN (SELECT codigo_sufijo_equipo
                                            FROM Detalle_solicitud
                                                WHERE id_equipo = %s
                                                    AND codigo_sufijo_equipo IS NOT NULL)
    """
    cursor.execute(sql_query,(datos_equipo["codigo"],datos_equipo["id"]))
    datos_equipo["cantidad_prestados"] = cursor.fetchone()["cantidad_prestados"]

    sql_query = """
        SELECT count(*) AS cantidad_total
            FROM Equipo_diferenciado
                WHERE codigo_equipo = %s
    """
    cursor.execute(sql_query,(datos_equipo["codigo"],))
    datos_equipo["cantidad_total"] = cursor.fetchone()["cantidad_total"]

    sql_query = """
        SELECT count(*) AS cantidad_funcionales
            FROM Equipo_diferenciado
                WHERE codigo_equipo = %s
                AND activo = 1
    """
    cursor.execute(sql_query,(datos_equipo["codigo"],))
    datos_equipo["cantidad_funcionales"] = cursor.fetchone()["cantidad_funcionales"]

    datos_equipo["cantidad_disponible"] = datos_equipo["cantidad_funcionales"]-datos_equipo["cantidad_prestados"]

    # Se obtiene la lista de equipos y usuarios para opción de modificar solicitud
    sql_query = """
        SELECT rut,nombres,apellidos
            FROM Usuario
                ORDER BY apellidos
    """
    cursor.execute(sql_query)
    lista_usuarios = cursor.fetchall()

    sql_query = """
        SELECT id,marca,modelo
            FROM Equipo
                GROUP BY codigo
    """
    cursor.execute(sql_query)
    lista_equipos = cursor.fetchall()

    # Se obtiene la lista de equipos disponibles para prestar
    sql_query = """
        SELECT Equipo_diferenciado.codigo_equipo,Equipo_diferenciado.codigo_sufijo
            FROM Equipo_diferenciado
                WHERE codigo_equipo = %s
                AND activo = 1
                AND codigo_sufijo NOT IN (SELECT codigo_sufijo_equipo
                                            FROM Detalle_solicitud
                                                WHERE id_equipo = %s
                                                    AND codigo_sufijo_equipo IS NOT NULL)
    """
    cursor.execute(sql_query,(datos_equipo["codigo"],datos_equipo["id"]))
    lista_equipos_prestamo = cursor.fetchall()

    return render_template("/vistas_gestion_solicitudes_prestamos/detalle_solicitud.html",
        datos_detalle_solicitud=datos_detalle_solicitud,
        datos_encabezado_solicitud=datos_encabezado_solicitud,
        datos_alumno=datos_alumno,
        datos_equipo=datos_equipo,
        lista_usuarios=lista_usuarios,
        lista_equipos=lista_equipos,
        lista_equipos_prestamo=lista_equipos_prestamo)

@mod.route("/rechazar_solicitud/<string:id_detalle>",methods=["POST"])
def rechazar_solicitud(id_detalle):
    if "usuario" not in session.keys():
        return redirect("/")
    if session["usuario"]["id_credencial"] != 3: # El usuario debe ser un administrador (Credencial = 3)
        return redirect("/")

    fecha_revision_solicitud = str(datetime.now().replace(microsecond=0))

    # Razón de rechazo de solicitud
    razon_rechazo = request.form.to_dict()["razon_rechazo"]
    razon_rechazo = razon_rechazo.replace("\n", "<br>") # Se modifica con new line de HTML para insertar en template

    # Se obtienen los datos al equipo y detalle de solicitud para notificar al usuario vía correo
    sql_query = """
        SELECT Equipo.marca,Equipo.modelo,Solicitud.fecha_registro,Solicitud.rut_alumno,Detalle_solicitud.id_solicitud
            FROM Equipo,Solicitud,Detalle_solicitud
                WHERE Solicitud.id = Detalle_solicitud.id_solicitud
                AND Detalle_solicitud.id_equipo = Equipo.id
                AND Detalle_solicitud.id = %s
    """
    cursor.execute(sql_query,(id_detalle,))
    datos_solicitud_rechazada = cursor.fetchone()

    # Si la solicitud no existe, se redirecciona a la sección de gestión de solicitudes
    # Además, se notifica al administrador

    if datos_solicitud_rechazada is None:
        flash("solicitud-no-encontrada")
        return redirect("/gestion_solicitudes_prestamos")

    # Si existe la solicitud, es marcada como rechazada (Historial)
    sql_query = """
        UPDATE Detalle_solicitud
            SET estado = 5
                WHERE id = %s
    """
    cursor.execute(sql_query,(id_detalle,))

    # Por último, se notifica al usuario sobre el rechazo de la solicitud
    # Se obtienen los datos del usuario
    sql_query = """
        SELECT nombres,email FROM Usuario
            WHERE rut = %s
    """
    cursor.execute(sql_query,(datos_solicitud_rechazada["rut_alumno"],)) # Modificar por el rut del solicitante
    datos_usuario = cursor.fetchone()

    direccion_template = os.path.normpath(os.path.join(os.getcwd(), "app/templates/vistas_gestion_solicitudes_prestamos/templates_mail/rechazo_solicitud.html"))
    archivo_html = open(direccion_template,encoding="utf-8").read()

    # Se reemplazan los datos correspondientes en el archivo html
    archivo_html = archivo_html.replace("%id_solicitud%",str(datos_solicitud_rechazada["id_solicitud"]))
    archivo_html = archivo_html.replace("%id_detalle%",id_detalle)
    archivo_html = archivo_html.replace("%nombre_usuario%",datos_usuario["nombres"])
    archivo_html = archivo_html.replace("%equipo_solicitado%",datos_solicitud_rechazada["marca"]+" "+datos_solicitud_rechazada["modelo"])
    archivo_html = archivo_html.replace("%fecha_registro%",str(datos_solicitud_rechazada["fecha_registro"]))
    archivo_html = archivo_html.replace("%fecha_revision_solicitud%",fecha_revision_solicitud)

    razon_rechazo = razon_rechazo.strip()
    if len(razon_rechazo) == 0:
        razon_rechazo = "** No se ha adjuntado un motivo de rechazo de solicitud. **"

    archivo_html = archivo_html.replace("%razon_rechazo%",razon_rechazo)

    enviar_correo_notificacion(archivo_html,datos_usuario["email"],"Rechazo de solicitud de préstamo",datos_usuario["email"])

    flash("solicitud-rechazada-correctamente")
    return redirect(redirect_url())

@mod.route("/aprobar_solicitud/<string:id_detalle>",methods=["POST"])
def aprobar_solicitud(id_detalle):
    if "usuario" not in session.keys():
        return redirect("/")
    if session["usuario"]["id_credencial"] != 3: # El usuario debe ser un administrador (Credencial = 3)
        return redirect("/")

    fecha_revision_solicitud = datetime.now()

    datos_formulario = request.form.to_dict()

    # Se obtienen los datos del equipo según el código de sufijo del formulario
    sql_query = """
        SELECT marca,modelo
            FROM Equipo
                WHERE codigo = %s
    """
    cursor.execute(sql_query,(datos_formulario["codigo_equipo"],))
    datos_equipo = cursor.fetchone()

    # Se verifica que el equipo aún se encuentre registrado
    sql_query = """
        SELECT count(*) AS cantidad_registrados
            FROM Equipo_diferenciado
                WHERE codigo_equipo = %s
                AND codigo_sufijo = %s
    """
    cursor.execute(sql_query,(datos_formulario["codigo_equipo"],datos_formulario["codigo_sufijo_prestado"]))
    cantidad_registrados = cursor.fetchone()["cantidad_registrados"]

    if not cantidad_registrados: # El equipo fue eliminado (coincidencia de tiempos)
        flash("equipo-no-existente")
        return redirect("/gestion_solicitudes_prestamos")

    # Se verifica que el equipo no haya sido prestado (coincidencia de tiempos)
    sql_query = """
        SELECT count(*) AS cantidad_prestados
            FROM Detalle_solicitud,Equipo
                WHERE Detalle_solicitud.id_equipo = Equipo.id
                AND Equipo.codigo = %s
                AND Detalle_solicitud.codigo_sufijo_equipo = %s
    """
    cursor.execute(sql_query,(datos_formulario["codigo_equipo"],datos_formulario["codigo_sufijo_prestado"]))
    equipo_disponible = not bool(cursor.fetchone()["cantidad_prestados"])

    if not equipo_disponible: # El equipo fue prestado mientras se recibía el formulario
        flash("equipo-prestado")
        return redirect(redirect_url())

    # En caso de que pueda prestarse:
    #Se agrega el código sufijo a la solicitud, la fecha de vencimiento y se modifica el estado a 'Por retirar'
    sql_query = """
        UPDATE Detalle_solicitud
            SET codigo_sufijo_equipo = %s,estado = 1,fecha_vencimiento = %s
                WHERE id = %s
    """
    cursor.execute(sql_query,(datos_formulario["codigo_sufijo_prestado"],datos_formulario["fecha_vencimiento_solicitud"],id_detalle))

    # Se notifica al usuario
    # Se obtienen los datos del usuario
    sql_query = """
        SELECT rut_alumno,fecha_registro,id FROM Solicitud
            WHERE id = %s
    """
    cursor.execute(sql_query,(datos_formulario["id_encabezado_solicitud"],))
    datos_encabezado_solicitud = cursor.fetchone()

    sql_query = """
        SELECT nombres,email FROM Usuario
            WHERE rut = %s
    """
    cursor.execute(sql_query,(datos_encabezado_solicitud["rut_alumno"],)) # Modificar por el rut del solicitante
    datos_usuario = cursor.fetchone()

    direccion_template = os.path.normpath(os.path.join(os.getcwd(), "app/templates/vistas_gestion_solicitudes_prestamos/templates_mail/aprobacion_solicitud.html"))
    archivo_html = open(direccion_template,encoding="utf-8").read()

    # Se reemplazan los datos correspondientes en el archivo html
    archivo_html = archivo_html.replace("%id_solicitud%",str(datos_encabezado_solicitud["id"]))
    archivo_html = archivo_html.replace("%id_detalle%",id_detalle)
    archivo_html = archivo_html.replace("%nombre_usuario%",datos_usuario["nombres"])
    archivo_html = archivo_html.replace("%equipo_solicitado%",datos_equipo["marca"]+" "+datos_equipo["modelo"])
    archivo_html = archivo_html.replace("%codigo_equipo%",datos_formulario["codigo_equipo"])
    archivo_html = archivo_html.replace("%codigo_sufijo%",datos_formulario["codigo_sufijo_prestado"])
    archivo_html = archivo_html.replace("%fecha_registro%",str(datos_encabezado_solicitud["fecha_registro"]))

    fecha_vencimiento = datetime.strptime(datos_formulario["fecha_vencimiento_solicitud"],"%Y-%m-%d").strftime("%d-%m-%Y")
    archivo_html = archivo_html.replace("%fecha_vencimiento_solicitud%",str(fecha_vencimiento))

    fecha_revision_solicitud = str(fecha_revision_solicitud.date())+" "+str(fecha_revision_solicitud.hour)+":"+str(fecha_revision_solicitud.minute)
    archivo_html = archivo_html.replace("%fecha_revision_solicitud%",fecha_revision_solicitud)

    enviar_correo_notificacion(archivo_html,datos_usuario["email"],"Aprobación de solicitud de préstamo",datos_usuario["email"])

    flash("solicitud-aprobada-correctamente")
    return redirect(redirect_url())

@mod.route("/eliminar_solicitud/<string:id_detalle>",methods=["POST"])
def eliminar_solicitud(id_detalle):
    # Permite eliminar el detalle de solicitud con id = id_solicitud

    # Se obtiene el ID de solicitud (id general) para determinar si se debe eliminar el encabezado
    sql_query = """
        SELECT id_solicitud
            FROM Detalle_solicitud
                WHERE id = %s
    """
    cursor.execute(sql_query,(id_detalle,))
    datos_encabezado = cursor.fetchone()

    # Se retorna en caso de que se haya eliminado el detalle de solicitud
    if datos_encabezado is None:
        return redirect(redirect_url())

    id_encabezado_solicitud = datos_encabezado["id_solicitud"]

    # Se elimina el detalle de solicitud
    sql_query = """
        DELETE FROM
            Detalle_solicitud
                WHERE id = %s
    """
    cursor.execute(sql_query,(id_detalle,))

    # Se verifica si existen más detalles asociados a la solicitud
    # Si no existen registros, entonces se elimina el encabezado

    sql_query = """
        SELECT COUNT(*) as cantidad_detalles
            FROM Detalle_solicitud
                WHERE id_solicitud = %s
    """
    cursor.execute(sql_query,(id_encabezado_solicitud,))
    cantidad_detalles = cursor.fetchone()["cantidad_detalles"]

    if cantidad_detalles == 0:
        sql_query = """
            DELETE FROM
                Solicitud
                    WHERE id = %s
        """
        cursor.execute(sql_query,(id_encabezado_solicitud,))

    flash("detalle-eliminado")
    return redirect("/gestion_solicitudes_prestamos")

@mod.route("/cancelar_solicitud/<string:id_detalle>",methods=["POST"])
def cancelar_solicitud(id_detalle):
    # Al cancelar una solicitud, se debe liberar el equipo reservado
    # ID de estado 'cancelada': 7

    datos_formulario = request.form.to_dict()

    fecha_cancelacion_solicitud = datetime.now()
    fecha_cancelacion_solicitud = str(fecha_cancelacion_solicitud.date())+" "+str(fecha_cancelacion_solicitud.hour)+":"+str(fecha_cancelacion_solicitud.minute)

    # Se libera el codigo de sufijo de equipo del detalle de solicitud y se modifica el estado
    sql_query = """
        UPDATE Detalle_solicitud
            SET estado = 7,codigo_sufijo_equipo = NULL
                WHERE id = %s
    """
    cursor.execute(sql_query,(id_detalle,))

    # Se obtienen los datos necesarios para el correo

    # Datos del usuario
    # Se obtienen los datos del usuario
    sql_query = """
        SELECT rut_alumno,fecha_registro,id FROM Solicitud
            WHERE id = %s
    """
    cursor.execute(sql_query,(datos_formulario["id_encabezado_solicitud"],))
    datos_encabezado_solicitud = cursor.fetchone()

    sql_query = """
        SELECT nombres,email FROM Usuario
            WHERE rut = %s
    """
    cursor.execute(sql_query,(datos_encabezado_solicitud["rut_alumno"],)) # Modificar por el rut del solicitante
    datos_usuario = cursor.fetchone()

    # Se obtienen los datos del equipo según el código de sufijo del formulario
    sql_query = """
        SELECT marca,modelo
            FROM Equipo
                WHERE codigo = %s
    """
    cursor.execute(sql_query,(datos_formulario["codigo_equipo"],))
    datos_equipo = cursor.fetchone()

    direccion_template = os.path.normpath(os.path.join(os.getcwd(), "app/templates/vistas_gestion_solicitudes_prestamos/templates_mail/cancelacion_solicitud.html"))
    archivo_html = open(direccion_template,encoding="utf-8").read()

    # Se reemplazan los datos correspondientes en el archivo html
    archivo_html = archivo_html.replace("%id_solicitud%",str(datos_encabezado_solicitud["id"]))
    archivo_html = archivo_html.replace("%id_detalle%",id_detalle)
    archivo_html = archivo_html.replace("%nombre_usuario%",datos_usuario["nombres"])
    archivo_html = archivo_html.replace("%equipo_solicitado%",datos_equipo["marca"]+" "+datos_equipo["modelo"])
    archivo_html = archivo_html.replace("%codigo_equipo%",datos_formulario["codigo_equipo"])
    archivo_html = archivo_html.replace("%codigo_sufijo%",datos_formulario["codigo_sufijo_equipo"])
    archivo_html = archivo_html.replace("%fecha_registro%",str(datos_encabezado_solicitud["fecha_registro"]))

    archivo_html = archivo_html.replace("%fecha_cancelacion_solicitud%",str(fecha_cancelacion_solicitud))

    razon_cancelacion = datos_formulario["razon_cancelacion"]
    razon_cancelacion = razon_cancelacion.strip()
    razon_cancelacion = razon_cancelacion.replace("\n", "<br>")

    if len(razon_cancelacion) == 0:
        razon_cancelacion = "** No se ha adjuntado una razón de cancelación de solicitud. **"

    archivo_html = archivo_html.replace("%razon_cancelacion%",razon_cancelacion)

    enviar_correo_notificacion(archivo_html,datos_usuario["email"],"Cancelación de solicitud de préstamo",datos_usuario["email"])

    flash("solicitud-cancelada")
    return redirect(redirect_url())

@mod.route("/entregar_equipo/<string:id_detalle>",methods=["POST"])
def entregar_equipo(id_detalle):
    # Se entrega el equipo al usuario solicitante, cambiando el estado de la solicitud
    # y agregando las fechas de inicio y término correspondientes al préstamo
    datos_formulario = request.form.to_dict()

    # Fecha en la que se marcó el retiro del equipo
    fecha_retiro = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fecha_inicio_prestamo = datetime.strptime(fecha_retiro,"%Y-%m-%d %H:%M:%S")

    # Se comprueba si el administrador especificó otra fecha
    # En caso de que no se haya especificado, se calcula la fecha según la cantidad de días especificada por el equipo
    if len(datos_formulario["fecha_termino_prestamo"]) == 0:
        fecha_termino_prestamo = fecha_inicio_prestamo + timedelta(days=int(datos_formulario["dias_max_prestamo"]))
        dia_habil = fecha_termino_prestamo.weekday() # Entrega día de la semana en formato 0-4: Lunes-Viernes / 5-6: Sábado-Domingo

        if dia_habil >= 5: # Si el día de la fecha de término es un fin de semana se debe recalcular
            if dia_habil == 6:
                delta = 1
            else:
                delta = 2
            # Se calcula la fecha de término como el Lunes inmediatamente siguiente al fin de semana
            fecha_termino_prestamo = fecha_termino_prestamo + timedelta(days=delta)
    else:
        fecha_termino_prestamo = datetime.strptime(datos_formulario["fecha_termino_prestamo"],"%Y-%m-%d")

    # Se agrega la hora a la fecha de término, por default (18:30 PM)
    fecha_termino_prestamo = fecha_termino_prestamo.replace(hour=18,minute=30,second=0)

    # Se actualizan los datos del detalle de la solicitud
    sql_query = """
        UPDATE Detalle_solicitud
            SET fecha_inicio = %s,fecha_termino = %s,estado = 2,fecha_vencimiento = NULL
                WHERE id = %s
    """
    cursor.execute(sql_query,(fecha_inicio_prestamo,fecha_termino_prestamo,id_detalle))

    # Se envía un nuevo correo al usuario con las indicaciones correspondientes
    sql_query = """
        SELECT Detalle_solicitud.*,Solicitud.rut_alumno,Solicitud.fecha_registro as fecha_registro_encabezado,Equipo.marca,Equipo.modelo,Equipo.codigo AS codigo_equipo
            FROM Solicitud,Detalle_solicitud,Equipo
                WHERE Detalle_solicitud.id_solicitud = Solicitud.id
                AND Detalle_solicitud.id_equipo = Equipo.id
                AND Detalle_solicitud.id = %s
    """
    cursor.execute(sql_query,(id_detalle,))
    datos_generales_solicitud = cursor.fetchone() # Info de solicitud + detalle + equipo

    # Se obtienen los datos del usuario
    sql_query = """
        SELECT nombres,email
            FROM Usuario
                WHERE rut = %s
    """
    cursor.execute(sql_query,(datos_generales_solicitud["rut_alumno"],))
    datos_usuario = cursor.fetchone()

    direccion_template = os.path.normpath(os.path.join(os.getcwd(), "app/templates/vistas_gestion_solicitudes_prestamos/templates_mail/indicaciones_retiro_equipo.html"))
    archivo_html = open(direccion_template,encoding="utf-8").read()

    # Se reemplazan los datos correspondientes en el archivo html
    archivo_html = archivo_html.replace("%id_solicitud%",str(datos_formulario["id_encabezado_solicitud"]))
    archivo_html = archivo_html.replace("%id_detalle%",id_detalle)
    archivo_html = archivo_html.replace("%nombre_usuario%",datos_usuario["nombres"])
    archivo_html = archivo_html.replace("%equipo_prestado%",datos_generales_solicitud["marca"]+" "+datos_generales_solicitud["modelo"])
    archivo_html = archivo_html.replace("%codigo_equipo%",datos_generales_solicitud["codigo_equipo"])
    archivo_html = archivo_html.replace("%codigo_sufijo%",datos_generales_solicitud["codigo_sufijo_equipo"])
    archivo_html = archivo_html.replace("%fecha_registro%",str(datos_generales_solicitud["fecha_registro_encabezado"]))
    archivo_html = archivo_html.replace("%fecha_retiro_equipo%",str(fecha_retiro))

    # Fechas de inicio y término de préstamo
    fecha_inicio_prestamo = str(datetime.strptime(str(fecha_inicio_prestamo),"%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M:%S"))
    fecha_termino_prestamo = str(datetime.strptime(str(fecha_termino_prestamo),"%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M:%S"))
    archivo_html = archivo_html.replace("%fecha_inicio_prestamo%",fecha_inicio_prestamo)
    archivo_html = archivo_html.replace("%fecha_termino_prestamo%",fecha_termino_prestamo)

    enviar_correo_notificacion(archivo_html,datos_usuario["email"],"Comprobante de retiro de equipo",datos_usuario["email"])

    flash("retiro-correcto")
    return redirect(redirect_url())

@mod.route("/devolucion_equipo/<string:id_detalle>",methods=["POST"])
def devolucion_equipo(id_detalle):

    datos_formulario = request.form.to_dict()

    if len(datos_formulario["fecha_devolucion_equipo"]) == 0:
        fecha_devolucion_equipo = datetime.now().replace(microsecond=0)
    else:
        fecha_devolucion_equipo = datos_formulario["fecha_devolucion_equipo"]

    # Se obtienen los datos de la solicitud para notificación vía correo electrónico
    sql_query = """
        SELECT Detalle_solicitud.*,Solicitud.rut_alumno,Solicitud.fecha_registro as fecha_registro_encabezado,Equipo.marca,Equipo.modelo,Equipo.codigo AS codigo_equipo
            FROM Solicitud,Detalle_solicitud,Equipo
                WHERE Detalle_solicitud.id_solicitud = Solicitud.id
                AND Detalle_solicitud.id_equipo = Equipo.id
                AND Detalle_solicitud.id = %s
    """
    cursor.execute(sql_query,(id_detalle,))
    datos_generales_solicitud = cursor.fetchone() # Info de solicitud + detalle + equipo

    # Se modifica el detalle de solicitud
    sql_query = """
        UPDATE Detalle_solicitud
            SET codigo_sufijo_equipo = NULL,estado = 4,fecha_devolucion = %s
                WHERE id = %s
    """
    cursor.execute(sql_query,(fecha_devolucion_equipo,id_detalle))

    # Se obtienen los datos del usuario
    sql_query = """
        SELECT nombres,email
            FROM Usuario
                WHERE rut = %s
    """
    cursor.execute(sql_query,(datos_generales_solicitud["rut_alumno"],))
    datos_usuario = cursor.fetchone()

    direccion_template = os.path.normpath(os.path.join(os.getcwd(), "app/templates/vistas_gestion_solicitudes_prestamos/templates_mail/comprobante_devolucion.html"))
    archivo_html = open(direccion_template,encoding="utf-8").read()

    archivo_html = archivo_html.replace("%id_solicitud%",str(datos_formulario["id_encabezado_solicitud"]))
    archivo_html = archivo_html.replace("%id_detalle%",id_detalle)
    archivo_html = archivo_html.replace("%nombre_usuario%",datos_usuario["nombres"])
    archivo_html = archivo_html.replace("%equipo_devuelto%",datos_generales_solicitud["marca"]+" "+datos_generales_solicitud["modelo"])
    archivo_html = archivo_html.replace("%codigo_equipo%",datos_generales_solicitud["codigo_equipo"])
    archivo_html = archivo_html.replace("%codigo_sufijo%",datos_generales_solicitud["codigo_sufijo_equipo"])
    archivo_html = archivo_html.replace("%fecha_devolucion_equipo%",str(fecha_devolucion_equipo))

    enviar_correo_notificacion(archivo_html,datos_usuario["email"],"Comprobante de devolución de equipo",datos_usuario["email"])
    flash("equipo-devuelto")
    return redirect(redirect_url())

@mod.route("/finalizar_solicitud/<string:id_detalle>",methods=["GET"])
def finalizar_solicitud(id_detalle):
    if "usuario" not in session.keys():
        return redirect("/")
    if session["usuario"]["id_credencial"] != 3: # El usuario debe ser un administrador (Credencial = 3)
        return redirect("/")

    # Se actualiza el estado del detalle de solicitud a Finalizada.
    sql_query = """
        UPDATE Detalle_solicitud
            SET estado = 6
                WHERE id = %s
    """
    cursor.execute(sql_query,(id_detalle,))

    flash("detalle-finalizado")
    return redirect(redirect_url())

# ======== EXPORTACIONES DE SOLICITUDES ============
@mod.route("/exportar_solicitudes/<int:id_exportacion>",methods=["GET"])
def exportar_solicitudes(id_exportacion):
    # El ID de exportación corresponde a la lista que se va a exportar
    # 1: Solicitudes entrantes
    # 2: Préstamos activos
    # 3: Historial de solicitudes

    # Se ejecuta la consulta SQL según el ID de exportación

    if id_exportacion == 2:
        sql_query = """
            SELECT Solicitud.fecha_registro,Detalle_solicitud.*,
            Usuario.rut AS rut_solicitante,Usuario.email,Usuario.nombres,Usuario.apellidos,
            Equipo.codigo,Equipo.nombre AS nombre_equipo,Equipo.modelo AS modelo_equipo,Equipo.marca AS marca_equipo,
            Estado_detalle_solicitud.nombre AS nombre_estado
                FROM Solicitud,Detalle_solicitud,Usuario,Equipo,Estado_detalle_solicitud
                    WHERE Solicitud.id = Detalle_solicitud.id_solicitud
                    AND Solicitud.rut_alumno = Usuario.rut
                    AND Detalle_solicitud.id_equipo = Equipo.id
                    AND Detalle_solicitud.estado = Estado_detalle_solicitud.id
                    AND Detalle_solicitud.estado >= 1
                    AND Detalle_solicitud.estado <= 6
        """
        cursor.execute(sql_query)
        lista_detalles = cursor.fetchall()

    wb = Workbook() # Instancia de libro Excel
    ws = wb.active # Hoja activa para escribir

    if id_exportacion == 1:
        ws.title = "Lista de solicitudes de préstamos por revisar"
    elif id_exportacion == 2:
        ws.title = "Lista de solicitudes de préstamos activas"
    else:
        ws.title = "Historial de solicitudes de préstamos"

    # Columnas de la tabla
    lista_columnas = ["ID solicitud","IDD","RUT del solicitante","Nombres","Apellidos","Nombre de equipo","Marca","Modelo","Estado de solicitud","Fecha de inicio","Fecha de término","Fecha de registro"]
    lista_keys = ["id_solicitud","id","rut_solicitante","nombres","apellidos","nombre_equipo","marca_equipo","modelo_equipo","nombre_estado","fecha_inicio","fecha_termino","fecha_registro"]

    # Información de la consulta
    ws.merge_cells("A1:B1")
    celda = ws.cell(row=1,column=1)
    celda.value = "Fecha de consulta"

    celda = ws.cell(row=1,column=3)
    celda.value = str(datetime.now().replace(microsecond=0))

    for i in range(len(lista_columnas)):
        celda = ws.cell(row=3,column=i+1)
        celda.value = lista_columnas[i]

    # Se agregan los registros
    index_row = 4
    index_column = 1
    for detalle in lista_detalles:
        for key in lista_keys:
            celda = ws.cell(row=index_row,column=index_column)
            celda.value = detalle[key]
            index_column += 1
        index_column = 1
        index_row += 1

    direccion_archivo = os.path.normpath(os.path.join(os.getcwd(), "app/static/files/exportacion_solicitudes.xlsx"))
    wb.save(direccion_archivo)

    return send_file(direccion_archivo,as_attachment=True)

# ========= ESTADÍSTICAS GENERALES ===================
@mod.route("/estadisticas_generales",methods=["GET"])
def estadisticas_generales():
    if "usuario" not in session.keys():
        return redirect("/")
    if session["usuario"]["id_credencial"] != 3: # El usuario debe ser un administrador (Credencial = 3)
        return redirect("/")

    # Se obtiene las cantidades según estados de solicitudes de préstamos
    cantidades_solicitudes_estados = {}
    # Solicitudes entrantes (por revisar)
    sql_query = """
        SELECT COUNT(*) as cantidad_solicitudes_entrantes
            FROM Detalle_solicitud
                WHERE estado = 0
    """
    cursor.execute(sql_query)
    cantidades_solicitudes_estados["solicitudes_entrantes"] = cursor.fetchone()["cantidad_solicitudes_entrantes"]

    # Retiros pendientes de solicitudes aprobadas
    sql_query = """
        SELECT COUNT(*) as cantidad_retiros_pendientes
            FROM Detalle_solicitud
                WHERE estado = 1
    """
    cursor.execute(sql_query)
    cantidades_solicitudes_estados["retiros_pendientes"] = cursor.fetchone()["cantidad_retiros_pendientes"]

    # Préstamos activos (en posesión)
    sql_query = """
        SELECT COUNT(*) as cantidad_en_posesion
            FROM Detalle_solicitud
                WHERE estado = 2
    """
    cursor.execute(sql_query)
    cantidades_solicitudes_estados["en_posesion"] = cursor.fetchone()["cantidad_en_posesion"]

    # Préstamos con atrasos
    sql_query = """
        SELECT COUNT(*) as cantidad_con_atrasos
            FROM Detalle_solicitud
                WHERE estado = 3
    """
    cursor.execute(sql_query)
    cantidades_solicitudes_estados["con_atrasos"] = cursor.fetchone()["cantidad_con_atrasos"]

    # Préstamos finalizados (Se contabilizan los que se encuentran aún en estado 'devuelto')
    sql_query = """
        SELECT COUNT(*) as cantidad_finalizados
            FROM Detalle_solicitud
                WHERE estado = 4
                OR estado = 6
    """
    cursor.execute(sql_query)
    cantidades_solicitudes_estados["finalizados"] = cursor.fetchone()["cantidad_finalizados"]

    # Solicitudes rechazadas y canceladas
    # Préstamos finalizados (Se contabilizan los que se encuentran aún en estado 'devuelto')
    sql_query = """
        SELECT COUNT(*) as cantidad_rechazadas_canceladas
            FROM Detalle_solicitud
                WHERE estado = 5
                OR estado = 7
    """
    cursor.execute(sql_query)
    cantidades_solicitudes_estados["rechazadas_canceladas"] = cursor.fetchone()["cantidad_rechazadas_canceladas"]

    return render_template("/estadisticas/estadisticas_generales.html",
        cantidades_solicitudes_estados=cantidades_solicitudes_estados)

@mod.route("/consultar_estadisticas_solicitudes",methods=["POST"])
def consultar_estadisticas_solicitudes():
    # Se obtiene la lista de equipos que se han solicitado en cualquiera de los estados
    # durante las fechas "desde" (limite_inferior) y "hasta" (limite_superior) indicadas.

    datos_formulario = request.form.to_dict()
    datos_formulario["id_consulta"] = int(datos_formulario["id_consulta"])

    if datos_formulario["id_consulta"] == 1:
        # Consulta 1: Se obtienen la cantidad de solicitudes para cada equipo según las fechas
        # del filtro
        # Se agrega la hora 23:59 a las fechas para cubrir los días límites en su totalidad
        datos_formulario["limite_inferior"] = str(datetime.strptime(datos_formulario["limite_inferior"],"%Y-%m-%d").replace(hour=0,minute=0))
        datos_formulario["limite_superior"] = str(datetime.strptime(datos_formulario["limite_superior"],"%Y-%m-%d").replace(hour=23,minute=59))

        sql_query = """
            SELECT Equipo.id,Equipo.nombre,Equipo.modelo,Equipo.marca,COUNT(*) AS cantidad_solicitudes
                FROM Equipo,Detalle_solicitud,Solicitud
                    WHERE Solicitud.id = Detalle_solicitud.id_solicitud
                    AND Detalle_solicitud.id_equipo = Equipo.id
                    AND Solicitud.fecha_registro >= %s
                    AND Solicitud.fecha_registro <= %s
                    GROUP BY Equipo.id
        """
        cursor.execute(sql_query,(datos_formulario["limite_inferior"],datos_formulario["limite_superior"]))
        resultados_consulta = cursor.fetchall()

        # Se transforma el formato para posterior visualización con librerías gráficas
        data = []
        for registro in resultados_consulta:
            equipo = registro["nombre"]+" "+registro["modelo"]+" "+registro["marca"]
            data.append([equipo,registro["cantidad_solicitudes"]])

        return jsonify(data)

    elif datos_formulario["id_consulta"] == 2:
        # Consulta 2: Se obtienen las cantidades para cada equipo solicitado según las fechas del formulario
        # en los distintos estados de solicitudes (por revisar, por retirar, etc)

        # Se agrega la hora 23:59 a las fechas para cubrir los días límites en su totalidad
        datos_formulario["limite_inferior"] = str(datetime.strptime(datos_formulario["limite_inferior"],"%Y-%m-%d").replace(hour=0,minute=0))
        datos_formulario["limite_superior"] = str(datetime.strptime(datos_formulario["limite_superior"],"%Y-%m-%d").replace(hour=23,minute=59))

        # Se obtiene la lista de estados para poder realizar las consultar por cada una de ellas
        sql_query = """
            SELECT Estado_detalle_solicitud.id,Estado_detalle_solicitud.nombre
                FROM Estado_detalle_solicitud
        """
        cursor.execute(sql_query)
        lista_estados = cursor.fetchall()

        # Se obtienen los equipos según las fechas del formulario para categorías en el gráfico
        sql_query = """
            SELECT Equipo.id,Equipo.nombre,Equipo.marca,Equipo.modelo
                FROM Equipo,Detalle_solicitud,Solicitud
                    WHERE Equipo.id = Detalle_solicitud.id_equipo
                    AND Detalle_solicitud.id_solicitud = Solicitud.id
                    AND Solicitud.fecha_registro >= %s
                    AND Solicitud.fecha_registro <= %s
                    GROUP BY Equipo.id
        """
        cursor.execute(sql_query,(datos_formulario["limite_inferior"],datos_formulario["limite_superior"]))
        lista_equipos = cursor.fetchall()

        if not len(lista_equipos):
            return jsonify([])

        # Se crea la lista con las categorías para posteriormente ponerlas en el gráfico
        lista_categorias = []
        lista_categorias.append("Equipo")
        for equipo in lista_equipos:
            label_equipo = ""+equipo["nombre"]+" "+equipo["marca"]+" "+equipo["modelo"]
            lista_categorias.append(label_equipo)
        dict_chart = {'role':'annotation'} # Necesario según documentación de Google Chart
        lista_categorias.append(dict(dict_chart))

        # Para cada equipo y para cada estado se obtiene la cantidad de detalles de solicitudes asociados
        # según la fecha de registro

        datos_grafico = {}
        for estado in lista_estados:
            datos_grafico[estado["nombre"]] = []
            datos_grafico[estado["nombre"]].append(estado["nombre"])
            for equipo in lista_equipos:
                sql_query = """
                    SELECT COUNT(*) AS cantidad_equipos
                        FROM Equipo,Detalle_solicitud,Solicitud,Estado_detalle_solicitud
                            WHERE Equipo.id = Detalle_solicitud.id_equipo
                            AND Detalle_solicitud.id_solicitud = Solicitud.id
                            AND Solicitud.fecha_registro >= %s
                            AND Solicitud.fecha_registro <= %s
                            AND Detalle_solicitud.estado = Estado_detalle_solicitud.id
                            AND Estado_detalle_solicitud.id = %s
                            AND Equipo.id = %s
                            GROUP BY Equipo.id
                """
                cursor.execute(sql_query,(datos_formulario["limite_inferior"],datos_formulario["limite_superior"],estado["id"],equipo["id"]))
                cantidad_equipos = cursor.fetchone()
                if cantidad_equipos is None:
                    cantidad_equipos = 0
                else:
                    cantidad_equipos = cantidad_equipos["cantidad_equipos"]

                datos_grafico[estado["nombre"]].append(cantidad_equipos)
            datos_grafico[estado["nombre"]].append("")

        # Se crea el objeto final con todos los datos según lo establecido en la API de Google Charts
        datos_finales = []
        datos_finales.append(lista_categorias)
        for estado in datos_grafico:
            datos_finales.append(datos_grafico[estado])

        return jsonify(datos_finales)

# Rutas solicitudes ágiles
@mod.route("/registrar_solicitud_agil",methods=["POST"])
def registrar_solicitud_agil():
    # Permite registrar solicitudes por parte del administrador (pensado para casos informales)
    # La solicitud se crea con todos los detalles y pasa automáticamente al estado de posesión
    rut_usuario = request.form.get("rut_usuario")
    fecha_termino = request.form.get("fecha_termino")
    lista_equipos_seleccionados = request.form.getlist("id_equipo_seleccionado")
    lista_codigos_sufijos_equipos_seleccionados = request.form.getlist("codigo_sufijo_equipo_seleccionado")

    # Se comprueba que el RUT sea válido
    sql_query = """
        SELECT COUNT(*) AS cantidad_usuarios
            FROM Usuario
                WHERE rut = %s
    """
    cursor.execute(sql_query,(rut_usuario,))
    usuario_existente = bool(cursor.fetchone()["cantidad_usuarios"])

    # Si el rut ingresado no coincide con ningún usuario, se notifica el error.
    if not usuario_existente:
        flash("usuario-no-existente")
        return redirect(redirect_url())

    # Se verifica que ambas listas (de equipos y códigos) coincidan en la cantidad de elementos
    if len(lista_equipos_seleccionados) != len(lista_codigos_sufijos_equipos_seleccionados):
        flash("error-listas-agiles")
        return redirect(redirect_url())

    # Se modifican las fechas para agregar la hora
    fecha_inicio = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fecha_termino_especificada = False

    if len(fecha_termino):
        fecha_termino_especificada = True
        # Si se especificó una fecha en el formulario, se considera esa.
        # En caso contrario, se registra para cada uno de los detalles según la cantidad de días
        fecha_termino = datetime.strptime(fecha_termino,"%Y-%m-%d").replace(hour=18,minute=30,second=0)

    # Se crea la solicitud de préstamo con el rut y la fecha correspondiente
    sql_query = """
        INSERT INTO Solicitud (rut_alumno,fecha_registro)
            VALUES (%s,%s)
    """
    cursor.execute(sql_query,(rut_usuario,fecha_inicio))
    id_solicitud = cursor.lastrowid # Se obtiene el id de solicitud recién creada

    # Se crean los detalles de solicitud según las listas de equipos y códigos
    for id_equipo,codigo_sufijo in zip(lista_equipos_seleccionados,lista_codigos_sufijos_equipos_seleccionados):

        if not fecha_termino_especificada:
            sql_query = """
                SELECT dias_max_prestamo
                    FROM Equipo
                        WHERE id = %s
            """
            cursor.execute(sql_query,(id_equipo,))
            cantidad_dias_prestamo = cursor.fetchone()["dias_max_prestamo"]

            if cantidad_dias_prestamo is None:
                cantidad_dias_prestamo = 5

            fecha_termino = datetime.strptime(fecha_inicio,"%Y-%m-%d %H:%M:%S") + timedelta(days=cantidad_dias_prestamo)
            dia_habil = fecha_termino.weekday() # Entrega día de la semana en formato 0-4: Lunes-Viernes / 5-6: Sábado-Domingo

            if dia_habil >= 5: # Si el día de la fecha de término es un fin de semana se debe recalcular
                if dia_habil == 6:
                    delta = 1
                else:
                    delta = 2
                # Se calcula la fecha de término como el Lunes inmediatamente siguiente al fin de semana
                fecha_termino += timedelta(days=delta)

            fecha_termino = fecha_termino.replace(hour=18,minute=30,second=0)

        sql_query = """
            INSERT INTO Detalle_solicitud
                (id_solicitud,id_equipo,fecha_inicio,fecha_termino,estado,codigo_sufijo_equipo)
                    VALUES (%s,%s,%s,%s,2,%s)
        """
        cursor.execute(sql_query,(id_solicitud,id_equipo,fecha_inicio,fecha_termino,codigo_sufijo))

    flash("solicitud-registrada")
    return redirect(redirect_url())

# Canasta de solicitud
@mod.route("/gestion_solicitudes_prestamos/canasta_solicitud/<string:id_solicitud>",methods=["GET"])
def detalle_canasta_solicitud(id_solicitud):
    if "usuario" not in session.keys():
        return redirect("/")
    if session["usuario"]["id_credencial"] != 3: # El usuario debe ser un administrador (Credencial = 3)
        return redirect("/")

    # Se obtiene la solicitud
    sql_query = """
        SELECT *
            FROM Solicitud
                WHERE id = %s
    """
    cursor.execute(sql_query,(id_solicitud,))
    solicitud = cursor.fetchone()

    # Se obtienen los datos principales del usuario
    sql_query = """
        SELECT rut,nombres,apellidos,email
            FROM Usuario
                WHERE rut = %s
    """
    cursor.execute(sql_query,(solicitud["rut_alumno"],))
    datos_usuario = cursor.fetchone()

    # Se obtienen cada uno de los detalles asociados a la solicitud
    sql_query = """
        SELECT Detalle_solicitud.*,Equipo.codigo AS codigo_equipo,Equipo.nombre AS nombre_equipo,Equipo.modelo AS modelo_equipo,Equipo.marca AS marca_equipo,
            Equipo.dias_max_prestamo AS dias_max_prestamo_equipo,Estado_detalle_solicitud.nombre AS nombre_estado
            FROM Detalle_solicitud,Equipo,Estado_detalle_solicitud
                WHERE id_solicitud = %s
                AND Detalle_solicitud.id_equipo = Equipo.id
                AND Detalle_solicitud.estado = Estado_detalle_solicitud.id
    """
    cursor.execute(sql_query,(id_solicitud,))
    lista_detalles_solicitud = cursor.fetchall()

    # Se almacena como llave el ID del detalle de solicitud y como valor la lista de equipos
    # disponibles para prestar
    equipos_disponibles = {}
    for detalle_solicitud in lista_detalles_solicitud:
        # En caso de que el detalle se encuentre aún pendiente, se obtiene la lista
        # de equipos disponibles para prestar
        if detalle_solicitud["estado"] == 0:
            sql_query = """
                SELECT Equipo_diferenciado.codigo_equipo,Equipo_diferenciado.codigo_sufijo
                    FROM Equipo_diferenciado,Equipo
                        WHERE Equipo.codigo = Equipo_diferenciado.codigo_equipo
                        AND Equipo.id = %s
                        AND activo = 1
                        AND codigo_sufijo NOT IN (SELECT codigo_sufijo_equipo
                                                    FROM Detalle_solicitud
                                                        WHERE id_equipo = %s
                                                            AND codigo_sufijo_equipo IS NOT NULL)
            """
            cursor.execute(sql_query,(detalle_solicitud["id_equipo"],detalle_solicitud["id_equipo"]))
            equipos_disponibles[detalle_solicitud["id"]] = cursor.fetchall()


    return render_template("/vistas_gestion_solicitudes_prestamos/detalle_canasta_solicitud.html",
        solicitud=solicitud,
        equipos_disponibles=equipos_disponibles,
        datos_usuario=datos_usuario,
        lista_detalles_solicitud=lista_detalles_solicitud)
