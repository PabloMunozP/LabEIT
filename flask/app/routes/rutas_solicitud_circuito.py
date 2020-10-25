from flask import Flask,Blueprint,render_template,request,redirect,url_for,flash,session,jsonify,make_response,send_file
from config import db,BASE_DIR
import os, time, bcrypt
import mysql.connector
import rut_chile
import glob
from datetime import datetime, timedelta
from jinja2 import Environment
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN

mod = Blueprint("rutas_solicitud_circuito",__name__)

def redirect_url(default='index'): # Redireccionamiento desde donde vino la request
    return request.args.get('next') or \
           request.referrer or \
           url_for(default)


def consultar_lista_circuito():
    query = ('''
        SELECT *
        FROM Circuito
    ''')
    cursor = db.query(query, None)
    return cursor.fetchall()


def consultar_cursos():
    query = ('''
        SELECT id,
            codigo_udp,
            nombre
        FROM Curso
    ''')
    # cursor.execute(query)
    cursor = db.query(query,None)
    return cursor.fetchall()
    
    
@mod.route("/solicitudes_prestamos_circuitos")
def solicitudes_prestamos_circuitos():
    if 'usuario' not in session: # Si no está logeado, redirigie al login
        return redirect('/')
    return render_template('solicitudes_prestamos_circuitos/preview.html',
                           lista_circuitos = consultar_lista_circuito(),
                           lista_cursos = consultar_cursos())

@mod.route("/agregar_al_carro_circuito",methods=['POST']) # Ajax
def agregar_al_carro_circuito():
    if request.method == "POST" :
        id_circuito = request.form["id_circuito"]
        cantidad = request.form["cantidad"]
        nombre = request.form["nombre_circuito"]
        
        if id_circuito and cantidad:
            if 'carro_circuito' not in session: # si no hay carrito
                session["carro_circuito"] = {}
        
            if id_circuito not in session["carro_circuito"].keys(): # Si el circuito no esta en el carro, lo agrega
                session["carro_circuito"][id_circuito] = {} # Crea un diccionario carro = { id : {}}
                session["carro_circuito"][id_circuito]['id'] = str(id_circuito) # carro = { id : {'id': 'id'}}
                session["carro_circuito"][id_circuito]['cantidad'] = int(cantidad)  # carro = { id : {'id': str , 'cantidad' : int}}
                session["carro_circuito"][id_circuito]['nombre'] = nombre #agrega el nombre
                
            else: # Si el circuito está en el carro, le suma la cantidad
                session["carro_circuito"][id_circuito]['cantidad'] = int(cantidad)    
                
            return render_template("solicitudes_prestamos_circuitos/tablas/lista_carro_circuito.html")
        
        else:
            return jsonify({'error':'missing data!'})
        
    return jsonify({'error':'missing data!'})

@mod.route("/eliminar_carro_circuito",methods=['POST']) # Ajax
def eliminar_carro_circuito():
    if request.method == "POST":
        id_circuito = request.form["id_circuito"]
        del session["carro_circuito"][id_circuito]
        if len(session["carro_circuito"]) == 0:
            del session["carro_circuito"]
            
        return render_template("solicitudes_prestamos_circuitos/tablas/lista_carro_circuito.html")
    
    return jsonify({'error':'missing data!'})


@mod.route("/vaciar_carro_circuito",methods=['POST']) #Ajax
def vaciar_carro_circuito():
    if request.method == "POST":
        del session["carro_circuito"]
        return render_template("solicitudes_prestamos_circuitos/tablas/lista_carro_circuito.html")
    
    return jsonify({'error':'missing data!'})





def generar_solicitud_alumno(rut_alumno, motivo, id_curso_motivo, rut_profesor): # funcion para generar la solicitud de circuitos
    hora_registro = datetime.now()
    query = ('''INSERT INTO Solicitud_circuito (rut_alumno, motivo, id_curso_motivo, rut_profesor, fecha_registro)
                VALUES (%s, %s, %s, %s,  %s) ''') # Query para crear la solicitud general
    cursor = db.query(query,(rut_alumno, motivo, id_curso_motivo, rut_profesor, hora_registro))
    # cursor.execute(query,(rut_alumno, motivo, id_curso_motivo, rut_profesor, hora_registro))
    id_solicitud = cursor.lastrowid # Se obtiene el id de solicitud recién creada
    
    query = ('''INSERT INTO Detalle_solicitud_circuito (id_solicitud_circuito, id_circuito, cantidad, estado)
                VALUES (%s,%s,%s,0) ''') # Query para generar el detalle de la solicitud
    
    for ID,item in session["carro_circuito"].items():
        
        cursor = db.query(query, (id_solicitud, ID, item["cantidad"]))
        
    return  


@mod.route("/registrar_solicitud_circuito",methods=['POST']) # Ajax
def registrar_solicitud_circuito():
    if request.method == "POST" and 'usuario' in session: # Si hay una sesión
        try:
            
            datos_solicitud = request.form.to_dict()
            generar_solicitud_alumno(session["usuario"]["rut"],
                                    datos_solicitud["descripcion_motivo"],
                                    datos_solicitud["curso_id"],
                                    None) # Genera una solicitud
            del session["carro_circuito"] # Elimina el carro
            flash('registro-correcto')
        
        except:
            flash('registro-error')
            
        return redirect('/solicitudes_prestamos_circuitos')
    return jsonify({'error':'missing data!'})





# ---------------------------------------------------------------- #

def consultar_solictudes_por_id(IDS):
    query = ('''
            SELECT Detalle_solicitud_circuito.id AS IDD,
                Detalle_solicitud_circuito.id_solicitud_circuito AS IDS,
                Detalle_solicitud_circuito.cantidad,
                Detalle_solicitud_circuito.id_circuito AS id_componente,
                Detalle_solicitud_circuito.fecha_inicio,
                Detalle_solicitud_circuito.fecha_termino,
                Detalle_solicitud_circuito.fecha_vencimiento,
                Detalle_solicitud_circuito.renovaciones,
                Solicitud_circuito.fecha_registro,
                Solicitud_circuito.rut_alumno AS rut,
                Estado_detalle_solicitud.nombre AS estado,
                CASE WHEN Solicitud_circuito.id_curso_motivo = 0 THEN 'Personal'
                    ELSE CONCAT(Curso.codigo_udp, ' ',Curso.nombre)
                    END AS curso_motivo,
                Solicitud_circuito.motivo,
                Usuario.nombres,
                Usuario.apellidos,
                Usuario.email,
                Circuito.nombre as componente,
                Circuito.cantidad - Circuito.prestados AS disponibles,
                Circuito.cantidad AS total,
                Circuito.dias_max_prestamo,
                Circuito.dias_renovacion,
                Circuito.descripcion
            FROM Detalle_solicitud_circuito
                LEFT JOIN Solicitud_circuito 
                    ON Solicitud_circuito.id = Detalle_solicitud_circuito.id_solicitud_circuito
                LEFT JOIN Usuario 
                    ON Usuario.rut = Solicitud_circuito.rut_alumno
                LEFT JOIN Circuito
                    ON Circuito.id = Detalle_solicitud_circuito.id_circuito
                LEFT JOIN Curso
                    ON Solicitud_circuito.id_curso_motivo = Curso.id
                LEFT JOIN Estado_detalle_solicitud
                    ON Estado_detalle_solicitud.id = Detalle_solicitud_circuito.estado
            WHERE Solicitud_circuito.id = %s
             ''')
    cursor = db.query(query,(IDS,))
    return cursor.fetchall()

def consultar_solictudes_pendientes():
    query = ('''
            SELECT Detalle_solicitud_circuito.id AS IDD,
                Detalle_solicitud_circuito.id_solicitud_circuito AS IDS,
                Detalle_solicitud_circuito.cantidad,
                Detalle_solicitud_circuito.id_circuito AS id_componente,
                Solicitud_circuito.fecha_registro,
                Solicitud_circuito.rut_alumno AS rut,
                CASE WHEN Solicitud_circuito.id_curso_motivo = 0 THEN 'Personal'
                    ELSE CONCAT(Curso.codigo_udp, ' ',Curso.nombre)
                    END AS curso_motivo,
                Solicitud_circuito.motivo,
                Usuario.nombres,
                Usuario.apellidos,
                Usuario.email,
                Circuito.nombre as componente,
                Circuito.cantidad - Circuito.prestados AS disponibles,
                Circuito.cantidad AS total,
                Circuito.dias_max_prestamo,
                Circuito.dias_renovacion,
                Circuito.descripcion
            FROM Detalle_solicitud_circuito
                LEFT JOIN Solicitud_circuito 
                    ON Solicitud_circuito.id = Detalle_solicitud_circuito.id_solicitud_circuito
                LEFT JOIN Usuario 
                    ON Usuario.rut = Solicitud_circuito.rut_alumno
                LEFT JOIN Circuito
                    ON Circuito.id = Detalle_solicitud_circuito.id_circuito
                LEFT JOIN Curso
                    ON Solicitud_circuito.id_curso_motivo = Curso.id
            WHERE Detalle_solicitud_circuito.estado = 0
             ''')
    cursor = db.query(query, None)
    return cursor.fetchall()

def consultar_solictudes_activas():
    query = ('''
            SELECT Detalle_solicitud_circuito.id AS IDD,
                Detalle_solicitud_circuito.id_solicitud_circuito AS IDS,
                Detalle_solicitud_circuito.cantidad,
                Detalle_solicitud_circuito.id_circuito AS id_componente,
                Detalle_solicitud_circuito.fecha_inicio,
                Detalle_solicitud_circuito.fecha_termino,
                Detalle_solicitud_circuito.fecha_vencimiento,
                Detalle_solicitud_circuito.renovaciones,
                Solicitud_circuito.fecha_registro,
                Solicitud_circuito.rut_alumno AS rut,
                Estado_detalle_solicitud.nombre AS estado,
                CASE WHEN Solicitud_circuito.id_curso_motivo = 0 THEN 'Personal'
                    ELSE CONCAT(Curso.codigo_udp, ' ',Curso.nombre)
                    END AS curso_motivo,
                Solicitud_circuito.motivo,
                Usuario.nombres,
                Usuario.apellidos,
                Usuario.email,
                Circuito.nombre as componente,
                Circuito.cantidad - Circuito.prestados AS disponibles,
                Circuito.cantidad AS total,
                Circuito.dias_max_prestamo,
                Circuito.dias_renovacion,
                Circuito.descripcion
            FROM Detalle_solicitud_circuito
                LEFT JOIN Solicitud_circuito 
                    ON Solicitud_circuito.id = Detalle_solicitud_circuito.id_solicitud_circuito
                LEFT JOIN Usuario 
                    ON Usuario.rut = Solicitud_circuito.rut_alumno
                LEFT JOIN Circuito
                    ON Circuito.id = Detalle_solicitud_circuito.id_circuito
                LEFT JOIN Curso
                    ON Solicitud_circuito.id_curso_motivo = Curso.id
                LEFT JOIN Estado_detalle_solicitud
                    ON Estado_detalle_solicitud.id = Detalle_solicitud_circuito.estado
            WHERE Detalle_solicitud_circuito.estado IN (1,2,3)
             ''')
    cursor = db.query(query, None)
    return cursor.fetchall()

def consultar_solictudes_historial():
    query = ('''
            SELECT Detalle_solicitud_circuito.id AS IDD,
                Detalle_solicitud_circuito.id_solicitud_circuito AS IDS,
                Detalle_solicitud_circuito.cantidad,
                Detalle_solicitud_circuito.id_circuito AS id_componente,
                Detalle_solicitud_circuito.fecha_inicio,
                Detalle_solicitud_circuito.fecha_termino,
                Detalle_solicitud_circuito.fecha_vencimiento,
                Detalle_solicitud_circuito.fecha_devolucion,
                Detalle_solicitud_circuito.renovaciones,
                Solicitud_circuito.fecha_registro,
                Solicitud_circuito.rut_alumno AS rut,
                Estado_detalle_solicitud.nombre AS estado,
                CASE WHEN Solicitud_circuito.id_curso_motivo = 0 THEN 'Personal'
                    ELSE CONCAT(Curso.codigo_udp, ' ',Curso.nombre)
                    END AS curso_motivo,
                Solicitud_circuito.motivo,
                Usuario.nombres,
                Usuario.apellidos,
                Usuario.email,
                Circuito.nombre as componente,
                Circuito.cantidad - Circuito.prestados AS disponibles,
                Circuito.cantidad AS total,
                Circuito.dias_max_prestamo,
                Circuito.dias_renovacion,
                Circuito.descripcion
            FROM Detalle_solicitud_circuito
                LEFT JOIN Solicitud_circuito 
                    ON Solicitud_circuito.id = Detalle_solicitud_circuito.id_solicitud_circuito
                LEFT JOIN Usuario 
                    ON Usuario.rut = Solicitud_circuito.rut_alumno
                LEFT JOIN Circuito
                    ON Circuito.id = Detalle_solicitud_circuito.id_circuito
                LEFT JOIN Curso
                    ON Solicitud_circuito.id_curso_motivo = Curso.id
                LEFT JOIN Estado_detalle_solicitud
                    ON Estado_detalle_solicitud.id = Detalle_solicitud_circuito.estado
            WHERE Detalle_solicitud_circuito.estado NOT IN (0,1,2,3)
             ''')
    cursor = db.query(query, None)
    return cursor.fetchall()

def consultar_solcitudes_prestamos():
    cursor = db.query('''
                SELECT Solicitud_circuito.id,
                    Solicitud_circuito.rut_alumno,
                    Solicitud_circuito.motivo,
                    Solicitud_circuito.id_curso_motivo,
                    Solicitud_circuito.fecha_registro,
                    Usuario.email,
                    CONCAT(Usuario.nombres, ' ',Usuario.apellidos) AS nombre,
                    COUNT(Detalle_solicitud_circuito.id_solicitud_circuito) AS componentes_solicitados,
                    CASE WHEN Solicitud_circuito.id_curso_motivo = 0 THEN 'Personal'
                        ELSE CONCAT(Curso.codigo_udp, ' ',Curso.nombre)
                        END AS curso_motivo
                FROM Solicitud_circuito
                LEFT JOIN Usuario 
                    ON Usuario.rut = Solicitud_circuito.rut_alumno
                LEFT JOIN Detalle_solicitud_circuito
                    ON Detalle_solicitud_circuito.id_solicitud_circuito = Solicitud_circuito.id
                LEFT JOIN Curso
                    ON Solicitud_circuito.id_curso_motivo = Curso.id
                GROUP BY Detalle_solicitud_circuito.id_solicitud_circuito
                   ''', None)
    return cursor.fetchall()

@mod.route("/gestion_solicitudes_prestamos_circuitos")
def gestion_solicitudes_prestamos_circuitos():
    if 'usuario' not in session or session["usuario"]["id_credencial"] != 3: # Si no es administrador
        return redirect('/')
    if 'carro_circuito_admin' in session: # Si hau un carro para solicitud agil, lo borra
        del session["carro_circuito_admin"]
    if 'flash' in session:
        flash(session["flash"])
        del session["flash"]
    return render_template('vistas_gestion_solicitudes_circuitos/main.html',
                           lista_solicitudes_pendientes = consultar_solictudes_pendientes(),
                           lista_solicitudes_activas = consultar_solictudes_activas(),
                           lista_solicitudes_historial = consultar_solictudes_historial(),
                           lista_solicitudes_prestamos = consultar_solcitudes_prestamos(),
                           lista_componentes = consultar_lista_circuito(),
                           lista_cursos = consultar_cursos())

# Aprueba una solicitud detalle que está en estado de pendiente 
@mod.route("/gestion_solicitudes_prestamos_circuitos/aprobar_solcitud_detalle",methods=['POST']) # AJAX
def gestion_aprobar_solicitud_detalle():
    if request.method == "POST" and session["usuario"]["id_credencial"] == 3: # Si el metodo es POST y la sesion es nivel administrador
        try:
            IDD = request.form["id_solicitud_detalle"]
            fecha_vencimiento_solicitud = request.form["fecha_vencimiento_solicitud"]
            fecha_vencimiento_solicitud = datetime.strptime(fecha_vencimiento_solicitud, "%Y-%m-%d") # Cambio el formato de la fecha
            fecha_vencimiento_solicitud = str(fecha_vencimiento_solicitud.replace(hour=18, minute=30, second=0)) # Cambia la hora a las 18:30
            
            query = ('''SELECT Circuito.id,
                            Circuito.dias_max_prestamo,
                            Circuito.cantidad - Circuito.prestados AS disponibles,
                            Circuito.prestados,
                            Detalle_solicitud_circuito.cantidad AS solicitados
                        FROM Circuito
                        LEFT JOIN Detalle_solicitud_circuito
                            ON Detalle_solicitud_circuito.id_circuito = Circuito.id
                        WHERE Detalle_solicitud_circuito.id = %s''') # Query para consultar la informacion del componente antes de actualizar la solicitud
            cursor = db.query(query,(IDD,)) 
            circuito_data = cursor.fetchone() # Guarda la informacion del componente
            
            if circuito_data == None: # Si no se encuentra informacion del componente envia el error
                return jsonify({'error':'no existe componente'}) # Retorna que hay un error en la información
            
            elif int(circuito_data["solicitados"]) <= int(circuito_data["disponibles"]): # Comprueba que la disponibilidad del componente (si hay solicitados <= disponibles)
                cursor = db.query('UPDATE Circuito SET prestados = %s WHERE id = %s;', (int(circuito_data["solicitados"]) + int(circuito_data["prestados"]),circuito_data["id"])) # Actualiza la cantidad de componentes prestados
                cursor = db.query('UPDATE Detalle_solicitud_circuito SET estado = 1, fecha_vencimiento = %s WHERE id = %s;',(fecha_vencimiento_solicitud, IDD)) # Actualiza la solicitud detalle
                ###############
                # Enviar correo
                ###############
                session["flash"] = 'sol-aprobada-correctamente'
                return jsonify({'sucess':'succes', 'solicitud':IDD})
                
            else:
                session["flash"] = 'error-solicitud'
                return jsonify({'error':'no hay suficientes componentes!'})
        except:
            session["flash"] = 'error-BDD'
            return jsonify({'error':'BDD'})
    return jsonify({'error':'acceso denegado'})

@mod.route("/gestion_solicitudes_prestamos_circuitos/borrar_solcitud_detalle",methods=['POST']) # AJAX
def gestion_borrar_solicitud_detalle():
    if request.method == "POST" and session["usuario"]["id_credencial"] == 3:
        try:
            IDD = request.form["id_solicitud_detalle"]
            IDS = request.form["id_solicitud"]
            
            query = ('''DELETE FROM Detalle_solicitud_circuito
                        WHERE Detalle_solicitud_circuito.id = %s''') # borra la solicitud detalle circuito
            db.query(query,(IDD,))
            
            query = ('''SELECT * FROM Detalle_solicitud_circuito
                        WHERE Detalle_solicitud_circuito.id_solicitud_circuito = %s''') # consulta los elementos con esa ID solicutud
            cursor = db.query(query,(IDS,)) # consulta la solicitud
            
            if len(cursor.fetchall()) < 1: # si los resultados de esa ID solicitud son < 1 los borra
                query = ('''DELETE FROM Solicitud_circuito
                            WHERE Solicitud_circuito.id = %s''')
                cursor = db.query(query,(IDS,))
            session["flash"] = 'sol-eliminada-correctamente'
            return jsonify({'nice':'nice!'})
        except:
            session["flash"] = 'error-BDD'
            return jsonify({'error':'BDD'})
    
    return jsonify({'error':'missing data!'})
 

@mod.route("/gestion_solicitudes_prestamos_circuitos/rechazar_solicitud_detalle",methods=['POST']) # AJAX
def rechazar_solcitud_detalle():
    if request.method == 'POST' and session["usuario"]["id_credencial"]== 3:
        
        try:
            IDD = request.form["id_solicitud_detalle"]
            motivo = request.form["motivo"]
            db.query('UPDATE Detalle_solicitud_circuito SET estado = 5, fecha_rechazo = %s, fecha_vencimiento = %s WHERE id = %s;',(datetime.now(), None,IDD))
            session["flash"] = 'rechazo-correcto'
            ###############
            # Enviar correo con el motivo
            ###############
            
            return jsonify({'nice':'nice!'})
        
        except:
            session["flash"] = 'error-BDD'
            return jsonify({'error':'BDD'})
        
    return jsonify({'error':'missing data!'})

@mod.route("/gestion_solicitudes_prestamos_circuitos/entregar_solicitud_detalle",methods=['POST']) # AJAX
def entregar_solicitud_detalle():
    if request.method == 'POST' and session["usuario"]["id_credencial"] == 3:
        try:
            IDD = request.form["id_solicitud_detalle"]
            fecha_termino = request.form["fecha_devolucion_solicitud"]
            fecha_inicio = datetime.now()

            
            query = ('''SELECT Circuito.id,
                            Circuito.dias_max_prestamo
                        FROM Circuito
                        RIGHT JOIN Detalle_solicitud_circuito
                            ON Detalle_solicitud_circuito.id_circuito = Circuito.id
                        WHERE Detalle_solicitud_circuito.id = %s''')
            cursor = db.query(query,(IDD,))
            circuito_data = cursor.fetchone()
            
            if fecha_termino != '': # Si hay una fecha establecida
                fecha_termino = datetime.strptime(fecha_termino, "%Y-%m-%d") # Cambia el formato de la fecha
                fecha_termino = str(fecha_termino.replace(hour=18, minute=30, second=0)) # Cambia la hora a las 18:30
                
            else: # Si no hay una fecha establecida
                fecha_termino = datetime.now() + timedelta(days=int(circuito_data["dias_max_prestamo"])) # Establece la fecha de termino a los dias predeterminados del equipo
                fecha_termino = str(fecha_termino.replace(hour=18, minute=30, second=0)) # Cambia la hora a las 18:30
                
            query = ('''UPDATE Detalle_solicitud_circuito 
                        SET estado = 2,
                            fecha_inicio = %s,
                            fecha_termino = %s, 
                            fecha_vencimiento = %s
                        WHERE id = %s;''')
            cursor = db.query(query,(fecha_inicio, fecha_termino, None, IDD))
            #################
            # Enviar correo ?
            #################
            session["flash"] = 'sol-actualizada-entrega'
            return jsonify({'nice':'nice!'})
        
        except:
            session["flash"] = 'error-BDD'
            return jsonify({'error':'BDD'})
        
    return jsonify({'error':'missing data!'})



@mod.route("/gestion_solicitudes_prestamos_circuitos/devolver_solicitud_detalle",methods=['POST']) # AJAX
def devolver_solicitud_detalle():
    if request.method == 'POST' and session["usuario"]["id_credencial"] == 3:
        try:
            IDD = request.form["id_solicitud_detalle"]
            fecha_devolucion = request.form["fecha_devolucion"]
            
            query = ('''SELECT Circuito.id,
                            Circuito.prestados,
                            Detalle_solicitud_circuito.cantidad AS solicitados
                        FROM Circuito
                        LEFT JOIN Detalle_solicitud_circuito
                            ON Detalle_solicitud_circuito.id_circuito = Circuito.id
                        WHERE Detalle_solicitud_circuito.id = %s
                        ''') # Query para obtener datos de la solicitud y componente
            cursor = db.query(query,(IDD,))
            circuito_data = cursor.fetchone() 
            
            query = ('''UPDATE Detalle_solicitud_circuito
                            SET estado = 4,
                                fecha_devolucion = %s
                            WHERE id = %s;''') # Actualiza la solicitud detalle como devuelto
            
            if fecha_devolucion == '': # Si en el formulario no hay una fecha designada
                cursor = db.query(query,(datetime.now(), # Le da fecha y hora actual
                                    IDD))
                
            else:                      # Si en el formario hay una fecha designada
                cursor = db.query(query,(fecha_devolucion, # Usa la fecha del formulario
                                    IDD))
            cursor = db.query('UPDATE Circuito SET Circuito.prestados = %s WHERE Circuito.id = %s',(int(circuito_data["prestados"])-int(circuito_data["solicitados"]),
                                                                                                circuito_data["id"]))
            # Actualiza los prestados = prestados - solicitados (los devuelve)
            
            session["flash"] = 'sol-actualizada-entrega'
            return jsonify({'nice':'nice!'})
        
        except:
            session["flash"] = 'error-BDD'
            return jsonify({'error':'BDD'})
            
    return jsonify({'error':'missing data!'})       
    
@mod.route("/gestion_solicitudes_prestamos_circuitos/cancelar_activa_solicitud_detalle",methods=['POST']) # AJAX
def cancelar_solicitud_detalle():
    if 'usuario' not in session or session["usuario"]["id_credencial"] != 3: # Si no es administrador
        return redirect('/')
    if request.method == 'POST':
        try:
            IDD = request.form["id_solicitud_detalle"]
            motivo = request.form["motivo"]
            
            query = ('''SELECT Circuito.id,
                            Circuito.prestados,
                            Detalle_solicitud_circuito.cantidad AS solicitados
                        FROM Circuito
                        LEFT JOIN Detalle_solicitud_circuito
                            ON Detalle_solicitud_circuito.id_circuito = Circuito.id
                        WHERE Detalle_solicitud_circuito.id = %s''') # Query para obtener datos de la solicitud y componente
            cursor = db.query(query,(IDD,)) # Ejecuta la query
            circuito_data = cursor.fetchone() # Almacena el resutado
            query = ('''UPDATE Detalle_solicitud_circuito
                            SET estado = 7,
                                fecha_cancelacion = %s,
                                fecha_vencimiento = NULL
                            WHERE id = %s;''') # Actualiza la solicitud detalle como cancelada 
            
            db.query(query,(datetime.now(),IDD)) # Establece la fecha para la solicitud
            db.query('UPDATE Circuito SET Circuito.prestados = %s WHERE Circuito.id = %s',
                        (int(circuito_data["prestados"])-int(circuito_data["solicitados"]),
                            circuito_data["id"])) # Actualiza los prestados = prestados - solicitados (los devuelve)
            
            
            
            # Correo ??
            session["flash"] = 'cancelar-activa'
            return jsonify({'nice':'nice!'})
        except:
            session["flash"] = 'error-BDD'
            return jsonify({'error':'BDD'})
               
    return jsonify({'error':'missing data!'})



@mod.route("/gestion_solicitudes_prestamos_circuitos/consultar_tabla_solicitud",methods=['POST']) # AJAX
def consultar_tabla_solicitud():
    if request.method == "POST" and session["usuario"]["id_credencial"] == 3:
        IDS = request.form["id_solicitud"]
        return render_template('vistas_gestion_solicitudes_circuitos/tablas/sub_detalle.html',
                               solicitudes = consultar_solictudes_por_id(IDS))
    return jsonify({'error':'missing data!'})




@mod.route("/gestion_solicitudes_prestamos_circuitos/agregar_carro",methods=['POST']) # Ajax
def gestion_agregar_carro():
    if request.method == "POST" and session["usuario"]["id_credencial"] == 3:
        id_circuito = request.form["id_circuito"]
        disponibles = request.form["disponibles"]
        cantidad = 0
        nombre = request.form["nombre_circuito"]
        if id_circuito and nombre:
            if 'carro_circuito_admin' not in session: # si no hay carrito
                session["carro_circuito_admin"] = {}
            
            if id_circuito not in session["carro_circuito_admin"].keys(): # Si el circuito no esta en el carro, lo agrega
                session["carro_circuito_admin"][id_circuito] = {} # Crea un diccionario carro = { id : {}}
                session["carro_circuito_admin"][id_circuito]['id'] = str(id_circuito) # carro = { id : {'id': 'id'}}
                session["carro_circuito_admin"][id_circuito]['cantidad'] = int(cantidad)  # carro = { id : {'id': str , 'cantidad' : int}}
                session["carro_circuito_admin"][id_circuito]['nombre'] = nombre #agrega el nombre
                session["carro_circuito_admin"][id_circuito]['disponibles'] = disponibles # determina la disponibilidad
            else: # Si el circuito está en el carro, le suma la cantidad
                session["carro_circuito_admin"][id_circuito]['cantidad'] = int(cantidad)    
            return render_template("vistas_gestion_solicitudes_circuitos/tablas/sub_solicitud_agil.html")
        else:
            return jsonify({'error':'missing data!'})
    
    return jsonify({'error':'missing data!'})


@mod.route("/gestion_solicitudes_prestamos_circuitos/eliminar_carro",methods=['POST']) # Ajax
def gestion_eliminar_carro():
    if request.method == "POST" and session["usuario"]["id_credencial"] == 3:
        id_circuito = request.form["id_circuito"]
        del session["carro_circuito_admin"][id_circuito]
        if len(session["carro_circuito_admin"]) == 0:
            del session["carro_circuito_admin"]
        return render_template("vistas_gestion_solicitudes_circuitos/tablas/sub_solicitud_agil.html")
    return jsonify({'error':'missing data!'})


# Funcion para actualizar el carro, si es posible actualizar a js
@mod.route("/gestion_solicitudes_prestamos_circuitos/actualizar_carro",methods=['POST'])
def gestion_actualizar_carro():
    if request.method == "POST" and session["usuario"]["id_credencial"] == 3:
        id_circuito = request.form["id_circuito"]
        cantidad = request.form["cantidad"]
        if int(cantidad) > int(session["carro_circuito_admin"][id_circuito]['disponibles']):
            return jsonify({'error':'data error!'})
        else:
            session["carro_circuito_admin"][id_circuito]['cantidad'] = int(cantidad)
            return jsonify({'nice':'nice!'})
        
    

# funcion para generar la solicitud agil
def generar_solicitud_agil(rut_alumno, motivo, id_curso_motivo, rut_profesor): # funcion para generar la solicitud de circuitos
    # generar Solicitud_circuito
    hora_registro = datetime.now()
    query = ('''
            INSERT INTO Solicitud_circuito (rut_alumno, motivo, id_curso_motivo, rut_profesor, fecha_registro)
            VALUES (%s, %s, %s, %s,  %s)
            ''')
    cursor = db.query(query,(rut_alumno,
                          motivo,
                          id_curso_motivo,
                          rut_profesor,
                          hora_registro))
    id_solicitud = cursor.lastrowid # Se obtiene el id de solicitud recién creada
    
    query = (''' 
             INSERT INTO Detalle_solicitud_circuito (id_solicitud_circuito, id_circuito, cantidad, estado)
             VALUES (%s,%s,%s,0)
             ''')
    for ID,item in session["carro_circuito"].items():
        cursor = db.query(query,(id_solicitud,ID,item["cantidad"]))
    return

# Funcion para confirmar la solictud agil
@mod.route("/gestion_solicitudes_prestamos_circuitos/confirmar_solicitud_agil",methods=['POST']) # AJAX
def confirmar_solicitud_agil():

    if request.method == "POST" and session["usuario"]["id_credencial"] == 3:
        try:
            datos_solicitud = request.form.to_dict()
            fecha_registro = datetime.now()
            cursor = db.query(''' SELECT rut FROM Usuario WHERE rut = %s''',(datos_solicitud['rut_usuario'],)) # Consulta para comprobar que el usuario existe
            if len(cursor.fetchall()) > 0:
                
                
                query_insert_sol = ('''INSERT INTO Solicitud_circuito (rut_alumno, motivo, id_curso_motivo, rut_profesor, fecha_registro) VALUES (%s, %s, %s, %s,  %s)''') # Query agregar una solicitud
                query_select_circuito_data = ('''SELECT * FROM Circuito WHERE id = %s''') # Query para consultar la informacion del componente
                query_insert_detalle = ('''INSERT INTO Detalle_solicitud_circuito (id_solicitud_circuito, id_circuito, cantidad, estado, fecha_inicio, fecha_termino) VALUES (%s,%s,%s,2,%s,%s)''')
                query_update_prestados = ('UPDATE Circuito SET Circuito.prestados = %s WHERE Circuito.id = %s')
                
                cursor = db.query(query_insert_sol,(datos_solicitud['rut_usuario'],None,datos_solicitud['curso_id'],None,fecha_registro)) # Agrega la solictud a la base de dato
                id_solicitud = cursor.lastrowid # Se obtiene el id de solicitud recién creada
            
                
                for ID,item in session["carro_circuito_admin"].items():
                    if item['cantidad'] > 0: # Si la cantidad solicitada es < 0 la registra
                        
                        cursor = db.query(query_select_circuito_data,(ID,)) # Obtiene los datos del componente
                        circuito_data = cursor.fetchone() # Almacena los datos del componente
                        fecha_termino = fecha_registro+timedelta(days=int(circuito_data["dias_max_prestamo"])) # Obtiene la fecha de termino
                        cursor = db.query( query_insert_detalle, (id_solicitud,ID,item["cantidad"],fecha_registro,fecha_termino.replace(hour=18, minute=30, second = 0, microsecond = 0))) # Genera el registro de la solictud_detalle_circuito
                        
                        cursor = db.query(query_update_prestados, (int(item["cantidad"]) + int(circuito_data["prestados"]), ID))
                        ###############
                        # Enviar correo
                        ############### 
                session["flash"] = 'agil-correctamente'        
                return jsonify({'nice':'nice!'})
            else:
                return jsonify({'error':'user'})
        except:
            session["flash"] = 'error-BDD'
            return jsonify({'error':'BDD'})
    return jsonify({'error':'missing data!'})

@mod.route("/exportar_solicitudes_circuitos/<int:id_exportacion>",methods=["GET"])
def exportar_inventario(id_exportacion):
        if "usuario" not in session.keys():
            return redirect("/")
        if session["usuario"]["id_credencial"] != 3: # El usuario debe ser un administrador (Credencial = 3 'Admin')
            return redirect("/")

            # 1: Equipos agrupados
            # 2: Equipos separados
            # 3: Circuitos

        if id_exportacion == 1:
            query = ('''
            SELECT Detalle_solicitud_circuito.id_solicitud_circuito AS 'ID de solicitud',
                Detalle_solicitud_circuito.id AS 'ID de detalle',
                Usuario.rut AS 'RUT del solicitante',
                CONCAT(Usuario.nombres," ",Usuario.apellidos) AS 'Nombre del solicitante',
                Detalle_solicitud_circuito.id_circuito AS "Id circuito",
                Circuito.nombre AS 'Circuito solicitado',
                Detalle_solicitud_circuito.cantidad as "Cantidad",
                Solicitud_circuito.fecha_registro AS 'Fecha de registro',
                CASE WHEN Solicitud_circuito.id_curso_motivo = 0 THEN 'Personal'
                    ELSE CONCAT(Curso.codigo_udp, ' ',Curso.nombre)
                    END AS "Motivo"
            FROM Detalle_solicitud_circuito
                LEFT JOIN Solicitud_circuito
                    ON Solicitud_circuito.id = Detalle_solicitud_circuito.id_solicitud_circuito
                LEFT JOIN Usuario
                    ON Usuario.rut = Solicitud_circuito.rut_alumno
                LEFT JOIN Circuito
                    ON Circuito.id = Detalle_solicitud_circuito.id_circuito
                LEFT JOIN Curso
                    ON Solicitud_circuito.id_curso_motivo = Curso.id
            WHERE Detalle_solicitud_circuito.estado = 0
            ORDER BY Detalle_solicitud_circuito.id_solicitud_circuito
             ''')
            #cursor.execute(query)
            cursor = db.query(query,None)
            lista_detalles = cursor.fetchall()

        elif id_exportacion == 2:
            query = ('''
            SELECT Detalle_solicitud_circuito.id_solicitud_circuito AS 'ID de solicitud',
                Detalle_solicitud_circuito.id AS 'ID de detalle',
                Usuario.rut AS 'RUT del solicitante',
                CONCAT(Usuario.nombres," ",Usuario.apellidos) AS 'Nombre del solicitante',
                Detalle_solicitud_circuito.id_circuito AS "Id circuito",
                Circuito.nombre AS 'Circuito solicitado',
                Detalle_solicitud_circuito.cantidad as "Cantidad",
                Estado_detalle_solicitud.nombre AS "Estado",
                Solicitud_circuito.fecha_registro AS 'Fecha de registro',
                Detalle_solicitud_circuito.fecha_inicio AS'Fecha de inicio',
                Detalle_solicitud_circuito.fecha_termino AS 'Fecha de término',
                Detalle_solicitud_circuito.fecha_devolucion AS 'Fecha de devolución',
                Detalle_solicitud_circuito.fecha_vencimiento AS 'Fecha de vencimiento',
                Detalle_solicitud_circuito.renovaciones AS 'Cantidad de renovaciones',
                CASE WHEN Solicitud_circuito.id_curso_motivo = 0 THEN 'Personal'
                    ELSE CONCAT(Curso.codigo_udp, ' ',Curso.nombre)
                    END AS "Motivo"
            FROM Detalle_solicitud_circuito
                LEFT JOIN Solicitud_circuito
                    ON Solicitud_circuito.id = Detalle_solicitud_circuito.id_solicitud_circuito
                LEFT JOIN Usuario
                    ON Usuario.rut = Solicitud_circuito.rut_alumno
                LEFT JOIN Circuito
                    ON Circuito.id = Detalle_solicitud_circuito.id_circuito
                LEFT JOIN Curso
                    ON Solicitud_circuito.id_curso_motivo = Curso.id
                LEFT JOIN Estado_detalle_solicitud
                    ON Estado_detalle_solicitud.id = Detalle_solicitud_circuito.estado
            WHERE Detalle_solicitud_circuito.estado IN (1,2,3)
            ORDER BY Detalle_solicitud_circuito.id_solicitud_circuito
             ''')
            #cursor.execute(query)
            cursor = db.query(query,None)
            lista_detalles = cursor.fetchall()

        elif id_exportacion == 3:
            query = ('''
            SELECT Detalle_solicitud_circuito.id_solicitud_circuito AS 'ID de solicitud',
                Detalle_solicitud_circuito.id AS 'ID de detalle',
                Solicitud_circuito.rut_alumno AS "Rut solicitante",
                CONCAT(Usuario.nombres," ",Usuario.apellidos) AS 'Nombre del solicitante',
                Detalle_solicitud_circuito.id_circuito AS "Id circuito",
                Circuito.nombre AS 'Circuito solicitado',
                Detalle_solicitud_circuito.cantidad as "Cantidad",
                Estado_detalle_solicitud.nombre AS Estado,
                Solicitud_circuito.fecha_registro AS 'Fecha de registro',
                Detalle_solicitud_circuito.fecha_inicio AS'Fecha de inicio',
                Detalle_solicitud_circuito.fecha_termino AS 'Fecha de término',
                Detalle_solicitud_circuito.fecha_devolucion AS 'Fecha de devolución',
                Detalle_solicitud_circuito.fecha_vencimiento AS 'Fecha de vencimiento',
                Detalle_solicitud_circuito.fecha_rechazo AS 'Fecha de rechazo',
                Detalle_solicitud_circuito.fecha_cancelacion AS 'Fecha de cancelación',
                Detalle_solicitud_circuito.renovaciones AS 'Cantidad de renovaciones',
                CASE WHEN Solicitud_circuito.id_curso_motivo = 0 THEN 'Personal'
                    ELSE CONCAT(Curso.codigo_udp, ' ',Curso.nombre)
                    END AS Motivo
            FROM Detalle_solicitud_circuito
                LEFT JOIN Solicitud_circuito
                    ON Solicitud_circuito.id = Detalle_solicitud_circuito.id_solicitud_circuito
                LEFT JOIN Usuario
                    ON Usuario.rut = Solicitud_circuito.rut_alumno
                LEFT JOIN Circuito
                    ON Circuito.id = Detalle_solicitud_circuito.id_circuito
                LEFT JOIN Curso
                    ON Solicitud_circuito.id_curso_motivo = Curso.id
                LEFT JOIN Estado_detalle_solicitud
                    ON Estado_detalle_solicitud.id = Detalle_solicitud_circuito.estado
            WHERE Detalle_solicitud_circuito.estado NOT IN (0,1,2,3)
            ORDER BY Detalle_solicitud_circuito.id_solicitud_circuito
             ''')
            #cursor.execute(query)
            cursor = db.query(query,None)
            lista_detalles = cursor.fetchall()


        wb = Workbook() # Instancia de libro Excel
        ws = wb.active # Hoja activa para escribir
        if id_exportacion == 1:
            ws.title = "Sol. pendientes circuitos"
            nombre_archivo = "circuitos_solicitudes_pendientes.xlsx"
        elif id_exportacion == 2:
            ws.title = "Sol. activas circuitos"
            nombre_archivo = "circuitos_solicitudes_activas.xlsx"
        else:
            ws.title = "Historial sol. circuitos"
            nombre_archivo = "circuitos_solicitudes_historial.xlsx"


        borde_delgado = Border(
            left=Side(border_style=BORDER_THIN, color='00000000'),
            right=Side(border_style=BORDER_THIN, color='00000000'),
            top=Side(border_style=BORDER_THIN, color='00000000'),
            bottom=Side(border_style=BORDER_THIN, color='00000000'))

        lista_columnas = []
        if len(lista_detalles):
            lista_columnas = [nombre_columna for nombre_columna in lista_detalles[0].keys()]
        else:
            return redirect(redirect_url())

        for i in range(len(lista_columnas)):
            celda = ws.cell(row=2,column=i+1)
            celda.font = Font(bold=True,color="FFFFFF")
            celda.border = borde_delgado
            celda.fill = PatternFill("solid", fgColor="4D4D4D")
            celda.alignment = Alignment(horizontal="left")
            celda.value = lista_columnas[i]

        # Se agregan los registros
        index_row = 3
        index_column = 1
        for detalle in lista_detalles:
            for key in detalle:
                celda = ws.cell(row=index_row,column=index_column)
                celda.value = detalle[key]
                celda.border = borde_delgado
                celda.alignment = Alignment(horizontal="left")
                index_column += 1
            index_column = 1
            index_row += 1

        # Ajuste automático de columnas en Excel
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length

        direccion_archivo = os.path.normpath(os.path.join(BASE_DIR, "app/static/files/exportaciones/"+nombre_archivo))
        wb.save(direccion_archivo)

        return send_file(direccion_archivo,as_attachment=True)
